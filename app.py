import os
import re
import sqlite3
import hashlib
import secrets
import string
import csv
import io
import json
import time
import threading
import uuid
import requests as http_requests
from urllib.parse import urlparse
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, request, jsonify, render_template, session, g, send_from_directory, send_file
from flask_cors import CORS
from flask_compress import Compress
from werkzeug.utils import secure_filename
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    Workbook = None
    Font = PatternFill = Alignment = get_column_letter = None
    OPENPYXL_AVAILABLE = False

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', secrets.token_hex(32))
app.config['DATABASE'] = os.path.join(app.instance_path, 'sms_platform.db')
app.config['DATABASE_URL'] = os.environ.get('DATABASE_URL', '')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload
# Deployment environment label. Use APP_ENVIRONMENT=test on the test server to
# show a persistent warning ribbon so users do not mistake it for production.
app.config['APP_ENVIRONMENT'] = (os.environ.get('APP_ENVIRONMENT') or 'production').strip().lower()
app.config['APP_LABEL'] = os.environ.get('APP_LABEL') or (
    'ENTORNO DE PRUEBAS' if app.config['APP_ENVIRONMENT'] == 'test' else '')

# Cache for recently generated plaintext passwords.
# The database only stores bcrypt hashes; existing passwords cannot be recovered.
# This map keeps plaintext only for passwords generated/reset in the current
# process so they can be included in the next user export, then it is cleared.
EXPORTABLE_PASSWORDS = {}

# Enable Gzip compression
Compress(app)
app.config['COMPRESS_MIMETYPES'] = ['text/html', 'text/css', 'text/xml', 'application/json', 'application/javascript', 'text/javascript']
app.config['COMPRESS_LEVEL'] = 6
app.config['COMPRESS_MIN_SIZE'] = 500

CORS(app, supports_credentials=True, origins=[
    r"^https?://localhost(:\d+)?$",
    r"^https?://127\.0\.0\.1(:\d+)?$",
    r"^capacitor://localhost$",
    r"^http://localhost$",
    r"^https?://.*$",
])

# Ensure instance folder exists
os.makedirs(app.instance_path, exist_ok=True)

# ============================================================
# Database abstraction (SQLite + PostgreSQL)
# ============================================================

def get_db_type():
    """Return 'postgres' or 'sqlite' based on DATABASE_URL."""
    url = app.config.get('DATABASE_URL', '')
    if url and url.startswith(('postgresql://', 'postgres://')):
        return 'postgres'
    return 'sqlite'

class DBWrapper:
    """Unified database wrapper that works with both SQLite and PostgreSQL."""
    def __init__(self, conn, db_type):
        self.conn = conn
        self.db_type = db_type

    def execute(self, query, params=None):
        if params is None:
            params = ()
        if self.db_type == 'postgres':
            # Convert SQLite placeholders to PostgreSQL
            pg_query = query.replace('?', '%s')
            pg_query = pg_query.replace("datetime('now')", "NOW()")
            # PostgreSQL BOOLEAN columns: SQLite uses 1/0 integers, PG needs TRUE/FALSE
            import re
            pg_query = re.sub(r'\bis_active\s*=\s*1\b', 'is_active = TRUE', pg_query, flags=re.IGNORECASE)
            pg_query = re.sub(r'\bis_active\s*=\s*0\b', 'is_active = FALSE', pg_query, flags=re.IGNORECASE)
            # Convert params: if query touches is_active via placeholder, coerce int to bool
            if isinstance(params, (list, tuple)):
                params = list(params)
                # Detect is_active=? in the ORIGINAL query to map corresponding param
                for m in re.finditer(r'is_active\s*=\s*\?', query, flags=re.IGNORECASE):
                    # The '?' is at m.end()-1. Count '?' before it = zero-based param index.
                    qmark_pos = m.end() - 1
                    p_idx = query[:qmark_pos].count('?')
                    if p_idx < len(params):
                        v = params[p_idx]
                        if isinstance(v, int) and not isinstance(v, bool):
                            params[p_idx] = bool(v)
                params = tuple(params)
            # Always use RealDictCursor so rows behave like sqlite3.Row
            import psycopg2.extras
            cur = self.conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(pg_query, params)
            return CursorWrapper(cur, self.db_type)
        else:
            cur = self.conn.execute(query, params)
            return CursorWrapper(cur, self.db_type)

    def executescript(self, script):
        if self.db_type == 'postgres':
            # PostgreSQL doesn't have executescript, execute as single statement
            self.conn.cursor().execute(script)
        else:
            self.conn.executescript(script)

    def commit(self):
        self.conn.commit()

    def rollback(self):
        try:
            self.conn.rollback()
        except Exception:
            pass

    def close(self):
        self.conn.close()

class CursorWrapper:
    """Unified cursor wrapper that returns dict-like rows for both backends."""
    def __init__(self, cursor, db_type):
        self.cursor = cursor
        self.db_type = db_type

    @property
    def lastrowid(self):
        if self.db_type == 'postgres':
            # RETURNING id was not added; query for the last inserted OID fallback
            try:
                self.cursor.execute("SELECT LASTVAL()")
                row = self.cursor.fetchone()
                return row['lastval'] if isinstance(row, dict) else row[0]
            except Exception:
                return None
        return self.cursor.lastrowid

    def fetchone(self):
        row = self.cursor.fetchone()
        if row is None:
            return None
        # RealDictRow is already dict-like; sqlite3.Row supports both indexing and keys
        return row

    def fetchall(self):
        rows = self.cursor.fetchall()
        return rows

# Per-process PostgreSQL connection pool. Each gunicorn worker is its own
# process, so a module-level pool is safe and avoids opening a fresh TCP
# connection (DNS + handshake + backend startup) on every request.
_pg_pool = None

def _get_pg_pool():
    global _pg_pool
    if _pg_pool is None:
        import psycopg2.pool
        # sync worker handles one request at a time, so minconn=1 reuses a single
        # warm connection; maxconn=2 leaves small headroom. Each worker keeps its
        # own process-level pool, so total PG connections ≈ workers × maxconn.
        _pg_pool = psycopg2.pool.SimpleConnectionPool(
            1, 2, app.config['DATABASE_URL']
        )
    return _pg_pool

def get_db():
    if 'db' not in g:
        db_type = get_db_type()
        if db_type == 'postgres':
            pool = _get_pg_pool()
            conn = pool.getconn()
            # Discard any aborted transaction left by a previous request so a
            # pooled connection starts clean.
            try:
                if conn.closed:
                    pool.putconn(conn, close=True)
                    conn = pool.getconn()
                conn.rollback()
            except Exception:
                try:
                    pool.putconn(conn, close=True)
                except Exception:
                    pass
                conn = pool.getconn()
                conn.rollback()
            conn.autocommit = False
            g.db = DBWrapper(conn, db_type)
            g._db_from_pool = True
        else:
            conn = sqlite3.connect(app.config['DATABASE'])
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA foreign_keys=ON")
            g.db = DBWrapper(conn, db_type)
            g._db_from_pool = False
    return g.db

@app.teardown_appcontext
def close_db(exception):
    db = g.pop('db', None)
    if db is not None:
        if g.pop('_db_from_pool', False) and get_db_type() == 'postgres':
            # Return the connection to the pool instead of closing it. Roll back
            # any uncommitted work so the next borrower gets a clean session.
            try:
                db.conn.rollback()
                _get_pg_pool().putconn(db.conn)
            except Exception:
                try:
                    _get_pg_pool().putconn(db.conn, close=True)
                except Exception:
                    pass
        else:
            db.close()

def init_db():
    """Initialize database schema. Supports both SQLite and PostgreSQL."""
    db_type = get_db_type()

    if db_type == 'postgres':
        import psycopg2
        conn = psycopg2.connect(app.config['DATABASE_URL'])
        conn.autocommit = True
        cur = conn.cursor()
        # Full schema matching SQLite version
        cur.execute("""
            CREATE TABLE IF NOT EXISTS user_categories (
                id SERIAL PRIMARY KEY,
                name VARCHAR(100) UNIQUE NOT NULL,
                retention_days INTEGER NOT NULL DEFAULT 0,
                is_default BOOLEAN NOT NULL DEFAULT FALSE,
                created_at TIMESTAMP NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMP NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(255) UNIQUE NOT NULL,
                password_hash VARCHAR(255) NOT NULL,
                full_name VARCHAR(255) NOT NULL DEFAULT '',
                role VARCHAR(20) NOT NULL DEFAULT 'team_member',
                team_creator_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                category_id INTEGER REFERENCES user_categories(id) ON DELETE SET NULL,
                is_active BOOLEAN NOT NULL DEFAULT TRUE,
                daily_limit INTEGER DEFAULT 0,
                permissions TEXT DEFAULT '',
                extnumber VARCHAR(50) DEFAULT NULL,
                country VARCHAR(5) DEFAULT NULL,
                created_at TIMESTAMP NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMP NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS contact_groups (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                description TEXT DEFAULT '',
                created_at TIMESTAMP NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS contacts (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                phone VARCHAR(50) NOT NULL,
                notes TEXT DEFAULT '',
                remark TEXT DEFAULT '',
                group_id INTEGER REFERENCES contact_groups(id) ON DELETE SET NULL,
                app_name VARCHAR(255) DEFAULT '',
                amount NUMERIC(14,2) DEFAULT 0,
                discount_amount NUMERIC(14,2) DEFAULT 0,
                payment_link TEXT DEFAULT '',
                created_at TIMESTAMP NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS templates (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                content TEXT NOT NULL,
                category VARCHAR(50) DEFAULT 'general',
                created_at TIMESTAMP NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMP NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS sms_records (
                id SERIAL PRIMARY KEY,
                phone VARCHAR(50) NOT NULL,
                contact_name VARCHAR(255) DEFAULT '',
                content TEXT NOT NULL,
                status VARCHAR(20) NOT NULL DEFAULT 'pending' CHECK(status IN ('pending', 'sent', 'failed', 'scheduled')),
                msgid VARCHAR(255) DEFAULT '',
                api_code INTEGER DEFAULT 0,
                api_msg TEXT DEFAULT '',
                scheduled_at TIMESTAMP,
                sent_at TIMESTAMP,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS system_settings (
                key VARCHAR(100) PRIMARY KEY,
                value TEXT NOT NULL DEFAULT '',
                updated_at TIMESTAMP NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS sms_api_configs (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) NOT NULL UNIQUE,
                country VARCHAR(10) NOT NULL DEFAULT '',
                domain VARCHAR(500) DEFAULT '',
                spid VARCHAR(255) DEFAULT '',
                api_pwd VARCHAR(255) DEFAULT '',
                sender_name VARCHAR(255) DEFAULT '',
                is_active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMP NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS sms_config (
                id SERIAL PRIMARY KEY,
                domain VARCHAR(500) DEFAULT '',
                spid VARCHAR(255) DEFAULT '',
                api_pwd VARCHAR(255) DEFAULT '',
                sender_name VARCHAR(255) DEFAULT '',
                updated_at TIMESTAMP NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS send_logs (
                id SERIAL PRIMARY KEY,
                action VARCHAR(100) NOT NULL,
                details TEXT DEFAULT '',
                status VARCHAR(20) NOT NULL DEFAULT 'info',
                created_at TIMESTAMP NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS role_permissions (
                role VARCHAR(20) PRIMARY KEY,
                permissions TEXT NOT NULL DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS team_config (
                id SERIAL PRIMARY KEY,
                team_admin_id INTEGER NOT NULL UNIQUE REFERENCES users(id),
                daily_sms_limit INTEGER DEFAULT 100,
                api_config_id INTEGER REFERENCES sms_api_configs(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS voice_config (
                id SERIAL PRIMARY KEY,
                provider VARCHAR(50) NOT NULL DEFAULT 'simulation',
                account_sid VARCHAR(255) DEFAULT '',
                auth_token VARCHAR(255) DEFAULT '',
                from_number VARCHAR(100) DEFAULT '',
                api_domain VARCHAR(500) DEFAULT '',
                voice_appid VARCHAR(255) DEFAULT '',
                voice_accesskey VARCHAR(255) DEFAULT '',
                voice_extnumber VARCHAR(100) DEFAULT '',
                ext_pool_mx VARCHAR(500) DEFAULT '',
                ext_pool_co VARCHAR(500) DEFAULT '',
                ext_pool_pe VARCHAR(500) DEFAULT '',
                api_domain_mx VARCHAR(500) DEFAULT '',
                appid_mx VARCHAR(255) DEFAULT '',
                accesskey_mx VARCHAR(255) DEFAULT '',
                from_number_mx VARCHAR(100) DEFAULT '',
                api_domain_co VARCHAR(500) DEFAULT '',
                appid_co VARCHAR(255) DEFAULT '',
                accesskey_co VARCHAR(255) DEFAULT '',
                from_number_co VARCHAR(100) DEFAULT '',
                api_domain_pe VARCHAR(500) DEFAULT '',
                appid_pe VARCHAR(255) DEFAULT '',
                accesskey_pe VARCHAR(255) DEFAULT '',
                from_number_pe VARCHAR(100) DEFAULT '',
                voice_token VARCHAR(255) DEFAULT '',
                voice_token_expiry BIGINT DEFAULT 0,
                token_mx VARCHAR(255) DEFAULT '',
                token_mx_expiry BIGINT DEFAULT 0,
                token_co VARCHAR(255) DEFAULT '',
                token_co_expiry BIGINT DEFAULT 0,
                token_pe VARCHAR(255) DEFAULT '',
                token_pe_expiry BIGINT DEFAULT 0,
                country VARCHAR(5) DEFAULT '',
                extra TEXT DEFAULT '',
                is_active BOOLEAN DEFAULT TRUE,
                updated_at TIMESTAMP NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS voice_configs (
                id SERIAL PRIMARY KEY,
                name VARCHAR(100) NOT NULL,
                country VARCHAR(5) NOT NULL DEFAULT '',
                provider VARCHAR(50) NOT NULL DEFAULT 'infin8linx',
                api_domain VARCHAR(500) DEFAULT '',
                voice_appid VARCHAR(255) DEFAULT '',
                voice_accesskey VARCHAR(255) DEFAULT '',
                from_number VARCHAR(100) DEFAULT '',
                voice_scheme VARCHAR(10) NOT NULL DEFAULT 'https',
                voice_token VARCHAR(255) DEFAULT '',
                voice_token_expiry BIGINT DEFAULT 0,
                is_active BOOLEAN DEFAULT TRUE,
                updated_at TIMESTAMP NOT NULL DEFAULT NOW(),
                UNIQUE(country)
            );

            CREATE TABLE IF NOT EXISTS voice_records (
                id SERIAL PRIMARY KEY,
                phone VARCHAR(50) NOT NULL,
                contact_name VARCHAR(255) DEFAULT '',
                script TEXT NOT NULL,
                status VARCHAR(20) NOT NULL DEFAULT 'pending' CHECK(status IN ('pending','initiated','ringing','answered','completed','failed','no-answer','busy','canceled')),
                call_sid VARCHAR(255) DEFAULT '',
                provider VARCHAR(50) DEFAULT '',
                extnumber VARCHAR(100) DEFAULT '',
                country VARCHAR(5) DEFAULT '',
                customuuid VARCHAR(64) DEFAULT '',
                provider_uuid VARCHAR(64) DEFAULT '',
                record_file VARCHAR(255) DEFAULT '',
                hangupcause INTEGER DEFAULT 0,
                answer_at TIMESTAMP,
                duration INTEGER DEFAULT 0,
                price NUMERIC(12,4) DEFAULT 0,
                error_msg TEXT DEFAULT '',
                initiated_at TIMESTAMP,
                finished_at TIMESTAMP,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP NOT NULL DEFAULT NOW()
            );
        """)

        # extensions catalog (per-country extension numbers)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS extensions (
                id SERIAL PRIMARY KEY,
                extnumber VARCHAR(100) NOT NULL,
                country VARCHAR(5) NOT NULL DEFAULT '',
                assigned_to INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP NOT NULL DEFAULT NOW(),
                UNIQUE(extnumber, country)
            );
        """)

        # Migrations for existing databases - add missing columns
        def pg_column_exists(table, column):
            cur.execute("SELECT 1 FROM information_schema.columns WHERE table_name=%s AND column_name=%s", (table, column))
            return cur.fetchone() is not None

        def pg_table_exists(table):
            cur.execute("SELECT 1 FROM information_schema.tables WHERE table_name=%s", (table,))
            return cur.fetchone() is not None

        # user_categories (clasificacion de empleados + retencion de contactos)
        if not pg_table_exists('user_categories'):
            cur.execute("""
                CREATE TABLE user_categories (
                    id SERIAL PRIMARY KEY,
                    name VARCHAR(100) UNIQUE NOT NULL,
                    retention_days INTEGER NOT NULL DEFAULT 0,
                    is_default BOOLEAN NOT NULL DEFAULT FALSE,
                    created_at TIMESTAMP NOT NULL DEFAULT NOW(),
                    updated_at TIMESTAMP NOT NULL DEFAULT NOW()
                )
            """)
        if not pg_column_exists('users', 'category_id'):
            cur.execute("ALTER TABLE users ADD COLUMN category_id INTEGER REFERENCES user_categories(id) ON DELETE SET NULL")
        cur.execute("SELECT id FROM user_categories WHERE is_default=TRUE LIMIT 1")
        if cur.fetchone() is None:
            cur.execute(
                "INSERT INTO user_categories (name, retention_days, is_default) VALUES (%s, %s, TRUE)",
                ('General', 7)
            )

        # Migration: retention days must be within 1..7. Clamp existing rows.
        try:
            cur.execute(
                "UPDATE user_categories SET retention_days = 7 WHERE retention_days > 7 "
                "OR retention_days < 1 OR retention_days IS NULL"
            )
        except Exception:
            db.rollback()

        # users migrations
        if not pg_column_exists('users', 'team_creator_id'):
            cur.execute("ALTER TABLE users ADD COLUMN team_creator_id INTEGER REFERENCES users(id) ON DELETE SET NULL")
        if not pg_column_exists('users', 'daily_limit'):
            cur.execute("ALTER TABLE users ADD COLUMN daily_limit INTEGER DEFAULT 0")
        if not pg_column_exists('users', 'permissions'):
            cur.execute("ALTER TABLE users ADD COLUMN permissions TEXT DEFAULT ''")
        if not pg_column_exists('users', 'last_login_ip'):
            cur.execute("ALTER TABLE users ADD COLUMN last_login_ip TEXT DEFAULT ''")
        if not pg_column_exists('users', 'last_login_at'):
            cur.execute("ALTER TABLE users ADD COLUMN last_login_at TEXT DEFAULT NULL")
        if not pg_column_exists('users', 'session_token'):
            cur.execute("ALTER TABLE users ADD COLUMN session_token TEXT DEFAULT ''")
        # Update role check constraint - drop old constraint, add new one
        try:
            cur.execute("ALTER TABLE users DROP CONSTRAINT IF EXISTS users_role_check")
            cur.execute("ALTER TABLE users ADD CONSTRAINT users_role_check CHECK (role IN ('admin', 'team_admin', 'team_member'))")
            cur.execute("UPDATE users SET role='team_member' WHERE role='employee'")
        except Exception:
            pass

        # contacts migrations
        if not pg_column_exists('contacts', 'remark'):
            cur.execute("ALTER TABLE contacts ADD COLUMN remark TEXT DEFAULT ''")
        if not pg_column_exists('contacts', 'created_by'):
            cur.execute("ALTER TABLE contacts ADD COLUMN created_by INTEGER")
        if not pg_column_exists('contacts', 'app_name'):
            cur.execute("ALTER TABLE contacts ADD COLUMN app_name VARCHAR(255) DEFAULT ''")
        if not pg_column_exists('contacts', 'amount'):
            cur.execute("ALTER TABLE contacts ADD COLUMN amount NUMERIC(14,2) DEFAULT 0")
        if not pg_column_exists('contacts', 'discount_amount'):
            cur.execute("ALTER TABLE contacts ADD COLUMN discount_amount NUMERIC(14,2) DEFAULT 0")
        if not pg_column_exists('contacts', 'payment_link'):
            cur.execute("ALTER TABLE contacts ADD COLUMN payment_link TEXT DEFAULT ''")

        # sms_records migrations
        if not pg_column_exists('sms_records', 'msgid'):
            cur.execute("ALTER TABLE sms_records ADD COLUMN msgid VARCHAR(255) DEFAULT ''")
        if not pg_column_exists('sms_records', 'api_code'):
            cur.execute("ALTER TABLE sms_records ADD COLUMN api_code INTEGER DEFAULT 0")
        if not pg_column_exists('sms_records', 'api_msg'):
            cur.execute("ALTER TABLE sms_records ADD COLUMN api_msg TEXT DEFAULT ''")

        # voice_config migrations (infin8linx provider fields).
        # Each entry is (column, base SQL type). The DEFAULT is appended below
        # (BIGINT -> 0, everything else -> '') so the type must NOT already
        # contain a DEFAULT clause (that would yield "TYPE DEFAULT x DEFAULT y",
        # invalid SQL, which crashes worker startup on existing PostgreSQL DBs).
        for _col, _type in (
            ('voice_appid', 'VARCHAR(255)'),
            ('voice_accesskey', 'VARCHAR(255)'),
            ('voice_extnumber', 'VARCHAR(100)'),
            ('ext_pool_mx', 'VARCHAR(500)'),
            ('ext_pool_co', 'VARCHAR(500)'),
            ('ext_pool_pe', 'VARCHAR(500)'),
            ('voice_token', 'VARCHAR(255)'),
            ('voice_token_expiry', 'BIGINT'),
            # Per-country Infinity credentials/pools
            ('api_domain_mx', 'VARCHAR(500)'),
            ('appid_mx', 'VARCHAR(255)'),
            ('accesskey_mx', 'VARCHAR(255)'),
            ('from_number_mx', 'VARCHAR(100)'),
            ('api_domain_co', 'VARCHAR(500)'),
            ('appid_co', 'VARCHAR(255)'),
            ('accesskey_co', 'VARCHAR(255)'),
            ('from_number_co', 'VARCHAR(100)'),
            ('api_domain_pe', 'VARCHAR(500)'),
            ('appid_pe', 'VARCHAR(255)'),
            ('accesskey_pe', 'VARCHAR(255)'),
            ('from_number_pe', 'VARCHAR(100)'),
            ('token_mx', 'VARCHAR(255)'),
            ('token_mx_expiry', 'BIGINT'),
            ('token_co', 'VARCHAR(255)'),
            ('token_co_expiry', 'BIGINT'),
            ('token_pe', 'VARCHAR(255)'),
            ('token_pe_expiry', 'BIGINT'),
            ('country', 'VARCHAR(5)'),
        ):
            if not pg_column_exists('voice_config', _col):
                _default = '0' if _type.upper().startswith('BIGINT') else "''"
                cur.execute(f"ALTER TABLE voice_config ADD COLUMN {_col} {_type} DEFAULT {_default}")
        if not pg_column_exists('voice_records', 'extnumber'):
            cur.execute("ALTER TABLE voice_records ADD COLUMN extnumber VARCHAR(100) DEFAULT ''")
        if not pg_column_exists('voice_records', 'country'):
            cur.execute("ALTER TABLE voice_records ADD COLUMN country VARCHAR(5) DEFAULT ''")
        # voice_records: CDR/callback correlation and call outcome fields
        if not pg_column_exists('voice_records', 'customuuid'):
            cur.execute("ALTER TABLE voice_records ADD COLUMN customuuid VARCHAR(64) DEFAULT ''")
        if not pg_column_exists('voice_records', 'provider_uuid'):
            cur.execute("ALTER TABLE voice_records ADD COLUMN provider_uuid VARCHAR(64) DEFAULT ''")
        if not pg_column_exists('voice_records', 'record_file'):
            cur.execute("ALTER TABLE voice_records ADD COLUMN record_file VARCHAR(255) DEFAULT ''")
        if not pg_column_exists('voice_records', 'hangupcause'):
            cur.execute("ALTER TABLE voice_records ADD COLUMN hangupcause INTEGER DEFAULT 0")
        if not pg_column_exists('voice_records', 'answer_at'):
            cur.execute("ALTER TABLE voice_records ADD COLUMN answer_at TIMESTAMP")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_voice_records_customuuid ON voice_records(customuuid)")
        # voice_configs: detected protocol (https by default; test connection
        # may persist 'http' if the provider answers plaintext on the port).
        if not pg_column_exists('voice_configs', 'voice_scheme'):
            cur.execute("ALTER TABLE voice_configs ADD COLUMN voice_scheme VARCHAR(10) NOT NULL DEFAULT 'https'")

        # users: per-user fixed extension (asignacion de telefono/ext fija)
        if not pg_column_exists('users', 'extnumber'):
            cur.execute("ALTER TABLE users ADD COLUMN extnumber VARCHAR(50) DEFAULT NULL")
        if not pg_column_exists('users', 'country'):
            cur.execute("ALTER TABLE users ADD COLUMN country VARCHAR(5) DEFAULT NULL")
        if not pg_column_exists('users', 'category_id'):
            cur.execute("ALTER TABLE users ADD COLUMN category_id INTEGER REFERENCES user_categories(id) ON DELETE SET NULL")

        # Create default admin
        cur.execute("SELECT id FROM users WHERE username='admin'")
        if cur.fetchone() is None:
            pw_hash = hashlib.sha256('admin123'.encode()).hexdigest()
            cur.execute(
                "INSERT INTO users (username, password_hash, full_name, role) VALUES (%s, %s, %s, %s)",
                ('admin', pw_hash, 'Administrador', 'admin')
            )
        # Create default sms_config
        cur.execute("SELECT id FROM sms_config LIMIT 1")
        if cur.fetchone() is None:
            cur.execute("INSERT INTO sms_config (domain, spid, api_pwd, sender_name) VALUES ('', '', '', '')")
        # Create default sms_api_configs
        cur.execute("SELECT id FROM sms_api_configs LIMIT 1")
        if cur.fetchone() is None:
            # Try to migrate from sms_config
            cur.execute("SELECT domain, spid, api_pwd, sender_name FROM sms_config LIMIT 1")
            old = cur.fetchone()
            if old:
                cur.execute("INSERT INTO sms_api_configs (name, country, domain, spid, api_pwd, sender_name) VALUES ('Mexico', 'MX', %s, %s, %s, %s)",
                            (old[0] or '', old[1] or '', old[2] or '', old[3] or ''))
            else:
                cur.execute("INSERT INTO sms_api_configs (name, country, domain, spid, api_pwd, sender_name) VALUES ('Mexico', 'MX', '', '', '', '')")
            cur.execute("INSERT INTO sms_api_configs (name, country, domain, spid, api_pwd, sender_name) VALUES ('Colombia', 'CO', '', '', '', '')")
        # Create default role_permissions. Seed team_admin/team_member with
        # sensible defaults so fresh installs do not show a blank sidebar;
        # admin permissions are hard-coded in /api/auth/me.
        cur.execute("SELECT role FROM role_permissions LIMIT 1")
        if cur.fetchone() is None:
            cur.execute("INSERT INTO role_permissions (role, permissions) VALUES (%s, %s)",
                        ('admin', '[]'))
            cur.execute("INSERT INTO role_permissions (role, permissions) VALUES (%s, %s)",
                        ('team_admin', json.dumps(DEFAULT_ROLE_PERMISSIONS['team_admin'])))
            cur.execute("INSERT INTO role_permissions (role, permissions) VALUES (%s, %s)",
                        ('team_member', json.dumps(DEFAULT_ROLE_PERMISSIONS['team_member'])))
        # Create default voice_config
        cur.execute("SELECT id FROM voice_config LIMIT 1")
        if cur.fetchone() is None:
            cur.execute("INSERT INTO voice_config (provider, account_sid, auth_token, from_number, is_active) VALUES ('simulation', '', '', '', TRUE)")

        # Seed voice_configs (one row per country, like sms_api_configs) and
        # migrate any per-country values from the legacy single voice_config row.
        cur.execute("SELECT COUNT(*) AS cnt FROM voice_configs")
        _cnt_row = cur.fetchone()
        _cnt = _cnt_row['cnt'] if isinstance(_cnt_row, dict) else (_cnt_row[0] if _cnt_row else 0)
        if _cnt == 0:
            cur.execute(
                "SELECT api_domain_mx, appid_mx, accesskey_mx, from_number_mx, token_mx, token_mx_expiry,"
                " api_domain_co, appid_co, accesskey_co, from_number_co, token_co, token_co_expiry,"
                " api_domain_pe, appid_pe, accesskey_pe, from_number_pe, token_pe, token_pe_expiry"
                " FROM voice_config ORDER BY id LIMIT 1"
            )
            legacy = cur.fetchone()
            defaults = [
                ('Mexico', 'mx'), ('Colombia', 'co'), ('Peru', 'pe'),
            ]
            offsets = {'mx': 0, 'co': 6, 'pe': 12}
            for name, cc in defaults:
                off = offsets[cc]
                domain = (legacy[off + 0] if legacy else '') or ''
                appid = (legacy[off + 1] if legacy else '') or ''
                accesskey = (legacy[off + 2] if legacy else '') or ''
                from_num = (legacy[off + 3] if legacy else '') or ''
                token = (legacy[off + 4] if legacy else '') or ''
                token_exp = int(legacy[off + 5] or 0) if legacy else 0
                provider = 'infin8linx' if (domain and appid and accesskey) else 'simulation'
                cur.execute(
                    "INSERT INTO voice_configs (name, country, provider, api_domain, voice_appid, voice_accesskey,"
                    " from_number, voice_token, voice_token_expiry, is_active) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE)",
                    (name, cc, provider, domain, appid, accesskey, from_num, token, token_exp)
                )

        # Performance indexes (idempotent). Critical for large teams:
        # the send_logs page and SMS records/statistics pages sort/filter by
        # these columns. Without indexes every request does a full table scan
        # and sort, which degrades linearly with row count.
        index_statements = [
            # Activity logs (admin log page): ORDER BY created_at DESC
            "CREATE INDEX IF NOT EXISTS idx_send_logs_created_at ON send_logs(created_at DESC)",
            # SMS records: list filtering by owner + time range, plus scheduling
            "CREATE INDEX IF NOT EXISTS idx_sms_records_created_by ON sms_records(created_by)",
            "CREATE INDEX IF NOT EXISTS idx_sms_records_created_at ON sms_records(created_at DESC)",
            "CREATE INDEX IF NOT EXISTS idx_sms_records_status_created_at ON sms_records(status, created_at DESC)",
            "CREATE INDEX IF NOT EXISTS idx_sms_records_status_sent_at ON sms_records(status, sent_at DESC)",
            "CREATE INDEX IF NOT EXISTS idx_sms_records_scheduled ON sms_records(status, scheduled_at)",
            # Contacts/groups lookups used by team-scoped queries
            "CREATE INDEX IF NOT EXISTS idx_contacts_created_by ON contacts(created_by)",
            "CREATE INDEX IF NOT EXISTS idx_contacts_group_id ON contacts(group_id)",
            "CREATE INDEX IF NOT EXISTS idx_voice_records_created_by ON voice_records(created_by)",
            "CREATE INDEX IF NOT EXISTS idx_voice_records_created_at ON voice_records(created_at DESC)",
            "CREATE INDEX IF NOT EXISTS idx_voice_records_status_created_at ON voice_records(status, created_at DESC)",
        ]
        # Use a session-level advisory lock so that multiple Gunicorn workers
        # booting at the same time do not race to create identically named
        # indexes (CREATE INDEX IF NOT EXISTS is not concurrency-safe under
        # concurrent DDL). Only one worker proceeds; others block here then
        # find all indexes already present.
        cur.execute("SELECT pg_advisory_xact_lock(hashtext('sms_platform_init_indexes'))")
        for stmt in index_statements:
            try:
                cur.execute(stmt)
            except Exception as e:
                # A missing table/column shouldn't block startup.
                print(f"Index warning: {e}")
        conn.commit()

        cur.close()
        conn.close()
    else:
        # SQLite
        db = sqlite3.connect(app.config['DATABASE'])
        db.executescript('''
            CREATE TABLE IF NOT EXISTS user_categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                retention_days INTEGER NOT NULL DEFAULT 0,
                is_default INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                updated_at TEXT NOT NULL DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                full_name TEXT NOT NULL DEFAULT '',
                role TEXT NOT NULL DEFAULT 'team_member' CHECK(role IN ('admin', 'team_admin', 'team_member')),
                team_creator_id INTEGER,
                category_id INTEGER,
                is_active INTEGER NOT NULL DEFAULT 1,
                daily_limit INTEGER DEFAULT 0,
                session_token TEXT DEFAULT '',
                extnumber TEXT DEFAULT NULL,
                country TEXT DEFAULT NULL,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                updated_at TEXT NOT NULL DEFAULT (datetime('now')),
                FOREIGN KEY (team_creator_id) REFERENCES users(id) ON DELETE SET NULL,
                FOREIGN KEY (category_id) REFERENCES user_categories(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS contact_groups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                description TEXT DEFAULT '',
                created_at TEXT NOT NULL DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                phone TEXT NOT NULL,
                notes TEXT DEFAULT '',
                remark TEXT DEFAULT '',
                group_id INTEGER,
                app_name TEXT DEFAULT '',
                amount REAL DEFAULT 0,
                discount_amount REAL DEFAULT 0,
                payment_link TEXT DEFAULT '',
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                FOREIGN KEY (group_id) REFERENCES contact_groups(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                content TEXT NOT NULL,
                category TEXT DEFAULT 'general',
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                updated_at TEXT NOT NULL DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS sms_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                phone TEXT NOT NULL,
                contact_name TEXT DEFAULT '',
                content TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending' CHECK(status IN ('pending', 'sent', 'failed', 'scheduled')),
                msgid TEXT DEFAULT '',
                api_code INTEGER DEFAULT 0,
                api_msg TEXT DEFAULT '',
                scheduled_at TEXT,
                sent_at TEXT,
                created_by INTEGER,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                FOREIGN KEY (created_by) REFERENCES users(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS system_settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS sms_api_configs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                country TEXT NOT NULL DEFAULT '',
                domain TEXT DEFAULT '',
                spid TEXT DEFAULT '',
                api_pwd TEXT DEFAULT '',
                sender_name TEXT DEFAULT '',
                is_active INTEGER DEFAULT 1,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                updated_at TEXT NOT NULL DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS send_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                action TEXT NOT NULL,
                details TEXT DEFAULT '',
                status TEXT NOT NULL DEFAULT 'info',
                created_at TEXT NOT NULL DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS role_permissions (
                role TEXT PRIMARY KEY,
                permissions TEXT NOT NULL DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS voice_config (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                provider TEXT NOT NULL DEFAULT 'simulation',
                account_sid TEXT DEFAULT '',
                auth_token TEXT DEFAULT '',
                from_number TEXT DEFAULT '',
                api_domain TEXT DEFAULT '',
                voice_appid TEXT DEFAULT '',
                voice_accesskey TEXT DEFAULT '',
                voice_extnumber TEXT DEFAULT '',
                ext_pool_mx TEXT DEFAULT '',
                ext_pool_co TEXT DEFAULT '',
                ext_pool_pe TEXT DEFAULT '',
                api_domain_mx TEXT DEFAULT '',
                appid_mx TEXT DEFAULT '',
                accesskey_mx TEXT DEFAULT '',
                from_number_mx TEXT DEFAULT '',
                api_domain_co TEXT DEFAULT '',
                appid_co TEXT DEFAULT '',
                accesskey_co TEXT DEFAULT '',
                from_number_co TEXT DEFAULT '',
                api_domain_pe TEXT DEFAULT '',
                appid_pe TEXT DEFAULT '',
                accesskey_pe TEXT DEFAULT '',
                from_number_pe TEXT DEFAULT '',
                voice_token TEXT DEFAULT '',
                voice_token_expiry INTEGER DEFAULT 0,
                token_mx TEXT DEFAULT '',
                token_mx_expiry INTEGER DEFAULT 0,
                token_co TEXT DEFAULT '',
                token_co_expiry INTEGER DEFAULT 0,
                token_pe TEXT DEFAULT '',
                token_pe_expiry INTEGER DEFAULT 0,
                extra TEXT DEFAULT '',
                is_active INTEGER DEFAULT 1,
                updated_at TEXT NOT NULL DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS voice_configs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                country TEXT NOT NULL DEFAULT '',
                provider TEXT NOT NULL DEFAULT 'infin8linx',
                api_domain TEXT DEFAULT '',
                voice_appid TEXT DEFAULT '',
                voice_accesskey TEXT DEFAULT '',
                from_number TEXT DEFAULT '',
                voice_scheme TEXT NOT NULL DEFAULT 'https',
                voice_token TEXT DEFAULT '',
                voice_token_expiry INTEGER DEFAULT 0,
                is_active INTEGER DEFAULT 1,
                updated_at TEXT NOT NULL DEFAULT (datetime('now')),
                UNIQUE(country)
            );
            CREATE TABLE IF NOT EXISTS voice_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                phone TEXT NOT NULL,
                contact_name TEXT DEFAULT '',
                script TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending' CHECK(status IN ('pending','initiated','ringing','answered','completed','failed','no-answer','busy','canceled')),
                call_sid TEXT DEFAULT '',
                provider TEXT DEFAULT '',
                extnumber TEXT DEFAULT '',
                country TEXT DEFAULT '',
                customuuid TEXT DEFAULT '',
                provider_uuid TEXT DEFAULT '',
                record_file TEXT DEFAULT '',
                hangupcause INTEGER DEFAULT 0,
                answer_at TEXT,
                duration INTEGER DEFAULT 0,
                price REAL DEFAULT 0,
                error_msg TEXT DEFAULT '',
                initiated_at TEXT,
                finished_at TEXT,
                created_by INTEGER,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                FOREIGN KEY (created_by) REFERENCES users(id) ON DELETE SET NULL
            );
            CREATE INDEX IF NOT EXISTS idx_voice_records_created_by ON voice_records(created_by);
            CREATE INDEX IF NOT EXISTS idx_voice_records_created_at ON voice_records(created_at DESC);
            CREATE INDEX IF NOT EXISTS idx_voice_records_status_created_at ON voice_records(status, created_at DESC);

            CREATE TABLE IF NOT EXISTS extensions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                extnumber TEXT NOT NULL,
                country TEXT NOT NULL DEFAULT '',
                assigned_to INTEGER,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                UNIQUE(extnumber, country),
                FOREIGN KEY (assigned_to) REFERENCES users(id) ON DELETE SET NULL
            );
        ''')
        # Create default admin if not exists
        cursor = db.execute("SELECT id FROM users WHERE username='admin'")
        if cursor.fetchone() is None:
            pw_hash = hashlib.sha256('admin123'.encode()).hexdigest()
            db.execute(
                "INSERT INTO users (username, password_hash, full_name, role) VALUES (?, ?, ?, ?)",
                ('admin', pw_hash, 'Administrador', 'admin')
            )
        # Create default sms_api_configs if not exists
        cursor = db.execute("SELECT id FROM sms_api_configs LIMIT 1")
        if cursor.fetchone() is None:
            db.execute("INSERT INTO sms_api_configs (name, country, domain, spid, api_pwd, sender_name) VALUES ('Mexico', 'MX', '', '', '', '')")
            db.execute("INSERT INTO sms_api_configs (name, country, domain, spid, api_pwd, sender_name) VALUES ('Colombia', 'CO', '', '', '', '')")
        # Create default role_permissions. Seed team_admin/team_member with
        # sensible defaults so fresh installs do not show a blank sidebar.
        cursor = db.execute("SELECT role FROM role_permissions LIMIT 1")
        if cursor.fetchone() is None:
            db.execute("INSERT INTO role_permissions (role, permissions) VALUES (?, ?)",
                       ('admin', '[]'))
            db.execute("INSERT INTO role_permissions (role, permissions) VALUES (?, ?)",
                       ('team_admin', json.dumps(DEFAULT_ROLE_PERMISSIONS['team_admin'])))
            db.execute("INSERT INTO role_permissions (role, permissions) VALUES (?, ?)",
                       ('team_member', json.dumps(DEFAULT_ROLE_PERMISSIONS['team_member'])))
            db.commit()
        # Create default voice_config if not exists
        cursor = db.execute("SELECT id FROM voice_config LIMIT 1")
        if cursor.fetchone() is None:
            db.execute("INSERT INTO voice_config (provider, account_sid, auth_token, from_number, is_active) VALUES ('simulation', '', '', '', 1)")
            db.commit()
        # Seed voice_configs (one row per country, like sms_api_configs) and
        # migrate any per-country values from the legacy single voice_config row.
        cursor = db.execute("SELECT COUNT(*) AS cnt FROM voice_configs")
        _cnt_row = cursor.fetchone()
        _cnt = _cnt_row['cnt'] if isinstance(_cnt_row, dict) else (_cnt_row[0] if _cnt_row else 0)
        if _cnt == 0:
            legacy = db.execute(
                "SELECT api_domain_mx, appid_mx, accesskey_mx, from_number_mx, token_mx, token_mx_expiry,"
                " api_domain_co, appid_co, accesskey_co, from_number_co, token_co, token_co_expiry,"
                " api_domain_pe, appid_pe, accesskey_pe, from_number_pe, token_pe, token_pe_expiry"
                " FROM voice_config ORDER BY id LIMIT 1"
            ).fetchone()
            defaults = [('Mexico', 'mx'), ('Colombia', 'co'), ('Peru', 'pe')]
            offsets = {'mx': 0, 'co': 6, 'pe': 12}
            for name, cc in defaults:
                off = offsets[cc]
                domain = (legacy[off + 0] if legacy else '') or ''
                appid = (legacy[off + 1] if legacy else '') or ''
                accesskey = (legacy[off + 2] if legacy else '') or ''
                from_num = (legacy[off + 3] if legacy else '') or ''
                token = (legacy[off + 4] if legacy else '') or ''
                token_exp = int(legacy[off + 5] or 0) if legacy else 0
                provider = 'infin8linx' if (domain and appid and accesskey) else 'simulation'
                db.execute(
                    "INSERT INTO voice_configs (name, country, provider, api_domain, voice_appid, voice_accesskey,"
                    " from_number, voice_token, voice_token_expiry, is_active) VALUES (?,?,?,?,?,?,?,?,?,1)",
                    (name, cc, provider, domain, appid, accesskey, from_num, token, token_exp)
                )
            db.commit()
        # Migration: add new columns if they don't exist
        try:
            db.execute("ALTER TABLE sms_config ADD COLUMN domain TEXT DEFAULT ''")
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE sms_config ADD COLUMN spid TEXT DEFAULT ''")
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE sms_config ADD COLUMN api_pwd TEXT DEFAULT ''")
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE sms_records ADD COLUMN msgid TEXT DEFAULT ''")
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE sms_records ADD COLUMN api_code INTEGER DEFAULT 0")
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE sms_records ADD COLUMN api_msg TEXT DEFAULT ''")
        except Exception:
            pass
        db.commit()
        # Migration: add remark column to contacts if not exists
        try:
            db.execute("ALTER TABLE contacts ADD COLUMN remark TEXT DEFAULT ''")
            db.commit()
        except Exception:
            pass
        # Migration: add created_by to contacts if missing (older SQLite DBs)
        try:
            db.execute("ALTER TABLE contacts ADD COLUMN created_by INTEGER")
            db.commit()
        except Exception:
            pass
        # Migration: contact extra fields (app/amount/discount/payment link)
        for _col, _type in (
            ("app_name", "TEXT DEFAULT ''"),
            ("amount", "REAL DEFAULT 0"),
            ("discount_amount", "REAL DEFAULT 0"),
            ("payment_link", "TEXT DEFAULT ''"),
        ):
            try:
                db.execute(f"ALTER TABLE contacts ADD COLUMN {_col} {_type}")
                db.commit()
            except Exception:
                pass
        # Migration: voice_config extra columns for infin8linx provider
        for _col, _type in (
            ("voice_appid", "TEXT DEFAULT ''"),
            ("voice_accesskey", "TEXT DEFAULT ''"),
            ("voice_extnumber", "TEXT DEFAULT ''"),
            ("ext_pool_mx", "TEXT DEFAULT ''"),
            ("ext_pool_co", "TEXT DEFAULT ''"),
            ("ext_pool_pe", "TEXT DEFAULT ''"),
            ("voice_token", "TEXT DEFAULT ''"),
            ("voice_token_expiry", "INTEGER DEFAULT 0"),
            ("api_domain_mx", "TEXT DEFAULT ''"),
            ("appid_mx", "TEXT DEFAULT ''"),
            ("accesskey_mx", "TEXT DEFAULT ''"),
            ("from_number_mx", "TEXT DEFAULT ''"),
            ("api_domain_co", "TEXT DEFAULT ''"),
            ("appid_co", "TEXT DEFAULT ''"),
            ("accesskey_co", "TEXT DEFAULT ''"),
            ("from_number_co", "TEXT DEFAULT ''"),
            ("api_domain_pe", "TEXT DEFAULT ''"),
            ("appid_pe", "TEXT DEFAULT ''"),
            ("accesskey_pe", "TEXT DEFAULT ''"),
            ("from_number_pe", "TEXT DEFAULT ''"),
            ("token_mx", "TEXT DEFAULT ''"),
            ("token_mx_expiry", "INTEGER DEFAULT 0"),
            ("token_co", "TEXT DEFAULT ''"),
            ("token_co_expiry", "INTEGER DEFAULT 0"),
            ("token_pe", "TEXT DEFAULT ''"),
            ("token_pe_expiry", "INTEGER DEFAULT 0"),
            ("country", "TEXT DEFAULT ''"),
        ):
            try:
                db.execute(f"ALTER TABLE voice_config ADD COLUMN {_col} {_type}")
                db.commit()
            except Exception:
                pass
        # Migration: record the extension used per call
        try:
            db.execute("ALTER TABLE voice_records ADD COLUMN extnumber TEXT DEFAULT ''")
            db.commit()
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE voice_records ADD COLUMN country TEXT DEFAULT ''")
            db.commit()
        except Exception:
            pass
        # Migration: CDR/callback correlation + outcome fields
        for _col, _type in (
            ("customuuid", "TEXT DEFAULT ''"),
            ("provider_uuid", "TEXT DEFAULT ''"),
            ("record_file", "TEXT DEFAULT ''"),
            ("hangupcause", "INTEGER DEFAULT 0"),
            ("answer_at", "TEXT"),
        ):
            try:
                db.execute(f"ALTER TABLE voice_records ADD COLUMN {_col} {_type}")
                db.commit()
            except Exception:
                pass
        try:
            db.execute("CREATE INDEX IF NOT EXISTS idx_voice_records_customuuid ON voice_records(customuuid)")
            db.commit()
        except Exception:
            pass
        # Migration: voice_configs detected protocol (https default)
        try:
            db.execute("ALTER TABLE voice_configs ADD COLUMN voice_scheme TEXT NOT NULL DEFAULT 'https'")
            db.commit()
        except Exception:
            pass
        # Migration: add team_creator_id to users if not exists
        try:
            db.execute("ALTER TABLE users ADD COLUMN team_creator_id INTEGER")
            db.commit()
        except Exception:
            pass
        # Migration: update old 'employee' role to 'team_member'
        try:
            db.execute("UPDATE users SET role='team_member' WHERE role='employee'")
            db.commit()
        except Exception:
            pass
        # Migration: add daily_limit to users if not exists
        try:
            db.execute("ALTER TABLE users ADD COLUMN daily_limit INTEGER DEFAULT 0")
            db.commit()
        except Exception:
            pass
        # Migration: add permissions column to users if not exists
        try:
            db.execute("ALTER TABLE users ADD COLUMN permissions TEXT DEFAULT ''")
            db.commit()
        except Exception:
            pass
        # Migration: add extnumber (fixed phone/extension) to users if not exists
        try:
            db.execute("ALTER TABLE users ADD COLUMN extnumber TEXT DEFAULT NULL")
            db.commit()
        except Exception:
            pass
        # Migration: add country (mx/co/pe) tag to users
        try:
            db.execute("ALTER TABLE users ADD COLUMN country TEXT DEFAULT NULL")
            db.commit()
        except Exception:
            pass
        # user_categories: clasificacion de empleados + retencion de contactos
        try:
            db.execute("""
                CREATE TABLE IF NOT EXISTS user_categories (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT UNIQUE NOT NULL,
                    retention_days INTEGER NOT NULL DEFAULT 0,
                    is_default INTEGER NOT NULL DEFAULT 0,
                    created_at TEXT NOT NULL DEFAULT (datetime('now')),
                    updated_at TEXT NOT NULL DEFAULT (datetime('now'))
                )
            """)
            db.execute("ALTER TABLE users ADD COLUMN category_id INTEGER")
            db.commit()
        except Exception:
            db.rollback()
        try:
            row = db.execute("SELECT id FROM user_categories WHERE is_default=1 LIMIT 1").fetchone()
            if not row:
                db.execute("INSERT INTO user_categories (name, retention_days, is_default) VALUES (?, ?, 1)", ('General', 7))
            db.commit()
        except Exception:
            db.rollback()
        # Migration: retention days must be within 1..7. Clamp existing rows.
        try:
            db.execute(
                "UPDATE user_categories SET retention_days = 7 "
                "WHERE retention_days > 7 OR retention_days < 1 OR retention_days IS NULL"
            )
            db.commit()
        except Exception:
            db.rollback()
        # Migration: recreate users table with new role CHECK constraint
        try:
            cursor = db.execute("SELECT sql FROM sqlite_master WHERE name='users'")
            schema_sql = cursor.fetchone()[0]
            if "'admin', 'employee'" in schema_sql:
                db.execute("PRAGMA foreign_keys=OFF")
                db.execute("DROP TABLE IF EXISTS users_old")
                db.execute("ALTER TABLE users RENAME TO users_old")
                db.execute("""CREATE TABLE users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    full_name TEXT NOT NULL DEFAULT '',
                    role TEXT NOT NULL DEFAULT 'team_member' CHECK(role IN ('admin', 'team_admin', 'team_member')),
                    is_active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL DEFAULT (datetime('now')),
                    updated_at TEXT NOT NULL DEFAULT (datetime('now')),
                    team_creator_id INTEGER,
                    permissions TEXT DEFAULT '',
                    extnumber TEXT DEFAULT NULL,
                    country TEXT DEFAULT NULL
                )""")
                db.execute("INSERT INTO users (id, username, password_hash, full_name, role, is_active, created_at, updated_at, team_creator_id, extnumber, country) SELECT id, username, password_hash, full_name, CASE WHEN role='employee' THEN 'team_member' ELSE role END, is_active, created_at, updated_at, team_creator_id, extnumber, country FROM users_old")
                db.execute("DROP TABLE users_old")
                db.execute("PRAGMA foreign_keys=ON")
                db.commit()
        except Exception as e:
            print(f"Migration users table error: {e}")
            db.execute("PRAGMA foreign_keys=ON")
        except Exception:
            pass
        # Migration: create team_config table
        try:
            db.execute("""
                CREATE TABLE IF NOT EXISTS team_config (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    team_admin_id INTEGER NOT NULL UNIQUE,
                    daily_sms_limit INTEGER DEFAULT 100,
                    api_config_id INTEGER DEFAULT NULL,
                    FOREIGN KEY (team_admin_id) REFERENCES users(id),
                    FOREIGN KEY (api_config_id) REFERENCES sms_api_configs(id) ON DELETE SET NULL
                )
            """)
            db.commit()
        except Exception:
            pass

        # Migration: create sms_api_configs table and migrate from sms_config
        try:
            cursor = db.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='sms_api_configs'")
            if cursor.fetchone() is None:
                db.execute('''CREATE TABLE sms_api_configs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE,
                    country TEXT NOT NULL DEFAULT '',
                    domain TEXT DEFAULT '',
                    spid TEXT DEFAULT '',
                    api_pwd TEXT DEFAULT '',
                    sender_name TEXT DEFAULT '',
                    is_active INTEGER DEFAULT 1,
                    created_at TEXT NOT NULL DEFAULT (datetime('now')),
                    updated_at TEXT NOT NULL DEFAULT (datetime('now'))
                )''')
                # Migrate existing sms_config data
                old_cursor = db.execute("SELECT * FROM sms_config LIMIT 1")
                old_config = old_cursor.fetchone()
                if old_config:
                    db.execute("INSERT INTO sms_api_configs (name, country, domain, spid, api_pwd, sender_name) VALUES ('Mexico', 'MX', ?, ?, ?, ?)",
                        (old_config['domain'] or '', old_config['spid'] or '', old_config['api_pwd'] or '', old_config['sender_name'] or ''))
                    db.execute("INSERT INTO sms_api_configs (name, country, domain, spid, api_pwd, sender_name) VALUES ('Colombia', 'CO', '', '', '', '')")
                else:
                    db.execute("INSERT INTO sms_api_configs (name, country, domain, spid, api_pwd, sender_name) VALUES ('Mexico', 'MX', '', '', '', '')")
                    db.execute("INSERT INTO sms_api_configs (name, country, domain, spid, api_pwd, sender_name) VALUES ('Colombia', 'CO', '', '', '', '')")
                db.commit()
        except Exception:
            pass

        # Migration: add api_config_id to team_config
        try:
            cursor = db.execute("PRAGMA table_info(team_config)")
            cols = [row[1] for row in cursor.fetchall()]
            if 'api_config_id' not in cols:
                db.execute("ALTER TABLE team_config ADD COLUMN api_config_id INTEGER DEFAULT NULL REFERENCES sms_api_configs(id) ON DELETE SET NULL")
                db.commit()
        except Exception:
            pass

        # Migration: add last_login_ip and last_login_at to users
        try:
            if db_type == 'postgres':
                cursor = db.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'users'")
                cols = [(row['column_name'] if isinstance(row, dict) else row[0]) for row in cursor.fetchall()]
                if 'last_login_ip' not in cols:
                    db.execute("ALTER TABLE users ADD COLUMN last_login_ip TEXT DEFAULT ''")
                if 'last_login_at' not in cols:
                    db.execute("ALTER TABLE users ADD COLUMN last_login_at TEXT DEFAULT NULL")
                if 'session_token' not in cols:
                    db.execute("ALTER TABLE users ADD COLUMN session_token TEXT DEFAULT ''")
            else:
                cursor = db.execute("PRAGMA table_info(users)")
                cols = [row[1] for row in cursor.fetchall()]
                if 'last_login_ip' not in cols:
                    db.execute("ALTER TABLE users ADD COLUMN last_login_ip TEXT DEFAULT ''")
                if 'last_login_at' not in cols:
                    db.execute("ALTER TABLE users ADD COLUMN last_login_at TEXT DEFAULT NULL")
                if 'session_token' not in cols:
                    db.execute("ALTER TABLE users ADD COLUMN session_token TEXT DEFAULT ''")
            db.commit()
        except Exception as e:
            print(f"Migration error: {e}")

        # Performance indexes (idempotent). Critical for large teams:
        # the send_logs page and SMS records/statistics pages sort/filter by
        # these columns. Without indexes every request does a full table scan
        # and sort, which degrades linearly with row count.
        index_statements = [
            # Activity logs (admin log page): ORDER BY created_at DESC
            "CREATE INDEX IF NOT EXISTS idx_send_logs_created_at ON send_logs(created_at DESC)",
            # SMS records: list filtering by owner + time range, plus scheduling
            "CREATE INDEX IF NOT EXISTS idx_sms_records_created_by ON sms_records(created_by)",
            "CREATE INDEX IF NOT EXISTS idx_sms_records_created_at ON sms_records(created_at DESC)",
            "CREATE INDEX IF NOT EXISTS idx_sms_records_status_created_at ON sms_records(status, created_at DESC)",
            "CREATE INDEX IF NOT EXISTS idx_sms_records_status_sent_at ON sms_records(status, sent_at DESC)",
            "CREATE INDEX IF NOT EXISTS idx_sms_records_scheduled ON sms_records(status, scheduled_at)",
            # Contacts/groups lookups used by team-scoped queries
            "CREATE INDEX IF NOT EXISTS idx_contacts_created_by ON contacts(created_by)",
            "CREATE INDEX IF NOT EXISTS idx_contacts_group_id ON contacts(group_id)",
        ]
        for stmt in index_statements:
            try:
                db.execute(stmt)
            except Exception as e:
                # A missing table (e.g. older schema) shouldn't block startup;
                # the CREATE TABLE statements above run before this.
                print(f"Index warning: {e}")
        try:
            db.commit()
        except Exception:
            pass

        db.close()

# ============================================================
# Auth helpers
# ============================================================

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def generate_password(length=10):
    alphabet = string.ascii_letters + string.digits
    while True:
        pwd = ''.join(secrets.choice(alphabet) for _ in range(length))
        # Asegurar que contenga al menos una mayuscula, una minuscula y un digito
        if (any(c.islower() for c in pwd) and any(c.isupper() for c in pwd)
                and any(c.isdigit() for c in pwd)):
            return pwd

# ---------- Configuracion de facturacion ----------
# Coste por SMS enviado (se cobra por intento/envio, exitoso o fallido).
BILLING_PRICE_KEY = 'sms_unit_price'
# Precio por defecto si el administrador aun no lo configura.
DEFAULT_SMS_UNIT_PRICE = 0.0

def get_sms_unit_price(db=None):
    """Devuelve el precio por SMS configurado por el administrador (float)."""
    own = db is None
    if own:
        db = get_db()
    try:
        row = db.execute(
            "SELECT value FROM system_settings WHERE key=?", (BILLING_PRICE_KEY,)
        ).fetchone()
        if row and row['value'] not in (None, ''):
            return round(float(row['value']), 6)
    except (ValueError, TypeError):
        pass
    return DEFAULT_SMS_UNIT_PRICE

def set_sms_unit_price(price, db=None):
    own = db is None
    if own:
        db = get_db()
    is_pg = getattr(db, 'db_type', 'sqlite') == 'postgres'
    ts = "CURRENT_TIMESTAMP" if is_pg else "datetime('now')"
    if is_pg:
        db.execute(
            "INSERT INTO system_settings(key, value, updated_at) VALUES(%s, %s, CURRENT_TIMESTAMP) "
            "ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = CURRENT_TIMESTAMP",
            (BILLING_PRICE_KEY, str(price))
        )
    else:
        db.execute(
            "INSERT INTO system_settings(key, value, updated_at) VALUES(?, ?, datetime('now')) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=datetime('now')",
            (BILLING_PRICE_KEY, str(price))
        )
    if own:
        db.commit()

def calc_cost(count, price):
    """Calcula el coste total redondeado a 2 decimales."""
    try:
        return round(float(count or 0) * float(price or 0), 2)
    except (ValueError, TypeError):
        return 0.0

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'No autorizado'}), 401
        user_row = get_db().execute("SELECT * FROM users WHERE id=? AND is_active=1", (session['user_id'],)).fetchone()
        if not user_row:
            session.clear()
            return jsonify({'error': 'Sesion expirada'}), 401
        user = dict(user_row)
        # Estrategia B: solo la ultima sesion iniciada es valida.
        # Si el token de la cookie no coincide con el de la base, se invalida.
        if not user.get('session_token') or session.get('session_token') != user.get('session_token'):
            session.clear()
            return jsonify({'error': 'Sesion invalidada. Se ha iniciado sesion en otro dispositivo.'}), 401
        g.user = user
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'No autorizado'}), 401
        user_row = get_db().execute("SELECT * FROM users WHERE id=? AND is_active=1", (session['user_id'],)).fetchone()
        if not user_row:
            session.clear()
            return jsonify({'error': 'Sesion expirada'}), 401
        user = dict(user_row)
        if not user.get('session_token') or session.get('session_token') != user.get('session_token'):
            session.clear()
            return jsonify({'error': 'Sesion invalidada. Se ha iniciado sesion en otro dispositivo.'}), 401
        if user['role'] != 'admin':
            return jsonify({'error': 'Permisos insuficientes'}), 403
        g.user = user
        return f(*args, **kwargs)
    return decorated

def manager_required(f):
    """Require admin or team_admin role."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'No autorizado'}), 401
        user_row = get_db().execute("SELECT * FROM users WHERE id=? AND is_active=1", (session['user_id'],)).fetchone()
        if not user_row:
            session.clear()
            return jsonify({'error': 'Sesion expirada'}), 401
        user = dict(user_row)
        if not user.get('session_token') or session.get('session_token') != user.get('session_token'):
            session.clear()
            return jsonify({'error': 'Sesion invalidada. Se ha iniciado sesion en otro dispositivo.'}), 401
        if user['role'] not in ('admin', 'team_admin'):
            return jsonify({'error': 'Permisos insuficientes'}), 403
        g.user = user
        return f(*args, **kwargs)
    return decorated

# ============================================================
# SMS API Integration (infin8linx)
# ============================================================

# Status code mapping from API docs
SMS_STATUS_CODES = {
    0: 'Exito',
    10: 'Parametros requeridos faltantes',
    11: 'Parametros invalidos',
    12: 'Cuenta/contrasena incorrectos',
    13: 'Saldo insuficiente',
    14: 'IP no en lista blanca',
    15: 'Limite de frecuencia excedido',
    16: 'Problema con el contenido',
    17: 'Error de codificacion',
    18: 'Problema con la firma',
    19: 'Error de plantilla',
    20: 'SMS no encontrado',
    21: 'Error del sistema',
    22: 'Numero invalido',
    23: 'Contenido sensible',
    24: 'Numero en lista negra',
}

def str_to_hex(s):
    """Convert a string to hex encoding (UTF-8)."""
    return s.encode('utf-8').hex()

def generate_api_pwd(spid, api_pwd, timestamp):
    """Generate encrypted password: MD5(spid + '00000000' + pwd + timestamp), lowercase."""
    raw = f"{spid}00000000{api_pwd}{timestamp}"
    return hashlib.md5(raw.encode('utf-8')).hexdigest()

def get_sms_api_config(config_id=None):
    """Get SMS API configuration from database. If config_id given, get that config; otherwise get first active config."""
    db = get_db()
    if config_id:
        config = db.execute("SELECT * FROM sms_api_configs WHERE id = ?", (config_id,)).fetchone()
    else:
        # Prefer a fully-configured active record (domain/spid/api_pwd all present)
        config = db.execute(
            "SELECT * FROM sms_api_configs WHERE is_active = 1 "
            "AND domain IS NOT NULL AND domain != '' "
            "AND spid IS NOT NULL AND spid != '' "
            "AND api_pwd IS NOT NULL AND api_pwd != '' "
            "ORDER BY id LIMIT 1"
        ).fetchone()
        # Fallback: if no fully-configured record, return the first active one
        if not config:
            config = db.execute(
                "SELECT * FROM sms_api_configs WHERE is_active = 1 ORDER BY id LIMIT 1"
            ).fetchone()
    if not config:
        return None
    return {
        'id': config['id'],
        'name': config['name'] or '',
        'country': config['country'] or '',
        'domain': config['domain'] or '',
        'spid': config['spid'] or '',
        'api_pwd': config['api_pwd'] or '',
        'sender_name': config['sender_name'] or '',
    }

def get_team_sms_config(user_id):
    """Get SMS API config for a user based on their team."""
    db = get_db()
    # Find user's team admin
    user = db.execute("SELECT id, role, team_creator_id FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        return get_sms_api_config()
    if user['role'] == 'admin':
        return get_sms_api_config()
    # Find team admin's config
    team_admin_id = user['team_creator_id']
    if not team_admin_id:
        return get_sms_api_config()
    tc = db.execute("SELECT api_config_id FROM team_config WHERE team_admin_id = ?", (team_admin_id,)).fetchone()
    if tc and tc['api_config_id']:
        return get_sms_api_config(tc['api_config_id'])
    return get_sms_api_config()

def normalize_phone(phone, default_country='+52'):
    """Normalize a phone number to E.164-ish form: +<digits>.

    Rules:
    - Strip spaces, dashes, parentheses.
    - If it already starts with '+' or '00', keep it (00 converted to +).
    - If it starts with the country code without '+' (e.g. '52155...' with 12+ digits),
      prepend '+'.
    - Otherwise prepend the default country (Mexico +52 by default), removing any
      leading domestic '0'/'1' long-distance prefix.
    Returns the normalized string; invalid (empty/too short) input returned as-is.
    """
    if not phone:
        return phone
    p = re.sub(r'[\s\-\(\)]+', '', str(phone))
    if not p:
        return phone
    if p.startswith('+'):
        digits = p[1:]
        return '+' + digits
    if p.startswith('00'):
        return '+' + p[2:]
    digits = p
    # Already has country code (>=12 digits for MX 52+10, or generic 11+ with 52 prefix)
    if len(digits) == 12 and digits.startswith('52'):
        return '+' + digits
    if len(digits) >= 11 and not digits.startswith('0') and not digits.startswith('1'):
        # Looks like it already includes a country code
        return '+' + digits
    # Domestic number: strip leading 0/1 prefix for MX, prepend default country
    if digits.startswith('0') or digits.startswith('1'):
        digits = digits.lstrip('01') or digits
    return default_country + digits


def phone_to_das(phone):
    """Convert a normalized phone (+52...) to the API 'das' format (0052...)."""
    p = normalize_phone(phone)
    if p.startswith('+'):
        return '00' + p[1:]
    return p


def _fmt_money(value):
    try:
        v = float(value or 0)
    except (TypeError, ValueError):
        v = 0.0
    return f"{v:.2f}"


def build_contact_template_cache(db, phones):
    """Bulk-load contact fields needed for template variable substitution.

    Returns a dict keyed by normalized phone with keys:
    name, app_name, amount, discount_amount, payment_link.
    """
    cache = {}
    norm_phones = []
    for raw in phones or []:
        p = normalize_phone((raw or '').strip())
        if p:
            norm_phones.append(p)
    if not norm_phones:
        return cache
    stripped = [p.lstrip('+') for p in norm_phones]
    placeholders = ','.join('?' for _ in stripped)
    try:
        rows = db.execute(
            "SELECT phone, name, app_name, amount, discount_amount, payment_link "
            "FROM contacts WHERE REPLACE(REPLACE(REPLACE(phone, ' ', ''), '-', ''), '+', '') IN (" + placeholders + ")",
            stripped
        ).fetchall()
    except Exception:
        rows = []
    for row in rows:
        r = dict(row)
        key = normalize_phone(r.get('phone') or '')
        entry = {
            'name': r.get('name') or '',
            'app_name': r.get('app_name') or '',
            'amount': r.get('amount') if r.get('amount') is not None else 0,
            'discount_amount': r.get('discount_amount') if r.get('discount_amount') is not None else 0,
            'payment_link': r.get('payment_link') or '',
        }
        cache[key] = entry
        cache[key.lstrip('+')] = entry
    return cache


def _normalize_extnumber(value):
    """Normalize an extension: strip, collapse internal whitespace, keep digits/letters/-/_/*#."""
    if value is None:
        return ''
    ext = ' '.join(str(value).strip().split())
    # Drop a header row such as "Extension" / "extension" / "ext".
    if ext.lower() in ('extension', 'extensiones', 'ext', 'extnumber', 'telefono', 'phone', '分机号'):
        return ''
    return ext


def _find_user_by_extnumber(extnumber, exclude_id=None):
    """Return the first active user already bound to the given extension, or None.

    Extensions are matched case-insensitively and after whitespace normalization.
    """
    ext = _normalize_extnumber(extnumber)
    if not ext:
        return None
    db = get_db()
    q = "SELECT * FROM users WHERE is_active = 1 AND LOWER(TRIM(extnumber)) = LOWER(?)"
    params = [ext]
    if exclude_id:
        q += " AND id != ?"
        params.append(exclude_id)
    return db.execute(q, tuple(params)).fetchone()


def _get_extension_pool(country=None):
    """Legacy fallback pool from the single voice_config row.

    Extensions are now managed in the dedicated `extensions` catalog (see the
    Extensiones page); this only supports older installs that never used it.
    The `country` argument is accepted for call-site compatibility but the
    catalog is the authoritative source.
    """
    try:
        db = get_db()
        row = db.execute("SELECT voice_extnumber FROM voice_config ORDER BY id LIMIT 1").fetchone()
        if row and row['voice_extnumber']:
            return _parse_extension_pool(row['voice_extnumber'])
    except Exception:
        pass
    return []


def normalize_country(country):
    """Normalize a country code to one of mx/co/pe or ''."""
    c = (country or '').strip().lower()
    return c if c in ('mx', 'co', 'pe') else ''


def allocate_extension(exclude_id=None, country=None):
    """Auto-assign a free extension for a country.

    Prefers the extensions catalog (managed from the Extensiones page); when
    the catalog is empty (older installs that only configured pool strings),
    falls back to the voice_config pool strings. Returns the chosen extension
    or '' if none is available. The caller must invoke
    _extensions_mark_assigned(ext, user_id, country) after the user is created.
    """
    import random as _random
    _extensions_seed_from_config()
    country = normalize_country(country)
    if _extensions_table_ready():
        picked = _extensions_allocate(country, exclude_id=exclude_id)
        if picked:
            return picked
    pool = _get_extension_pool(country)
    if not pool:
        return ''
    db = get_db()
    # Extensions already assigned to active users (normalized, lowercase).
    taken_rows = db.execute(
        "SELECT extnumber FROM users WHERE is_active = 1 AND extnumber IS NOT NULL AND TRIM(extnumber) <> ''"
    ).fetchall()
    taken = set()
    for row in taken_rows:
        val = row['extnumber'] if not isinstance(row, tuple) else row[0]
        if exclude_id:
            # The excluded user already holds its own extension; when re-assigning
            # for that user we should not count it as taken.
            owner = db.execute(
                "SELECT extnumber FROM users WHERE id = ?", (exclude_id,)
            ).fetchone()
            own = owner['extnumber'] if owner and not isinstance(owner, tuple) else ''
            if own and _normalize_extnumber(own).lower() == _normalize_extnumber(val).lower():
                continue
        taken.add(_normalize_extnumber(val).lower())
    free = [ext for ext in pool if _normalize_extnumber(ext).lower() not in taken]
    if not free:
        return ''
    return _random.choice(free)


# ---------------------------------------------------------------------------
# Extensions catalog (separate management page)
# ---------------------------------------------------------------------------

EXT_COUNTRIES = ('', 'mx', 'co', 'pe')


def _extensions_table_ready():
    """Return True if the extensions table exists."""
    db = get_db()
    if db.db_type == 'postgres':
        row = db.execute(
            "SELECT 1 FROM information_schema.tables WHERE table_name='extensions'"
        ).fetchone()
    else:
        row = db.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='extensions'"
        ).fetchone()
    return row is not None


def _extensions_seed_from_config(force=False):
    """Import legacy pool strings (ext_pool_mx/co/pe, voice_extnumber) into the
    extensions table on first use. Idempotent: only seeds when the table is empty
    (or force=True). Already-assigned extensions (found in users.extnumber) are
    linked via assigned_to.
    """
    if not _extensions_table_ready():
        return
    db = get_db()
    if not force:
        cnt = db.execute("SELECT COUNT(*) AS c FROM extensions").fetchone()
        if cnt and (cnt['c'] if not isinstance(cnt, tuple) else cnt[0]):
            return
    try:
        cfg = get_voice_config()
    except Exception:
        cfg = {}
    country_map = [
        ('mx', cfg.get('ext_pool_mx')),
        ('co', cfg.get('ext_pool_co')),
        ('pe', cfg.get('ext_pool_pe')),
        ('', cfg.get('voice_extnumber')),
    ]
    for country, raw in country_map:
        for ext in _parse_extension_pool(raw):
            norm = _normalize_extnumber(ext)
            if not norm:
                continue
            exists = db.execute(
                "SELECT id FROM extensions WHERE LOWER(TRIM(extnumber))=LOWER(?) AND country=?",
                (norm, country)
            ).fetchone()
            if exists:
                continue
            owner = _find_user_by_extnumber(norm)
            db.execute(
                "INSERT INTO extensions (extnumber, country, assigned_to) VALUES (?,?,?)",
                (norm, country, owner['id'] if owner else None)
            )
    db.commit()


def _extensions_allocate(country, exclude_id=None):
    """Pick and atomically mark a free extension for the country from the
    extensions catalog. Returns the extension string or '' if none available.
    """
    import random as _random
    db = get_db()
    country = normalize_country(country)
    rows = db.execute(
        "SELECT id, extnumber FROM extensions "
        "WHERE country=? AND (assigned_to IS NULL) "
        "ORDER BY id",
        (country,)
    ).fetchall()
    candidates = []
    for r in rows:
        ext = r['extnumber'] if not isinstance(r, tuple) else r[1]
        eid = r['id'] if not isinstance(r, tuple) else r[0]
        # Extra safety: ensure no active user holds it (assigned_to should be
        # NULL already, but users assigned before the table existed may exist).
        if _find_user_by_extnumber(ext, exclude_id=exclude_id):
            continue
        candidates.append((eid, ext))
    if not candidates:
        return ''
    eid, ext = _random.choice(candidates)
    return ext


def _extensions_mark_assigned(extnumber, user_id, country):
    """Link an extension row to a user after allocation."""
    if not _extensions_table_ready() or not extnumber:
        return
    db = get_db()
    country = normalize_country(country)
    norm = _normalize_extnumber(extnumber)
    row = db.execute(
        "SELECT id FROM extensions WHERE LOWER(TRIM(extnumber))=LOWER(?) AND country=?",
        (norm, country)
    ).fetchone()
    if row:
        eid = row['id'] if not isinstance(row, tuple) else row[0]
        db.execute("UPDATE extensions SET assigned_to=? WHERE id=?", (user_id, eid))
        db.commit()
    else:
        # Auto-register an extension that came from the legacy pool string.
        owner = _find_user_by_extnumber(norm)
        db.execute(
            "INSERT INTO extensions (extnumber, country, assigned_to) VALUES (?,?,?)",
            (norm, country, user_id if not owner else owner['id'])
        )
        db.commit()


def _extensions_release(extnumber):
    """Clear assigned_to for an extension when a user releases it."""
    if not _extensions_table_ready() or not extnumber:
        return
    db = get_db()
    norm = _normalize_extnumber(extnumber)
    db.execute(
        "UPDATE extensions SET assigned_to=NULL WHERE LOWER(TRIM(extnumber))=LOWER(?)",
        (norm,)
    )
    db.commit()


def apply_template_vars(text, phone, contact_names=None, contact_cache=None):
    """Replace {nombre}/{telefono}/{app_name}/{amount}/{discount}/{payment_link} placeholders."""
    if not text:
        return text
    raw_phone = (phone or '').strip()
    norm_phone = normalize_phone(raw_phone)
    name = ''
    if contact_names:
        name = contact_names.get(raw_phone, '') or contact_names.get(norm_phone, '') or ''
    app_name = ''
    amount = 0
    discount_amount = 0
    payment_link = ''
    if contact_cache:
        entry = contact_cache.get(norm_phone) or contact_cache.get(norm_phone.lstrip('+'))
        if entry:
            if not name:
                name = entry.get('name', '')
            app_name = entry.get('app_name', '')
            amount = entry.get('amount', 0)
            discount_amount = entry.get('discount_amount', 0)
            payment_link = entry.get('payment_link', '')
    msg = text.replace('{nombre}', name).replace('{telefono}', norm_phone)
    msg = msg.replace('{app_name}', app_name)
    msg = msg.replace('{amount}', _fmt_money(amount))
    msg = msg.replace('{discount}', _fmt_money(discount_amount))
    msg = msg.replace('{payment_link}', payment_link)
    return msg


def is_sms_api_configured(user_id=None):
    """Check if SMS API is properly configured."""
    if user_id:
        config = get_team_sms_config(user_id)
    else:
        config = get_sms_api_config()
    if not config:
        return False
    return bool(config['domain'] and config['spid'] and config['api_pwd'])

def sms_api_send_single(phone, content, config=None):
    """Send a single SMS via the API.
    GET /sms/send?spid=xx&pwd=xx&das=xx&timestamp=xx&sm=xx
    Returns: dict with code, msg, data (msgid, parts, state)
    """
    if config is None:
        config = get_sms_api_config()
    if not config or not (config.get('domain') and config.get('spid') and config.get('api_pwd')):
        return {'code': -1, 'msg': 'API SMS no configurada', 'data': None}

    timestamp = str(int(time.time()))
    pwd = generate_api_pwd(config['spid'], config['api_pwd'], timestamp)
    sm_hex = str_to_hex(content)

    # Normalize and convert phone format: +52... -> 0052...
    das = phone_to_das(phone)

    params = {
        'spid': config['spid'],
        'pwd': pwd,
        'timestamp': timestamp,
        'das': das,
        'sm': sm_hex,
    }
    if config['sender_name']:
        params['senderid'] = config['sender_name']

    # Build full URL with query string
    from urllib.parse import urlencode
    url = f"https://{config['domain']}/sms/send?{urlencode(params)}"

    try:
        resp = http_requests.get(url, timeout=15)
        result = resp.json()
        return result
    except http_requests.exceptions.Timeout:
        return {'code': -2, 'msg': 'Tiempo de espera agotado', 'data': None}
    except http_requests.exceptions.ConnectionError:
        return {'code': -3, 'msg': 'Error de conexion con el proveedor', 'data': None}
    except Exception as e:
        return {'code': -4, 'msg': f'Error: {str(e)}', 'data': None}

def sms_api_send_batch(phone_content_pairs, config=None):
    """Send multiple SMS via the API (max 200 per request).
    POST /sms/rsend Content-Type: application/x-www-form-urlencoded
    phone_content_pairs: list of (phone, content) tuples
    Returns: dict with code, msg, data (list of {das, msgid, parts, state})
    """
    if config is None:
        config = get_sms_api_config()
    if not config or not (config.get('domain') and config.get('spid') and config.get('api_pwd')):
        return {'code': -1, 'msg': 'API SMS no configurada', 'data': None}

    timestamp = str(int(time.time()))
    pwd = generate_api_pwd(config['spid'], config['api_pwd'], timestamp)

    # Build dasm: phone,hex_content/phone,hex_content...
    parts = []
    for phone, content in phone_content_pairs:
        hex_content = str_to_hex(content)
        # Normalize and convert phone format: +52... -> 0052...
        das = phone_to_das(phone)
        parts.append(f"{das},{hex_content}")
    dasm = '/'.join(parts)

    payload = {
        'spid': config['spid'],
        'pwd': pwd,
        'timestamp': timestamp,
        'dasm': dasm,
    }
    if config['sender_name']:
        payload['senderid'] = config['sender_name']

    url = f"https://{config['domain']}/sms/rsend"

    try:
        resp = http_requests.post(url, data=payload, timeout=30)
        result = resp.json()
        return result
    except http_requests.exceptions.Timeout:
        return {'code': -2, 'msg': 'Tiempo de espera agotado', 'data': None}
    except http_requests.exceptions.ConnectionError:
        return {'code': -3, 'msg': 'Error de conexion con el proveedor', 'data': None}
    except Exception as e:
        return {'code': -4, 'msg': f'Error: {str(e)}', 'data': None}

def sms_api_query_status(msgids, config=None):
    """Query SMS delivery status.
    GET /sms/state?spid=xx&pwd=xx&timestamp=xx&msgid=xx
    msgids: list of msgid strings (max 100)
    Returns: dict with code, msg, data (list of {msgid, state})
    state: 0=no receipt, 1=success, 2=failed
    """
    if config is None:
        config = get_sms_api_config()
    if not config or not (config.get('domain') and config.get('spid') and config.get('api_pwd')):
        return {'code': -1, 'msg': 'API SMS no configurada', 'data': None}

    timestamp = str(int(time.time()))
    pwd = generate_api_pwd(config['spid'], config['api_pwd'], timestamp)
    msgid_str = ','.join(msgids[:100])

    url = f"https://{config['domain']}/sms/state"
    params = {
        'spid': config['spid'],
        'pwd': pwd,
        'timestamp': timestamp,
        'msgid': msgid_str,
    }

    try:
        resp = http_requests.get(url, params=params, timeout=15)
        result = resp.json()
        return result
    except http_requests.exceptions.Timeout:
        return {'code': -2, 'msg': 'Tiempo de espera agotado', 'data': None}
    except http_requests.exceptions.ConnectionError:
        return {'code': -3, 'msg': 'Error de conexion con el proveedor', 'data': None}
    except Exception as e:
        return {'code': -4, 'msg': f'Error: {str(e)}', 'data': None}

def sms_api_check_charset(content, config=None):
    """Check SMS content charset and billing parts.
    GET /sms/charset?spid=xx&pwd=xx&sm=xx&timestamp=xx
    Returns: dict with code, msg, data (charset, parts, single, detail)
    """
    if config is None:
        config = get_sms_api_config()
    if not config or not (config.get('domain') and config.get('spid') and config.get('api_pwd')):
        return {'code': -1, 'msg': 'API SMS no configurada', 'data': None}

    timestamp = str(int(time.time()))
    pwd = generate_api_pwd(config['spid'], config['api_pwd'], timestamp)
    sm_hex = str_to_hex(content)

    url = f"https://{config['domain']}/sms/charset"
    params = {
        'spid': config['spid'],
        'pwd': pwd,
        'timestamp': timestamp,
        'sm': sm_hex,
    }

    try:
        resp = http_requests.get(url, params=params, timeout=15)
        result = resp.json()
        return result
    except http_requests.exceptions.Timeout:
        return {'code': -2, 'msg': 'Tiempo de espera agotado', 'data': None}
    except http_requests.exceptions.ConnectionError:
        return {'code': -3, 'msg': 'Error de conexion con el proveedor', 'data': None}
    except Exception as e:
        return {'code': -4, 'msg': f'Error: {str(e)}', 'data': None}

# ============================================================
# Frontend route
# ============================================================

@app.after_request
def add_cache_headers(response):
    """Add cache headers: HTML must not be cached (so asset version bumps take effect);
    static files under /static/ use 1 hour cache with ETag validation."""
    if request.path == '/' or request.path.startswith('/api/'):
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    elif request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'public, max-age=3600'
    return response

@app.route('/')
def index():
    return render_template(
        'index.html',
        app_environment=app.config['APP_ENVIRONMENT'],
        app_label=app.config['APP_LABEL'],
    )

# ============================================================
# Auth API
# ============================================================

ROLE_LABELS = {
    'admin': 'Administrador del Sistema',
    'team_admin': 'Administrador de Equipo',
    'team_member': 'Miembro de Equipo'
}

# Paises soportados para extensiones/agentes (marcado y pool por pais).
COUNTRY_LABELS = {
    'mx': 'Mexico',
    'co': 'Colombia',
    'pe': 'Peru'
}

@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    if not username or not password:
        return jsonify({'error': 'Usuario y contrasena son requeridos'}), 400
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    if not user or user['password_hash'] != hash_password(password):
        return jsonify({'error': 'Credenciales invalidas'}), 401
    if not user['is_active']:
        return jsonify({'error': 'Cuenta desactivada. Contacte al administrador.'}), 403

    # Single-session policy: generate a new token for this login.
    # Any older session with a different token is invalidated.
    session_token = secrets.token_urlsafe(32)
    session['user_id'] = user['id']
    session['role'] = user['role']
    session['session_token'] = session_token

    # Record login IP and rotate session token
    login_ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()
    db.execute(
        "UPDATE users SET last_login_ip=?, last_login_at=datetime('now'), session_token=? WHERE id=?",
        (login_ip, session_token, user['id'])
    )
    db.commit()
    return jsonify({
        'user': {
            'id': user['id'],
            'username': user['username'],
            'full_name': user['full_name'],
            'role': user['role'],
            'role_label': ROLE_LABELS.get(user['role'], user['role'])
        }
    })

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'message': 'Sesion cerrada'})

@app.route('/api/auth/me', methods=['GET'])
@login_required
def get_me():
    db = get_db()
    # Get permissions from role_permissions table
    permissions = []
    perms_configured = True
    try:
        role = g.user['role']
        if role == 'admin':
            # Admin always has all permissions
            permissions = ['dashboard', 'contacts', 'groups', 'templates', 'send', 'records', 'calls', 'content-search', 'users', 'my-account', 'my-team', 'all-teams', 'config', 'voice-config', 'retention', 'role-permissions']
            perms_configured = True
        else:
            # Get permissions from role_permissions table
            row = db.execute("SELECT permissions FROM role_permissions WHERE role = ?", (role,)).fetchone()
            raw_perms = None
            has_explicit_config = False
            if row:
                perms_raw = row['permissions']
                if isinstance(perms_raw, (list, dict)):
                    # PostgreSQL JSON/JSONB column already parsed by psycopg2
                    raw_perms = perms_raw if isinstance(perms_raw, list) else []
                    has_explicit_config = True
                else:
                    perms_raw = perms_raw or ''
                    import json as _json
                    try:
                        raw_perms = _json.loads(perms_raw) if perms_raw else []
                        # An empty string / no JSON means the role was never
                        # configured; treat that as "use defaults". A literal
                        # "[]" means an admin explicitly removed all pages.
                        has_explicit_config = bool(perms_raw)
                    except Exception:
                        raw_perms = []
            # Fall back to role defaults only when the row is missing or the
            # stored value is empty (never configured). This prevents a blank
            # sidebar for team_admin/team_member on fresh databases while still
            # respecting an admin who explicitly sets permissions to [].
            if has_explicit_config:
                permissions = raw_perms or []
                perms_configured = True
            else:
                permissions = list(DEFAULT_ROLE_PERMISSIONS.get(role, []))
                perms_configured = False
    except Exception as e:
        import traceback
        app.logger.error(f"Error loading permissions for role {g.user.get('role')}: {e}\n{traceback.format_exc()}")
        permissions = []
        perms_configured = False

    # Get team country code
    team_country = ''
    try:
        if g.user['role'] == 'admin':
            # Admin: get first active config country
            cfg = db.execute("SELECT country FROM sms_api_configs WHERE is_active = 1 LIMIT 1").fetchone()
            if cfg:
                team_country = cfg['country'] or ''
        else:
            team_admin_id = g.user.get('team_creator_id')
            if team_admin_id:
                tc = db.execute("SELECT api_config_id FROM team_config WHERE team_admin_id = ?", (team_admin_id,)).fetchone()
                if tc and tc['api_config_id']:
                    cfg = db.execute("SELECT country FROM sms_api_configs WHERE id = ?", (tc['api_config_id'],)).fetchone()
                    if cfg:
                        team_country = cfg['country'] or ''
    except Exception as e:
        app.logger.error(f"Error loading team country: {e}")

    country_codes = {'MX': '+52', 'CO': '+57', 'US': '+1', 'BR': '+55'}
    team_country_code = country_codes.get(team_country, '+52')

    # Categoria del empleado (define retencion de contactos)
    category_id = None
    category_name = None
    category_retention_days = 0
    try:
        if 'category_id' in g.user.keys():
            category_id = g.user['category_id']
        if category_id:
            crow = db.execute("SELECT name, retention_days FROM user_categories WHERE id=?", (category_id,)).fetchone()
            if crow:
                category_name = crow['name'] if isinstance(crow, dict) else crow[0]
                rd = crow['retention_days'] if isinstance(crow, dict) else crow[1]
                category_retention_days = int(rd or 0)
    except Exception:
        pass

    return jsonify({
        'user': {
            'id': g.user['id'],
            'username': g.user['username'],
            'full_name': g.user['full_name'],
            'role': g.user['role'],
            'role_label': ROLE_LABELS.get(g.user['role'], g.user['role']),
            'permissions': permissions,
            'permsConfigured': perms_configured,
            'extnumber': g.user['extnumber'] if 'extnumber' in g.user.keys() else None,
            'country': normalize_country(g.user['country']) if 'country' in g.user.keys() else '',
            'country_label': COUNTRY_LABELS.get(normalize_country(g.user['country'])) if 'country' in g.user.keys() else '',
            'category_id': category_id,
            'category_name': category_name,
            'category_retention_days': category_retention_days,
            'team_country_code': team_country_code,
            'environment': app.config['APP_ENVIRONMENT'],
            'app_label': app.config['APP_LABEL'],
        }
    })

# ============================================================
# User Categories (clasificacion de empleados + retencion de contactos)
# ============================================================

def get_user_categories(only_active_counts=False):
    """Return all user categories ordered. retention_days=0 means 'forever'."""
    db = get_db()
    rows = db.execute(
        "SELECT id, name, retention_days, is_default, created_at, updated_at "
        "FROM user_categories ORDER BY is_default DESC, name ASC"
    ).fetchall()
    cats = []
    for r in rows:
        cd = dict(r)
        cd['is_default'] = bool(cd.get('is_default'))
        cats.append(cd)
    if only_active_counts:
        for c in cats:
            uc = db.execute("SELECT COUNT(*) AS c FROM users WHERE category_id=?", (c['id'],)).fetchone()
            c['user_count'] = int((uc['c'] if isinstance(uc, dict) else uc[0]) if uc else 0)
    return cats


def get_default_category_id():
    db = get_db()
    row = db.execute("SELECT id FROM user_categories WHERE is_default=TRUE LIMIT 1").fetchone()
    if row:
        return row['id'] if isinstance(row, dict) else row[0]
    row = db.execute("SELECT id FROM user_categories ORDER BY id ASC LIMIT 1").fetchone()
    return (row['id'] if isinstance(row, dict) else row[0]) if row else None


def resolve_category_id(raw):
    """Validate and normalize a category id from request input. Returns int or None."""
    if raw is None or raw == '' or str(raw).strip() == '':
        return None
    try:
        cid = int(raw)
    except (TypeError, ValueError):
        return None
    db = get_db()
    row = db.execute("SELECT id FROM user_categories WHERE id=?", (cid,)).fetchone()
    return cid if row else None


@app.route('/api/user-categories', methods=['GET'])
@login_required
def list_user_categories():
    # Managers configure retention; members only need the list for user forms.
    # Include user counts for any manager.
    counts = g.user['role'] in ('admin', 'team_admin')
    return jsonify({'categories': get_user_categories(only_active_counts=counts)})


@app.route('/api/user-categories', methods=['POST'])
@login_required
@manager_required
def create_user_category():
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'El nombre es obligatorio'}), 400
    if len(name) > 100:
        return jsonify({'error': 'El nombre es demasiado largo'}), 400
    try:
        retention_days = int(data.get('retention_days', 7))
    except (TypeError, ValueError):
        return jsonify({'error': 'Dias de retencion invalidos'}), 400
    if retention_days < 1 or retention_days > 7:
        return jsonify({'error': 'Los dias de retencion deben estar entre 1 y 7'}), 400
    db = get_db()
    existing = db.execute("SELECT id FROM user_categories WHERE name=?", (name,)).fetchone()
    if existing:
        return jsonify({'error': 'Ya existe una categoria con ese nombre'}), 409
    cur = db.execute(
        "INSERT INTO user_categories (name, retention_days, is_default) VALUES (?, ?, ?)",
        (name, retention_days, False)
    )
    db.commit()
    return jsonify({'message': 'Categoria creada', 'id': cur.lastrowid}), 201


@app.route('/api/user-categories/<int:cat_id>', methods=['PUT'])
@login_required
@manager_required
def update_user_category(cat_id):
    data = request.get_json() or {}
    db = get_db()
    cat = db.execute("SELECT * FROM user_categories WHERE id=?", (cat_id,)).fetchone()
    if not cat:
        return jsonify({'error': 'Categoria no encontrada'}), 404
    name = (data.get('name') or '').strip()
    if 'retention_days' in data:
        try:
            retention_days = int(data.get('retention_days'))
        except (TypeError, ValueError):
            return jsonify({'error': 'Dias de retencion invalidos'}), 400
        if retention_days < 1 or retention_days > 7:
            return jsonify({'error': 'Los dias de retencion deben estar entre 1 y 7'}), 400
    else:
        retention_days = cat['retention_days']
    if not name:
        name = cat['name']
    if len(name) > 100:
        return jsonify({'error': 'El nombre es demasiado largo'}), 400
    dup = db.execute("SELECT id FROM user_categories WHERE name=? AND id<>?", (name, cat_id)).fetchone()
    if dup:
        return jsonify({'error': 'Ya existe una categoria con ese nombre'}), 409
    if get_db_type() == 'postgres':
        db.execute(
            "UPDATE user_categories SET name=%s, retention_days=%s, updated_at=NOW() WHERE id=%s",
            (name, retention_days, cat_id)
        )
    else:
        db.execute(
            "UPDATE user_categories SET name=?, retention_days=?, updated_at=datetime('now') WHERE id=?",
            (name, retention_days, cat_id)
        )
    db.commit()
    return jsonify({'message': 'Categoria actualizada'})


@app.route('/api/user-categories/<int:cat_id>', methods=['DELETE'])
@login_required
@manager_required
def delete_user_category(cat_id):
    db = get_db()
    cat = db.execute("SELECT * FROM user_categories WHERE id=?", (cat_id,)).fetchone()
    if not cat:
        return jsonify({'error': 'Categoria no encontrada'}), 404
    is_default = bool(dict(cat).get('is_default'))
    if is_default:
        return jsonify({'error': 'No se puede eliminar la categoria por defecto'}), 400
    in_use = db.execute("SELECT COUNT(*) AS c FROM users WHERE category_id=?", (cat_id,)).fetchone()
    used = int((in_use['c'] if isinstance(in_use, dict) else in_use[0]) if in_use else 0)
    if used > 0:
        return jsonify({'error': f'No se puede eliminar: {used} usuario(s) pertenecen a esta categoria. Reasignelos primero.'}), 409
    db.execute("DELETE FROM user_categories WHERE id=?", (cat_id,))
    db.commit()
    return jsonify({'message': 'Categoria eliminada'})


# ============================================================
# Users API (role-based access control)
# ============================================================

@app.route('/api/users', methods=['GET'])
@manager_required
def list_users():
    db = get_db()
    current = g.user
    role_filter = request.args.get('role', '').strip()
    search = request.args.get('search', '').strip()

    # Base query with team creator name and category
    base_select = """
        SELECT u.id, u.username, u.full_name, u.role, u.team_creator_id,
               u.is_active, u.created_at, u.updated_at,
               u.last_login_ip, u.last_login_at, u.extnumber, u.country,
               u.category_id,
               c.name AS category_name, c.retention_days AS category_retention_days,
               tc.username AS team_creator_name, tc.full_name AS team_creator_fullname
        FROM users u
        LEFT JOIN users tc ON u.team_creator_id = tc.id
        LEFT JOIN user_categories c ON u.category_id = c.id
    """

    if current['role'] == 'admin':
        where_clauses = ["u.id != ?"]
        params = [current['id']]
        if role_filter:
            where_clauses.append("u.role = ?")
            params.append(role_filter)
        if search:
            where_clauses.append("(u.username LIKE ? OR u.full_name LIKE ?)")
            params.extend([f'%{search}%', f'%{search}%'])
        where_sql = " WHERE " + " AND ".join(where_clauses) if where_clauses else ""
        users = db.execute(base_select + where_sql + " ORDER BY u.created_at DESC", params).fetchall()
    else:
        # Team admin sees all team members (team_creator_id = current['id'] OR same team)
        where_clauses = ["(u.team_creator_id = ? OR (u.team_creator_id IN (SELECT team_creator_id FROM users WHERE id = ?) AND u.team_creator_id IS NOT NULL))"]
        params = [current['id'], current['id']]
        if role_filter:
            where_clauses.append("u.role = ?")
            params.append(role_filter)
        if search:
            where_clauses.append("(u.username LIKE ? OR u.full_name LIKE ?)")
            params.extend([f'%{search}%', f'%{search}%'])
        where_sql = " AND ".join(where_clauses)
        users = db.execute(base_select + " WHERE " + where_sql + " ORDER BY u.created_at DESC", params).fetchall()

    result = []
    for u in users:
        ud = dict(u)
        ud['role_label'] = ROLE_LABELS.get(ud['role'], ud['role'])
        ud['country'] = normalize_country(ud.get('country'))
        ud['country_label'] = COUNTRY_LABELS.get(ud['country'], '') if ud['country'] else ''
        # Parse permissions
        import json as _json
        perms_raw = ud.get('permissions', '') or ''
        try:
            ud['permissions'] = _json.loads(perms_raw) if perms_raw else []
        except Exception:
            ud['permissions'] = []
        # Admin always has all permissions
        if ud['role'] == 'admin':
            ud['permissions'] = ['dashboard', 'contacts', 'groups', 'templates', 'send', 'records', 'calls', 'content-search', 'users', 'my-account', 'my-team', 'all-teams', 'config', 'voice-config']
        # Team affiliation: show team creator name for team members
        if ud['team_creator_name']:
            ud['team_affiliation'] = ud['team_creator_fullname'] or ud['team_creator_name']
        else:
            ud['team_affiliation'] = None
        result.append(ud)
    return jsonify({'users': result})

@app.route('/api/users', methods=['POST'])
@manager_required
def create_user():
    data = request.get_json()
    current = g.user
    username = data.get('username', '').strip()
    password = (data.get('password') or '').strip()
    full_name = (data.get('full_name') or '').strip()
    role = data.get('role', 'team_member')
    if not username:
        return jsonify({'error': 'El usuario es requerido'}), 400
    generated_password = None
    if not password:
        # Auto-generate a 10-character alphanumeric password if not provided
        password = generate_password(10)
        generated_password = password
    if len(password) < 6:
        return jsonify({'error': 'La contrasena debe tener minimo 6 caracteres'}), 400
    # Role assignment rules
    if current['role'] == 'admin':
        # System admin can only create team_admin
        if role != 'team_admin':
            return jsonify({'error': 'El Administrador del Sistema solo puede crear Administradores de Equipo'}), 400
        team_creator_id = None
    elif current['role'] == 'team_admin':
        # Team admin can only create team_member
        if role != 'team_member':
            return jsonify({'error': 'Solo puede crear Miembros de Equipo'}), 400
        team_creator_id = current['id']
    else:
        return jsonify({'error': 'Permisos insuficientes'}), 403
    db = get_db()
    existing = db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
    if existing:
        return jsonify({'error': 'El nombre de usuario ya existe'}), 409
    api_config_id = data.get('api_config_id')
    # Las extensiones NO se asignan manualmente: se elige automaticamente una
    # libre del pool configurado si el administrador marca "asignar telefono".
    assign_extension = bool(data.get('assign_extension', False))
    extnumber = None
    country = normalize_country(data.get('country'))
    # Categoria del empleado (define los dias de retencion de sus contactos).
    # Si no se especifica, se asigna la categoria por defecto.
    category_id = resolve_category_id(data.get('category_id'))
    if category_id is None and not (data.get('category_id') is not None and str(data.get('category_id')).strip() != ''):
        category_id = get_default_category_id()
    if assign_extension:
        extnumber = allocate_extension(country=country)
        if not extnumber:
            label = {'mx': 'Mexico', 'co': 'Colombia', 'pe': 'Peru'}.get(country, 'el pool')
            return jsonify({'error': f'No hay extensiones disponibles para {label}. Pida al administrador del sistema que agregue mas extensiones en Configuracion de Voz.'}), 409
    cur = db.execute(
        "INSERT INTO users (username, password_hash, full_name, role, team_creator_id, extnumber, country, category_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (username, hash_password(password), full_name, role, team_creator_id, extnumber, country or None, category_id)
    )
    new_user_id = cur.lastrowid
    EXPORTABLE_PASSWORDS[int(new_user_id)] = password
    if extnumber:
        _extensions_mark_assigned(extnumber, new_user_id, country)
    # If creating a team_admin, create team_config with api_config_id
    if role == 'team_admin':
        db.execute(
            "INSERT INTO team_config (team_admin_id, api_config_id, daily_sms_limit) VALUES (?, ?, 100)",
            (new_user_id, api_config_id if api_config_id else None)
        )
    db.commit()
    response = {'message': 'Usuario creado exitosamente'}
    if generated_password:
        response['generated_password'] = generated_password
    return jsonify(response), 201

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@manager_required
def update_user(user_id):
    data = request.get_json()
    current = g.user
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        return jsonify({'error': 'Usuario no encontrado'}), 404
    # Permission checks
    if current['role'] == 'admin':
        # System admin can edit anyone except changing their own role
        pass
    elif current['role'] == 'team_admin':
        # Team admin can only edit their own team members
        if user['team_creator_id'] != current['id']:
            return jsonify({'error': 'Solo puede editar miembros de su propio equipo'}), 403
        # Cannot change role
        if 'role' in data and data['role'] != user['role']:
            return jsonify({'error': 'No puede cambiar el rol de un usuario'}), 403
    else:
        return jsonify({'error': 'Permisos insuficientes'}), 403
    full_name = data.get('full_name', user['full_name'])
    is_active = data.get('is_active', user['is_active'])
    password = data.get('password', None)
    updates = ["updated_at=datetime('now')"]
    params = []
    if full_name is not None:
        updates.append("full_name=?")
        params.append(full_name)
    # Only system admin can change roles
    if current['role'] == 'admin' and 'role' in data:
        new_role = data['role']
        if new_role in ('admin', 'team_admin', 'team_member'):
            updates.append("role=?")
            params.append(new_role)
    if is_active in (0, 1, True, False):
        updates.append("is_active=?")
        params.append(int(is_active))
    # Pais del agente (mx/co/pe). Cambiar de pais no reasigna la extension:
    # si tenia una y el pais cambia, el admin debe liberarla y reasignar.
    if 'country' in data:
        new_country = normalize_country(data.get('country'))
        updates.append("country=?")
        params.append(new_country or None)
    # Categoria del empleado (retencion de contactos). Se permite asignar o
    # desasignar (category_id=null -> usa la default en la practica).
    if 'category_id' in data:
        new_category = resolve_category_id(data.get('category_id'))
        updates.append("category_id=?")
        params.append(new_category)
    if password:
        if len(password) < 6:
            return jsonify({'error': 'La contrasena debe tener minimo 6 caracteres'}), 400
        updates.append("password_hash=?")
        params.append(hash_password(password))
        updates.append("session_token=?")
        params.append(None)
    # Gestion de extension: no se permite escribir el numero manualmente.
    # - assign_extension=true  -> asignar automaticamente una libre del catalogo
    # - release_extension=true -> liberar la extension actual
    # El valor manual de 'extnumber' se ignora deliberadamente.
    wants_assign = bool(data.get('assign_extension', False))
    wants_release = bool(data.get('release_extension', False))
    released_ext = None
    new_ext = None
    new_ext_country = None
    if wants_release:
        released_ext = _normalize_extnumber(user['extnumber'] if 'extnumber' in user.keys() else None)
        updates.append("extnumber=?")
        params.append(None)
    elif wants_assign:
        # Si ya tenia una extension, se conserva; si no, se asigna una libre.
        had_ext = _normalize_extnumber(user['extnumber'] if 'extnumber' in user.keys() else None)
        if not had_ext:
            target_country = normalize_country(data.get('country')) if 'country' in data else (
                normalize_country(user['country'] if 'country' in user.keys() else None))
            new_ext = allocate_extension(exclude_id=user_id, country=target_country)
            if not new_ext:
                label = {'mx': 'Mexico', 'co': 'Colombia', 'pe': 'Peru'}.get(target_country, 'el pool')
                return jsonify({'error': f'No hay extensiones disponibles para {label}. Pida al administrador del sistema que agregue mas extensiones en la pagina de Extensiones.'}), 409
            new_ext_country = target_country
            updates.append("extnumber=?")
            params.append(new_ext)
    params.append(user_id)
    db.execute(f"UPDATE users SET {', '.join(updates)} WHERE id=?", params)
    db.commit()
    if released_ext:
        _extensions_release(released_ext)
    if new_ext:
        _extensions_mark_assigned(new_ext, user_id, new_ext_country)
    return jsonify({'message': 'Usuario actualizado'})

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@manager_required
def delete_user(user_id):
    current = g.user
    db = get_db()
    if user_id == current['id']:
        return jsonify({'error': 'No puede eliminar su propia cuenta'}), 400
    user = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        return jsonify({'error': 'Usuario no encontrado'}), 404
    # Permission checks
    if current['role'] == 'admin':
        # System admin can delete team_admins (but not other system admins)
        if user['role'] == 'admin':
            return jsonify({'error': 'No puede eliminar a otro Administrador del Sistema'}), 403
    elif current['role'] == 'team_admin':
        # Team admin can only delete their own team members
        if user['team_creator_id'] != current['id']:
            return jsonify({'error': 'Solo puede eliminar miembros de su propio equipo'}), 403
        if user['role'] != 'team_member':
            return jsonify({'error': 'Solo puede eliminar Miembros de Equipo'}), 403
    else:
        return jsonify({'error': 'Permisos insuficientes'}), 403
    # Protection: cannot delete a team_admin who still has members under them
    if user['role'] == 'team_admin':
        member_count = db.execute(
            "SELECT COUNT(*) as cnt FROM users WHERE team_creator_id=?",
            (user_id,)
        ).fetchone()
        if member_count and member_count['cnt'] > 0:
            return jsonify({
                'error': f"Este Administrador de Equipo tiene {member_count['cnt']} miembro(s) a cargo. Elimine o reasigne los miembros primero."
            }), 400
    deleted_ext = _normalize_extnumber(user['extnumber'] if 'extnumber' in user.keys() else None)
    db.execute("DELETE FROM users WHERE id=?", (user_id,))
    db.commit()
    if deleted_ext:
        _extensions_release(deleted_ext)
    return jsonify({'message': 'Usuario eliminado'})

def _bulk_create_users_core(current_user, users, default_api_config_id, default_password=None, assign_extensions=False, default_country=None, default_category_id=None):
    """Core bulk-creation logic shared by JSON and Excel upload.

    Returns (created, errors). Each created item includes the plain-text
    password used (for the password-list export); existing/hashed passwords
    are never returned elsewhere.

    When assign_extensions is True, each created user gets a free extension
    auto-allocated from the pool of their country (u.country mx/co/pe, or the
    global fallback pool). No manual assignment is allowed.
    default_country (mx/co/pe/'') applies to rows that omit their own country.
    """
    db = get_db()
    default_country = normalize_country(default_country)
    # Categoria por defecto de la tanda; si no es valida, cae a la global.
    bulk_category_id = resolve_category_id(default_category_id)
    if bulk_category_id is None:
        bulk_category_id = get_default_category_id()

    if current_user['role'] == 'admin':
        role = 'team_admin'
        team_creator_id = None
    elif current_user['role'] == 'team_admin':
        role = 'team_member'
        team_creator_id = current_user['id']
    else:
        return None, None, None

    created = []
    errors = []
    existing_users = set(r['username'].lower() for r in db.execute("SELECT username FROM users").fetchall())
    _extensions_seed_from_config()
    use_catalog = _extensions_table_ready()
    # Extensions already taken (active users / catalog assignments).
    taken_extensions = set()
    if use_catalog:
        for r in db.execute(
            "SELECT LOWER(TRIM(extnumber)) AS e FROM extensions WHERE assigned_to IS NOT NULL"
        ).fetchall():
            taken_extensions.add((r['e'] if not isinstance(r, tuple) else r[0]) or '')
    for r in db.execute(
        "SELECT extnumber FROM users WHERE is_active = 1 AND extnumber IS NOT NULL AND TRIM(extnumber) != ''"
    ).fetchall():
        taken_extensions.add(_normalize_extension(r['extnumber'] if not isinstance(r, tuple) else r[0]).lower())
    # Build per-country free pools (catalog first, legacy pool string fallback).
    pools_by_country = {}
    for _cc in ('mx', 'co', 'pe', ''):
        if use_catalog:
            rows = db.execute("SELECT extnumber FROM extensions WHERE country=?", (_cc,)).fetchall()
            all_pool = [_normalize_extension(r['extnumber'] if not isinstance(r, tuple) else r[0]) for r in rows]
        else:
            all_pool = _get_extension_pool(_cc or None)
        pools_by_country[_cc] = {
            'all': all_pool,
            'free': [e for e in all_pool if _normalize_extnumber(e).lower() not in taken_extensions]
        }
    if assign_extensions:
        if not any(p['all'] for p in pools_by_country.values()):
            preflight_error = 'No hay extensiones configuradas. Pida al administrador del sistema que agregue extensiones en la pagina de Extensiones antes de asignar.'
            return None, [{'index': -1, 'username': '', 'error': preflight_error}], role
    seen_usernames = set()

    for idx, u in enumerate(users):
        if not isinstance(u, dict):
            errors.append({'index': idx, 'username': '', 'error': 'Formato invalido'})
            continue
        username = (u.get('username') or '').strip()
        full_name = (u.get('full_name') or '').strip()
        country = normalize_country(u.get('country')) or default_country
        api_config_id = u.get('api_config_id') or default_api_config_id
        # Las extensiones nunca se leen del archivo: se asignan automaticamente.
        extnumber = None

        if not username:
            errors.append({'index': idx, 'username': username, 'error': 'Usuario requerido'})
            continue

        # Solo el usuario es obligatorio. Si no hay contrasena (ni por fila ni por defecto),
        # se genera una combinacion irregular de 10 caracteres alfanumericos.
        password = (u.get('password') or '').strip() or (default_password or '').strip()
        generated_password = False
        if not password:
            password = generate_password(10)
            generated_password = True
        elif len(password) < 6:
            errors.append({'index': idx, 'username': username, 'error': 'Contrasena minimo 6 caracteres'})
            continue

        uname_lower = username.lower()
        if uname_lower in existing_users or uname_lower in seen_usernames:
            errors.append({'index': idx, 'username': username, 'error': 'El nombre de usuario ya existe'})
            continue

        if assign_extensions:
            pool_info = pools_by_country.get(country) or pools_by_country['']
            free_pool = pool_info['free']
            if not free_pool:
                label = COUNTRY_LABELS.get(country, 'el pool')
                errors.append({'index': idx, 'username': username, 'error': f'No quedan extensiones libres para {label}. Pida al administrador del sistema que agregue mas extensiones.'})
                continue
            import random as _random
            chosen = _random.choice(free_pool)
            free_pool.remove(chosen)
            extnumber = chosen
            taken_extensions.add(_normalize_extnumber(chosen).lower())

        try:
            # Categoria: por fila o la default de la creacion masiva; si no, la default global.
            row_cat = resolve_category_id(u.get('category_id'))
            if row_cat is None:
                row_cat = bulk_category_id
            db.execute(
                "INSERT INTO users (username, password_hash, full_name, role, team_creator_id, extnumber, country, category_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (username, hash_password(password), full_name, role, team_creator_id, extnumber, country or None, row_cat)
            )
            new_user_row = db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
            new_user_id = new_user_row['id']
            if role == 'team_admin':
                if not api_config_id:
                    errors.append({'index': idx, 'username': username, 'error': 'Configuracion API (pais) requerida'})
                    db.execute("DELETE FROM users WHERE id=?", (new_user_id,))
                    continue
                db.execute(
                    "INSERT INTO team_config (team_admin_id, api_config_id, daily_sms_limit) VALUES (?, ?, 100)",
                    (new_user_id, api_config_id)
                )
            EXPORTABLE_PASSWORDS[int(new_user_id)] = password
            if extnumber:
                _extensions_mark_assigned(extnumber, new_user_id, country)
            created.append({
                'id': new_user_id,
                'username': username,
                'full_name': full_name,
                'role': role,
                'country': country,
                'country_label': COUNTRY_LABELS.get(country, ''),
                'extnumber': extnumber or '',
                'password': password,
                'generated': generated_password
            })
            seen_usernames.add(uname_lower)
        except Exception as e:
            errors.append({'index': idx, 'username': username, 'error': 'Error al crear: ' + str(e)})

    db.commit()
    return created, errors, role


@app.route('/api/users/bulk', methods=['POST'])
@manager_required
def bulk_create_users():
    """Bulk create users.
    Admin -> creates team_admins (each needs api_config_id).
    Team_admin -> creates team_members under their management.
    Payload: { "users": [ {"username":..,"password":..,"full_name":..,"api_config_id":..}, ... ], "api_config_id": <optional default>, "default_password": <optional> }
    """
    data = request.get_json() or {}
    users = data.get('users', [])
    if not isinstance(users, list) or not users:
        return jsonify({'error': 'Debe proporcionar una lista de usuarios'}), 400
    if len(users) > 500:
        return jsonify({'error': 'No se pueden crear mas de 500 usuarios a la vez'}), 400

    default_api_config_id = data.get('api_config_id')
    default_password = data.get('default_password')
    assign_extensions = bool(data.get('assign_extensions', False))
    default_country = data.get('country')
    default_category_id = data.get('category_id')
    result = _bulk_create_users_core(g.user, users, default_api_config_id, default_password, assign_extensions=assign_extensions, default_country=default_country, default_category_id=default_category_id)
    if result[0] is None:
        # Special case: pre-validation error (e.g. empty pool when assigning)
        if len(result) > 2 and isinstance(result[1], list) and result[1] and result[1][0].get('index') == -1:
            return jsonify({'error': result[1][0]['error']}), 409
        return jsonify({'error': 'Permisos insuficientes'}), 403
    created, errors, role = result
    return jsonify({
        'message': 'Creacion masiva completada',
        'created_count': len(created),
        'error_count': len(errors),
        'created': created,
        'errors': errors
    }), 200


def _parse_excel_users(file_storage):
    """Parse an uploaded Excel file into a list of user dicts.

    Accepts columns (case/space/accent insensitive, first row as header):
    usuario/username, contrasena/password, nombre/full_name/nombre completo.
    Also accepts a 3-column sheet without header (username, password, full_name).

    NOTE: extension/telefono columns are intentionally ignored. Extensions are
    never taken from the file; they are auto-allocated by the system when the
    caller requests assign_extensions.
    """
    from openpyxl import load_workbook
    wb = load_workbook(filename=file_storage.stream, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    def normalize(s):
        if s is None:
            return ''
        return str(s).strip().lower().replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ü', 'u')

    header_map = {
        'usuario': 'username', 'username': 'username', 'user': 'username', 'cuenta': 'username',
        'contrasena': 'password', 'password': 'password', 'clave': 'password', 'pass': 'password',
        'nombre': 'full_name', 'nombre completo': 'full_name', 'nombre_completo': 'full_name', 'fullname': 'full_name', 'name': 'full_name', 'full name': 'full_name',
        'pais': 'country', 'país': 'country', 'country': 'country',
    }

    def parse_country(val):
        s = normalize(val)
        if s in ('mx', 'mexico', 'méxico', '52', '+52'):
            return 'mx'
        if s in ('co', 'colombia', '57', '+57'):
            return 'co'
        if s in ('pe', 'peru', 'perú', '51', '+51'):
            return 'pe'
        return ''

    first = rows[0]
    first_norm = [normalize(c) for c in first]
    has_header = any(h in header_map for h in first_norm)

    users = []
    if has_header:
        col_index = {}
        for ci, h in enumerate(first_norm):
            if h in header_map:
                col_index[header_map[h]] = ci
        for row in rows[1:]:
            if row is None or all(c is None or str(c).strip() == '' for c in row):
                continue
            def get(key):
                ci = col_index.get(key)
                if ci is None or ci >= len(row):
                    return ''
                return '' if row[ci] is None else str(row[ci])
            users.append({
                'username': get('username'),
                'password': get('password'),
                'full_name': get('full_name'),
                'country': parse_country(get('country')),
            })
    else:
        # No header: treat columns as username, password, full_name, country in order.
        # A trailing numeric column (legacy) is ignored (extensions are auto-allocated).
        for row in rows:
            if row is None or all(c is None or str(c).strip() == '' for c in row):
                continue
            users.append({
                'username': '' if row[0] is None else str(row[0]),
                'password': '' if len(row) < 2 or row[1] is None else str(row[1]),
                'full_name': '' if len(row) < 3 or row[2] is None else str(row[2]),
                'country': parse_country('' if len(row) < 4 or row[3] is None else row[3]),
            })
    wb.close()
    return users


@app.route('/api/users/bulk-import', methods=['POST'])
@manager_required
def bulk_import_users():
    """Bulk create users from an uploaded Excel file (.xlsx/.xls).

    Form fields: api_config_id (required for admin), default_password (optional).
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No se ha enviado ningun archivo'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Archivo invalido'}), 400
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'El archivo debe ser Excel (.xlsx o .xls)'}), 400

    try:
        users = _parse_excel_users(file)
    except Exception as e:
        return jsonify({'error': 'No se pudo leer el archivo Excel: ' + str(e)}), 400

    if not users:
        return jsonify({'error': 'El archivo no contiene usuarios validos'}), 400
    if len(users) > 500:
        return jsonify({'error': 'No se pueden crear mas de 500 usuarios a la vez'}), 400

    default_api_config_id = request.form.get('api_config_id')
    if default_api_config_id:
        try:
            default_api_config_id = int(default_api_config_id)
        except (TypeError, ValueError):
            default_api_config_id = None
    default_password = request.form.get('default_password')
    assign_extensions = request.form.get('assign_extensions', 'false').lower() in ('1', 'true', 'on', 'yes')
    default_country = request.form.get('country')
    default_category_id = request.form.get('category_id')

    result = _bulk_create_users_core(g.user, users, default_api_config_id, default_password, assign_extensions=assign_extensions, default_country=default_country, default_category_id=default_category_id)
    if result[0] is None:
        if result[1]:
            return jsonify({'error': result[1][0].get('error', 'Permisos insuficientes')}), 409
        return jsonify({'error': 'Permisos insuficientes'}), 403
    created, errors, role = result
    return jsonify({
        'message': 'Importacion completada',
        'created_count': len(created),
        'error_count': len(errors),
        'created': created,
        'errors': errors
    }), 200


@app.route('/api/users/bulk-password', methods=['POST'])
@manager_required
def bulk_update_passwords():
    """Bulk reset passwords for users, generating a random password per user.

    Admin can reset team_admins/team_members (not other admins).
    Team_admin can reset only their own team_members.
    Payload: { "ids": [1, 2, 3] } (each gets a unique random 10-char password).
    Cannot change own password here.
    """
    data = request.get_json() or {}
    ids = data.get('ids')
    # Backwards compatibility: accept { "items": [{"id":1}, ...] } as well
    if ids is None:
        items = data.get('items', [])
        if isinstance(items, list):
            ids = [it.get('id') for it in items if isinstance(it, dict)]
    if not isinstance(ids, list) or not ids:
        return jsonify({'error': 'Debe proporcionar una lista de usuarios'}), 400
    if len(ids) > 500:
        return jsonify({'error': 'No se pueden actualizar mas de 500 usuarios a la vez'}), 400

    db = get_db()
    current = g.user
    updated = []
    errors = []

    for idx, uid in enumerate(ids):
        if not uid:
            errors.append({'index': idx, 'error': 'ID de usuario requerido'})
            continue
        new_password = generate_password(10)

        row = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
        if not row:
            errors.append({'index': idx, 'id': uid, 'error': 'Usuario no encontrado'})
            continue
        if int(uid) == int(current['id']):
            errors.append({'index': idx, 'id': uid, 'username': row['username'], 'error': 'Use "Mi Cuenta" para cambiar su propia contrasena'})
            continue
        if current['role'] == 'team_admin':
            if row['role'] != 'team_member' or (row['team_creator_id'] is not None and int(row['team_creator_id']) != int(current['id'])):
                errors.append({'index': idx, 'id': uid, 'username': row['username'], 'error': 'Solo puede restablecer miembros de su equipo'})
                continue
        elif current['role'] == 'admin':
            if row['role'] == 'admin':
                errors.append({'index': idx, 'id': uid, 'username': row['username'], 'error': 'No puede modificar a otro administrador del sistema'})
                continue

        try:
            db.execute("UPDATE users SET password_hash=?, session_token=NULL WHERE id=?", (hash_password(new_password), uid))
            updated.append({'id': uid, 'username': row['username'], 'password': new_password})
            EXPORTABLE_PASSWORDS[int(uid)] = new_password
        except Exception as e:
            errors.append({'index': idx, 'id': uid, 'username': row['username'], 'error': 'Error al actualizar: ' + str(e)})

    db.commit()
    return jsonify({
        'message': 'Contrasenas actualizadas',
        'updated_count': len(updated),
        'error_count': len(errors),
        'updated': updated,
        'errors': errors
    }), 200


def _users_for_export(current_user):
    """Return the list of users visible to the current user, with joined fields."""
    db = get_db()
    if current_user['role'] == 'admin':
        rows = db.execute("""
            SELECT u.id, u.username, u.full_name, u.role, u.is_active, u.created_at,
                   u.team_creator_id, u.last_login_ip, u.last_login_at, u.extnumber, u.country,
                   c.username as creator_username,
                   tc.api_config_id, sac.name as config_name, sac.country as config_country
            FROM users u
            LEFT JOIN users c ON u.team_creator_id = c.id
            LEFT JOIN team_config tc ON u.id = tc.team_admin_id
            LEFT JOIN sms_api_configs sac ON tc.api_config_id = sac.id
            ORDER BY u.id
        """).fetchall()
    elif current_user['role'] == 'team_admin':
        rows = db.execute("""
            SELECT u.id, u.username, u.full_name, u.role, u.is_active, u.created_at,
                   u.team_creator_id, u.last_login_ip, u.last_login_at, u.extnumber
            FROM users u
            WHERE u.id=? OR u.team_creator_id=?
            ORDER BY u.id
        """, (current_user['id'], current_user['id'])).fetchall()
    else:
        rows = db.execute("""
            SELECT u.id, u.username, u.full_name, u.role, u.is_active, u.created_at,
                   u.last_login_ip, u.last_login_at, u.extnumber
            FROM users u WHERE u.id=?
        """, (current_user['id'],)).fetchall()
    return rows


def _build_users_workbook(rows, include_passwords=False, password_map=None, viewer_role='admin'):
    """Build an openpyxl Workbook for user export.

    include_passwords=True adds a 'Contrasena' column using password_map {id: plain},
    intended only for freshly created/reset passwords. Existing accounts show
    'Configurada (no visible)'.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = 'Usuarios'

    role_labels = {
        'admin': 'Administrador del Sistema',
        'team_admin': 'Administrador de Equipo',
        'team_member': 'Miembro de Equipo'
    }
    rows = [dict(r) for r in rows]

    headers = ['ID', 'Usuario', 'Nombre Completo', 'Rol', 'Extension', 'Pais Agente', 'Equipo/Admin', 'Pais/Config API', 'Estado', 'Creado', 'Ultimo Login IP', 'Ultimo Login']
    if include_passwords:
        headers.append('Contrasena')

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    password_map = password_map or {}
    for ri, r in enumerate(rows, start=2):
        creator = r['creator_username'] if viewer_role == 'admin' else None
        config_name = r['config_name'] if viewer_role == 'admin' else None
        config_country = r['config_country'] if viewer_role == 'admin' else None
        team_field = creator if creator else ('Si' if r['role'] == 'team_admin' else '-')
        config_field = ''
        if config_name:
            config_field = f"{config_name} ({config_country})" if config_country else config_name
        elif r['role'] == 'team_admin':
            config_field = '-'
        is_active = r['is_active']
        if hasattr(is_active, 'real'):
            is_active = bool(is_active)
        agent_country = normalize_country(r.get('country'))
        values = [
            r['id'],
            r['username'],
            r['full_name'] or '',
            role_labels.get(r['role'], r['role']),
            r.get('extnumber') or '',
            COUNTRY_LABELS.get(agent_country, '') if agent_country else '',
            team_field,
            config_field,
            'Activo' if is_active else 'Inactivo',
            str(r['created_at']) if r['created_at'] else '',
            r['last_login_ip'] or '',
            str(r['last_login_at']) if r['last_login_at'] else ''
        ]
        if include_passwords:
            values.append(password_map.get(r['id'], 'Configurada (no visible)'))
        for ci, v in enumerate(values, start=1):
            ws.cell(row=ri, column=ci, value=v)

    # Auto-ish column widths
    widths = [6, 18, 24, 26, 14, 14, 18, 20, 10, 22, 18, 22]
    if include_passwords:
        widths.append(18)
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + ci) if ci <= 26 else 'A' + chr(64 + ci - 26)].width = w

    return wb


@app.route('/api/users/template', methods=['GET'])
@login_required
def download_user_template():
    """Download an Excel template for bulk user creation."""
    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.comments import Comment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Usuarios'

    headers = ['usuario', 'contrasena', 'nombre_completo', 'pais']
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font

    # Usage hints as comments on the header cells
    ws.cell(row=1, column=1).comment = Comment(
        'OBLIGATORIO. Nombre de usuario (unico, sin espacios).', 'SMS Platform'
    )
    ws.cell(row=1, column=2).comment = Comment(
        'OPCIONAL. Minimo 6 caracteres. Si se deja vacio, el sistema generara automaticamente una clave de 10 caracteres (letras y numeros).',
        'SMS Platform'
    )
    ws.cell(row=1, column=3).comment = Comment(
        'OPCIONAL. Nombre completo del usuario.', 'SMS Platform'
    )
    ws.cell(row=1, column=4).comment = Comment(
        'OPCIONAL. Pais del agente para asignar su extension: mx (Mexico), co (Colombia) o pe (Peru).',
        'SMS Platform'
    )

    # Example rows
    examples = [
        ['juan.perez', 'Clave1234', 'Juan Perez', 'mx'],
        ['maria.lopez', '', 'Maria Lopez', 'co'],
        ['carlos.ruiz', '', 'Carlos Ruiz', 'pe'],
    ]
    for row in examples:
        ws.append(row)

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 10

    # Note row: extensions are assigned by the system, not from the file.
    note_row = len(examples) + 3
    note_cell = ws.cell(row=note_row, column=1,
                        value='Nota: la columna "pais" define de que pool de extensiones se asigna (mx/co/pe). Las extensiones no se escriben en el archivo; se asignan automaticamente (marque "Asignar extension" al importar).')
    note_cell.font = Font(italic=True, color='B45309')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='plantilla_usuarios.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/users/export', methods=['GET'])
@login_required
def export_users():
    """Export the visible user list as an .xlsx file (no password column).

    Plain-text passwords are only exported immediately after a batch password
    reset via /api/users/export-passwords. The general user list intentionally
    omits passwords to avoid leaking credentials.
    """
    from io import BytesIO
    from flask import send_file
    rows = _users_for_export(g.user)
    wb = _build_users_workbook(
        rows,
        include_passwords=False,
        viewer_role=g.user['role'],
        password_map={}
    )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='usuarios.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/users/export-passwords', methods=['POST'])
@manager_required
def export_passwords():
    """Export an .xlsx that includes plain-text passwords for the given ids.

    Only ids present in `ids` AND freshly known via `passwords` map (created/reset
    in the same operation) get their plain text. Others are marked as not visible.
    Payload: { "ids": [1,2], "passwords": {"1": "plain123"} }
    """
    from io import BytesIO
    from flask import send_file
    data = request.get_json() or {}
    ids = data.get('ids') or []
    passwords = data.get('passwords') or {}
    # Normalize keys to int
    try:
        id_set = set(int(i) for i in ids)
    except (TypeError, ValueError):
        return jsonify({'error': 'Lista de IDs invalida'}), 400
    password_map = {}
    for k, v in passwords.items():
        try:
            password_map[int(k)] = v
        except (TypeError, ValueError):
            continue

    rows = _users_for_export(g.user)
    # If ids provided, restrict to those; otherwise export all visible with known pw
    if id_set:
        rows = [r for r in rows if r['id'] in id_set]
    wb = _build_users_workbook(rows, include_passwords=True, password_map=password_map, viewer_role=g.user['role'])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='contrasenas_usuarios.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/users/bulk-delete', methods=['POST'])
@manager_required
def bulk_delete_users():
    """Bulk delete users. Payload: { "ids": [1,2,3] }
    Enforces the same permission rules as single delete.
    """
    data = request.get_json() or {}
    ids = data.get('ids', [])
    if not isinstance(ids, list) or not ids:
        return jsonify({'error': 'Debe proporcionar una lista de IDs'}), 400
    if len(ids) > 500:
        return jsonify({'error': 'No se pueden eliminar mas de 500 usuarios a la vez'}), 400

    current = g.user
    db = get_db()

    deleted = []
    errors = []
    for uid in ids:
        try:
            uid = int(uid)
        except (TypeError, ValueError):
            errors.append({'id': uid, 'error': 'ID invalido'})
            continue
        if uid == current['id']:
            errors.append({'id': uid, 'error': 'No puede eliminar su propia cuenta'})
            continue
        user = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
        if not user:
            errors.append({'id': uid, 'error': 'Usuario no encontrado'})
            continue
        # Permission checks (same rules as single delete)
        if current['role'] == 'admin':
            if user['role'] == 'admin':
                errors.append({'id': uid, 'username': user['username'], 'error': 'No puede eliminar a otro Administrador del Sistema'})
                continue
        elif current['role'] == 'team_admin':
            if user['team_creator_id'] != current['id']:
                errors.append({'id': uid, 'username': user['username'], 'error': 'No es miembro de su equipo'})
                continue
            if user['role'] != 'team_member':
                errors.append({'id': uid, 'username': user['username'], 'error': 'Solo puede eliminar Miembros de Equipo'})
                continue
        else:
            errors.append({'id': uid, 'error': 'Permisos insuficientes'})
            continue
        # Protection: cannot delete a team_admin who still has members under them
        if user['role'] == 'team_admin':
            member_count = db.execute(
                "SELECT COUNT(*) as cnt FROM users WHERE team_creator_id=?",
                (uid,)
            ).fetchone()
            if member_count and member_count['cnt'] > 0:
                errors.append({
                    'id': uid,
                    'username': user['username'],
                    'error': f"Tiene {member_count['cnt']} miembro(s) a cargo. Elimínelos o reasígnelos primero."
                })
                continue
        db.execute("DELETE FROM users WHERE id=?", (uid,))
        deleted.append({'id': uid, 'username': user['username']})

    db.commit()
    return jsonify({
        'message': 'Eliminacion masiva completada',
        'deleted_count': len(deleted),
        'error_count': len(errors),
        'deleted': deleted,
        'errors': errors
    }), 200

# ============================================================
# Permissions API (admin manages menu access for each user)
# ============================================================

# Available menu pages that can be assigned
AVAILABLE_PAGES = [
    {'id': 'dashboard', 'label': 'Panel Principal', 'icon': 'grid'},
    {'id': 'contacts', 'label': 'Contactos', 'icon': 'users'},
    {'id': 'groups', 'label': 'Grupos', 'icon': 'folder'},
    {'id': 'templates', 'label': 'Plantillas', 'icon': 'file-text'},
    {'id': 'send', 'label': 'Enviar SMS', 'icon': 'send'},
    {'id': 'records', 'label': 'Registros SMS', 'icon': 'activity'},
    {'id': 'calls', 'label': 'Llamadas', 'icon': 'phone'},
    {'id': 'content-search', 'label': 'Buscar Contenido', 'icon': 'search'},
    {'id': 'users', 'label': 'Usuarios', 'icon': 'user-plus'},
    {'id': 'my-account', 'label': 'Mi Cuenta', 'icon': 'user'},
    {'id': 'my-team', 'label': 'Mi Equipo', 'icon': 'users'},
    {'id': 'all-teams', 'label': 'Todos los Equipos', 'icon': 'bar-chart'},
    {'id': 'config', 'label': 'Configuracion API SMS', 'icon': 'settings'},
    {'id': 'voice-config', 'label': 'Configuracion Voz', 'icon': 'settings'},
    {'id': 'retention', 'label': 'Retencion de Contactos', 'icon': 'shield'},
]

# Default permissions per role when the role_permissions table has no explicit
# configuration (empty string or missing row). This avoids blank sidebars on
# fresh installations or after role_permissions were reset. Admins always get
# every page plus role-permissions management, handled explicitly in /api/auth/me.
DEFAULT_ROLE_PERMISSIONS = {
    'admin': [p['id'] for p in AVAILABLE_PAGES] + ['role-permissions'],
    'team_admin': [
        'dashboard', 'contacts', 'groups', 'templates', 'send', 'records',
        'calls', 'content-search', 'users', 'my-account', 'my-team', 'all-teams',
        'retention',
    ],
    'team_member': [
        'dashboard', 'contacts', 'groups', 'templates', 'send', 'records',
        'calls', 'my-account',
    ],
}

@app.route('/api/role-permissions', methods=['GET'])
@admin_required
def get_role_permissions():
    """Get permissions for all roles"""
    import json as _json
    db = get_db()

    # Get all roles and their permissions
    roles = db.execute("""
        SELECT role, permissions FROM role_permissions
        ORDER BY role
    """).fetchall()

    role_perms = []
    for r in roles:
        perms_raw = r['permissions'] or ''
        try:
            perms = _json.loads(perms_raw) if perms_raw else []
        except Exception:
            perms = []
        role_perms.append({
            'role': r['role'],
            'role_label': ROLE_LABELS.get(r['role'], r['role']),
            'permissions': perms
        })

    return jsonify({
        'available_pages': AVAILABLE_PAGES,
        'roles': role_perms
    })

@app.route('/api/role-permissions/<role>', methods=['PUT'])
@admin_required
def update_role_permissions(role):
    """Update permissions for a specific role"""
    import json as _json
    db = get_db()

    data = request.get_json()
    permissions = data.get('permissions', [])

    # Validate permissions
    valid_ids = [p['id'] for p in AVAILABLE_PAGES]
    for p in permissions:
        if p not in valid_ids:
            return jsonify({'error': 'Invalid permission: ' + p}), 400

    # Admin role always has all permissions
    if role == 'admin':
        permissions = [p['id'] for p in AVAILABLE_PAGES] + ['role-permissions']

    perms_json = _json.dumps(permissions)

    # Update or insert
    existing = db.execute("SELECT role FROM role_permissions WHERE role = ?", (role,)).fetchone()
    if existing:
        db.execute("UPDATE role_permissions SET permissions = ? WHERE role = ?", (perms_json, role))
    else:
        db.execute("INSERT INTO role_permissions (role, permissions) VALUES (?, ?)", (role, perms_json))
    db.commit()

    return jsonify({'success': True, 'role': role, 'permissions': permissions})

@app.route('/api/permissions', methods=['GET'])
@admin_required
def get_permissions():
    """Get all available pages and current permissions for all users"""
    import json as _json
    db = get_db()
    users = db.execute("""
        SELECT id, username, full_name, role, permissions
        FROM users
        ORDER BY role, username
    """).fetchall()

    user_perms = []
    for u in users:
        perms_raw = u['permissions'] or ''
        try:
            perms = _json.loads(perms_raw) if perms_raw else []
        except Exception:
            perms = []
        # Admin always has all permissions
        if u['role'] == 'admin':
            perms = [p['id'] for p in AVAILABLE_PAGES]
        user_perms.append({
            'id': u['id'],
            'username': u['username'],
            'full_name': u['full_name'],
            'role': u['role'],
            'role_label': ROLE_LABELS.get(u['role'], u['role']),
            'permissions': perms
        })

    return jsonify({
        'available_pages': AVAILABLE_PAGES,
        'users': user_perms
    })

@app.route('/api/permissions/<int:user_id>', methods=['PUT'])
@admin_required
def update_permissions(user_id):
    """Update permissions for a specific user"""
    import json as _json
    db = get_db()

    # Cannot modify admin permissions
    target = db.execute("SELECT id, role FROM users WHERE id=?", (user_id,)).fetchone()
    if not target:
        return jsonify({'error': 'Usuario no encontrado'}), 404
    if target['role'] == 'admin':
        return jsonify({'error': 'No se pueden modificar los permisos del administrador'}), 403

    data = request.get_json()
    permissions = data.get('permissions', [])

    # Validate permissions
    valid_ids = [p['id'] for p in AVAILABLE_PAGES]
    permissions = [p for p in permissions if p in valid_ids]

    db.execute("UPDATE users SET permissions=?, updated_at=datetime('now') WHERE id=?",
               (_json.dumps(permissions), user_id))
    db.commit()

    return jsonify({'message': 'Permisos actualizados', 'permissions': permissions})

# ============================================================
# Contacts API
# ============================================================

def _contact_visible_where(alias: str = "c") -> tuple[str, list]:
    """Return (WHERE clause fragment, params) for contacts visible to current user.

    admin          -> all contacts
    team_admin     -> own + team members' contacts
    team_member    -> own contacts only
    """
    uid = session.get('user_id')
    role = session.get('role')
    if role == 'admin':
        return "1=1", []
    if role == 'team_admin':
        return f"({alias}.created_by = ? OR {alias}.created_by IN (SELECT id FROM users WHERE team_creator_id = ?))", [uid, uid]
    return f"{alias}.created_by = ?", [uid]


def _can_manage_contact(contact: sqlite3.Row | None) -> bool:
    """Whether the current user is allowed to edit/delete the given contact."""
    if contact is None:
        return False
    role = session.get('role')
    if role == 'admin':
        return True
    owner = contact['created_by']
    if role == 'team_admin':
        return owner == session['user_id'] or owner in _team_member_ids(session['user_id'])
    return owner == session['user_id']


def _team_member_ids(team_admin_id: int) -> list[int]:
    db = get_db()
    return [r['id'] for r in db.execute(
        "SELECT id FROM users WHERE team_creator_id=?", (team_admin_id,)
    ).fetchall()]


def _parse_money(value):
    """Parse a numeric money field coming from JSON; return float or 0.0."""
    if value is None or value == '':
        return 0.0
    try:
        v = float(str(value).replace(',', '').strip())
        return round(v, 2)
    except (TypeError, ValueError):
        return 0.0


@app.route('/api/contacts', methods=['GET'])
@login_required
def list_contacts():
    db = get_db()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '').strip()
    group_id = request.args.get('group_id', '', type=str)
    offset = (page - 1) * per_page

    where, scope_params = _contact_visible_where("c")
    query = ("SELECT c.*, cg.name as group_name FROM contacts c "
             "LEFT JOIN contact_groups cg ON c.group_id = cg.id WHERE " + where)
    count_query = ("SELECT COUNT(*) as total FROM contacts c "
                   "LEFT JOIN contact_groups cg ON c.group_id = cg.id WHERE " + where)
    params = list(scope_params)
    if search:
        query += " AND (c.name LIKE ? OR c.phone LIKE ? OR c.notes LIKE ?)"
        count_query += " AND (c.name LIKE ? OR c.phone LIKE ? OR c.notes LIKE ?)"
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])
    if group_id:
        query += " AND c.group_id = ?"
        count_query += " AND c.group_id = ?"
        params.append(int(group_id))
    remark = request.args.get('remark', '').strip()
    if remark:
        query += " AND c.remark = ?"
        count_query += " AND c.remark = ?"
        params.append(remark)
    total = db.execute(count_query, params).fetchone()['total']
    query += " ORDER BY c.created_at DESC LIMIT ? OFFSET ?"
    params.extend([per_page, offset])
    contacts = db.execute(query, params).fetchall()
    return jsonify({
        'contacts': [dict(c) for c in contacts],
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page
    })

@app.route('/api/contacts', methods=['POST'])
@login_required
def create_contact():
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    phone = normalize_phone((data.get('phone') or '').strip())
    notes = (data.get('notes') or '').strip()
    remark = (data.get('remark') or '').strip()
    app_name = (data.get('app_name') or '').strip()[:255]
    amount = _parse_money(data.get('amount'))
    discount_amount = _parse_money(data.get('discount_amount'))
    payment_link = (data.get('payment_link') or '').strip()
    group_id = data.get('group_id', None)
    if not name or not phone:
        return jsonify({'error': 'Nombre y telefono son requeridos'}), 400
    db = get_db()
    user_id = session.get('user_id')
    user_role = session.get('role')
    if group_id:
        if user_role == 'team_member':
            group = db.execute(
                "SELECT id FROM contact_groups WHERE id=? AND (created_by=? OR created_by IS NULL)",
                (int(group_id), user_id),
            ).fetchone()
        else:
            group = db.execute("SELECT id FROM contact_groups WHERE id=?", (int(group_id),)).fetchone()
        if not group:
            group_id = None
    db.execute(
        "INSERT INTO contacts (name, phone, notes, remark, group_id, app_name, amount, discount_amount, payment_link, created_by) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (name, phone, notes, remark, group_id, app_name, amount, discount_amount, payment_link,
         session.get('user_id'))
    )
    db.commit()
    return jsonify({'message': 'Contacto creado', 'phone': phone}), 201


@app.route('/api/contacts/import-device', methods=['POST'])
@login_required
def import_device_contacts():
    """Bulk import contacts selected from the device address book."""
    payload = request.get_json(silent=True) or {}
    contacts = payload.get('contacts')
    group_id = payload.get('group_id')

    if not isinstance(contacts, list) or not contacts:
        return jsonify({'error': 'No se recibieron contactos validos'}), 400
    if len(contacts) > 2000:
        return jsonify({'error': 'Maximo 2000 contactos por importacion'}), 400

    db = get_db()
    if group_id:
        group = db.execute("SELECT id FROM contact_groups WHERE id=?", (int(group_id),)).fetchone()
        if not group:
            group_id = None

    created = 0
    skipped = 0
    seen = set()
    for item in contacts:
        if not isinstance(item, dict):
            skipped += 1
            continue
        name = (str(item.get('name') or '').strip())[:100]
        phone = normalize_phone((str(item.get('phone') or '').strip())[:30])
        notes = (str(item.get('notes') or '').strip())[:255]
        if not name:
            name = phone or 'Sin nombre'
        if not phone:
            skipped += 1
            continue
        key = phone.lstrip('+').replace(' ', '').replace('-', '')
        if key in seen:
            skipped += 1
            continue
        seen.add(key)
        existing = db.execute(
            "SELECT id FROM contacts WHERE REPLACE(REPLACE(REPLACE(phone, ' ', ''), '-', ''), '+', '') = ? LIMIT 1",
            (key.lstrip('+'),)
        ).fetchone()
        if existing:
            skipped += 1
            continue
        db.execute(
            "INSERT INTO contacts (name, phone, notes, remark, group_id, created_by) VALUES (?, ?, ?, ?, ?, ?)",
            (name, phone, notes, payload.get('remark', 'Contacto del dispositivo'), group_id, session.get('user_id'))
        )
        created += 1
    db.commit()
    return jsonify({'message': 'Contactos importados', 'created': created, 'skipped': skipped}), 201

@app.route('/api/contacts/<int:contact_id>', methods=['PUT'])
@login_required
def update_contact(contact_id):
    data = request.get_json()
    db = get_db()
    contact = db.execute("SELECT * FROM contacts WHERE id=?", (contact_id,)).fetchone()
    if not contact:
        return jsonify({'error': 'Contacto no encontrado'}), 404
    if not _can_manage_contact(contact):
        return jsonify({'error': 'No autorizado: solo puedes editar tus contactos o los de tu equipo'}), 403
    name = data.get('name', contact['name'])
    raw_phone = data.get('phone', contact['phone'])
    phone = normalize_phone(raw_phone) if raw_phone else contact['phone']
    if not phone:
        return jsonify({'error': 'Teléfono inválido'}), 400
    notes = data.get('notes', contact['notes'])
    remark = data.get('remark', contact['remark'])
    group_id = data.get('group_id', contact['group_id'])
    app_name = (str(data.get('app_name', contact['app_name'] or '') or '')).strip()[:255]
    amount = _parse_money(data.get('amount', contact['amount']))
    discount_amount = _parse_money(data.get('discount_amount', contact['discount_amount']))
    payment_link = (str(data.get('payment_link', contact['payment_link'] or '') or '')).strip()
    db.execute(
        "UPDATE contacts SET name=?, phone=?, notes=?, remark=?, group_id=?, app_name=?, amount=?, discount_amount=?, payment_link=? WHERE id=?",
        (name, phone, notes, remark, group_id, app_name, amount, discount_amount, payment_link, contact_id)
    )
    db.commit()
    return jsonify({'message': 'Contacto actualizado'})

@app.route('/api/contacts/<int:contact_id>', methods=['DELETE'])
@login_required
def delete_contact(contact_id):
    db = get_db()
    contact = db.execute("SELECT * FROM contacts WHERE id=?", (contact_id,)).fetchone()
    if not contact:
        return jsonify({'error': 'Contacto no encontrado'}), 404
    if not _can_manage_contact(contact):
        return jsonify({'error': 'No autorizado: solo puedes eliminar tus contactos o los de tu equipo'}), 403
    db.execute("DELETE FROM contacts WHERE id=?", (contact_id,))
    db.commit()
    return jsonify({'message': 'Contacto eliminado'})

@app.route('/api/contacts/template', methods=['GET'])
@login_required
def download_contact_template():
    """Download a CSV template for bulk contact import."""
    import io as _io
    from flask import send_file

    buf = _io.StringIO()
    # Write UTF-8 BOM so Excel opens the file with correct encoding
    buf.write('\ufeff')
    writer = csv.writer(buf)
    writer.writerow(['name', 'phone', 'notes', 'remark', 'app_name', 'amount', 'discount_amount', 'payment_link'])
    writer.writerow([
        'Juan Perez',
        '5215512345678',
        'Cliente interesado en promo MXN',
        'Dispuesto a pagar sin fondos',
        'App Recargas',
        '150.00',
        '20.00',
        'https://pago.ejemplo.com/juan',
    ])
    writer.writerow([
        'Maria Lopez',
        '5215587654321',
        'No molestar despues de las 20h',
        'No contactable',
        '',
        '',
        '',
        '',
    ])
    writer.writerow([
        'Carlos Ruiz',
        '5215511223344',
        '',
        'Promesa de pago',
        'App Prestamos',
        '1200.50',
        '100.00',
        'https://pago.ejemplo.com/carlos',
    ])

    data = buf.getvalue().encode('utf-8')
    return send_file(
        _io.BytesIO(data),
        as_attachment=True,
        download_name='plantilla_contactos.csv',
        mimetype='text/csv; charset=utf-8',
    )


@app.route('/api/contacts/import', methods=['POST'])
@login_required
def import_contacts():
    if 'file' not in request.files:
        return jsonify({'error': 'No se ha subido ningun archivo'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Archivo vacio'}), 400
    if not file.filename.endswith('.csv'):
        return jsonify({'error': 'Solo se aceptan archivos CSV'}), 400
    group_id = request.form.get('group_id', None)
    if group_id:
        group_id = int(group_id)
    try:
        content = file.stream.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        db = get_db()
        imported = 0
        errors = []
        for i, row in enumerate(reader, start=2):
            name = (row.get('name') or row.get('nombre') or '').strip()
            phone = normalize_phone((row.get('phone') or row.get('telefono') or row.get('tel') or '').strip())
            notes = (row.get('notes') or row.get('notas') or row.get('observaciones') or '').strip()
            remark = (row.get('remark') or row.get('nota') or '').strip()
            app_name = (row.get('app_name') or row.get('app') or '').strip()[:255]
            amount = _parse_money(row.get('amount') or row.get('monto'))
            discount_amount = _parse_money(row.get('discount_amount') or row.get('descuento'))
            payment_link = (row.get('payment_link') or row.get('link_pago') or row.get('url_pago') or '').strip()
            if not name or not phone:
                errors.append(f"Fila {i}: nombre y telefono son requeridos")
                continue
            db.execute(
                "INSERT INTO contacts (name, phone, notes, remark, group_id, app_name, amount, discount_amount, payment_link) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (name, phone, notes, remark, group_id, app_name, amount, discount_amount, payment_link)
            )
            imported += 1
        db.commit()
        result = {'imported': imported}
        if errors:
            result['errors'] = errors[:10]
        return jsonify(result), 201
    except Exception as e:
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}'}), 400

# ============================================================
# Groups API
# ============================================================

@app.route('/api/groups', methods=['GET'])
@login_required
def list_groups():
    db = get_db()
    groups = db.execute("""
        SELECT cg.*, COUNT(c.id) as contact_count
        FROM contact_groups cg
        LEFT JOIN contacts c ON c.group_id = cg.id
        GROUP BY cg.id
        ORDER BY cg.created_at DESC
    """).fetchall()
    return jsonify({'groups': [dict(gr) for gr in groups]})

@app.route('/api/groups', methods=['POST'])
@login_required
def create_group():
    data = request.get_json()
    name = data.get('name', '').strip()
    description = data.get('description', '').strip()
    if not name:
        return jsonify({'error': 'El nombre del grupo es requerido'}), 400
    db = get_db()
    db.execute("INSERT INTO contact_groups (name, description) VALUES (?, ?)", (name, description))
    db.commit()
    return jsonify({'message': 'Grupo creado'}), 201

@app.route('/api/groups/<int:group_id>', methods=['PUT'])
@login_required
def update_group(group_id):
    data = request.get_json()
    db = get_db()
    group = db.execute("SELECT * FROM contact_groups WHERE id=?", (group_id,)).fetchone()
    if not group:
        return jsonify({'error': 'Grupo no encontrado'}), 404
    name = data.get('name', group['name'])
    description = data.get('description', group['description'])
    db.execute("UPDATE contact_groups SET name=?, description=? WHERE id=?", (name, description, group_id))
    db.commit()
    return jsonify({'message': 'Grupo actualizado'})

@app.route('/api/groups/<int:group_id>', methods=['DELETE'])
@login_required
def delete_group(group_id):
    db = get_db()
    group = db.execute("SELECT * FROM contact_groups WHERE id=?", (group_id,)).fetchone()
    if not group:
        return jsonify({'error': 'Grupo no encontrado'}), 404
    db.execute("UPDATE contacts SET group_id=NULL WHERE group_id=?", (group_id,))
    db.execute("DELETE FROM contact_groups WHERE id=?", (group_id,))
    db.commit()
    return jsonify({'message': 'Grupo eliminado'})

# ============================================================
# Templates API
# ============================================================

@app.route('/api/templates', methods=['GET'])
@login_required
def list_templates():
    db = get_db()
    search = request.args.get('search', '').strip()
    category = request.args.get('category', '').strip()
    query = "SELECT * FROM templates WHERE 1=1"
    params = []
    if search:
        query += " AND (name LIKE ? OR content LIKE ?)"
        params.extend([f'%{search}%', f'%{search}%'])
    if category:
        query += " AND category = ?"
        params.append(category)
    query += " ORDER BY updated_at DESC"
    templates = db.execute(query, params).fetchall()
    return jsonify({'templates': [dict(t) for t in templates]})

@app.route('/api/templates', methods=['POST'])
@login_required
def create_template():
    data = request.get_json()
    name = data.get('name', '').strip()
    content = data.get('content', '').strip()
    category = data.get('category', 'general').strip()
    if not name or not content:
        return jsonify({'error': 'Nombre y contenido son requeridos'}), 400
    db = get_db()
    db.execute(
        "INSERT INTO templates (name, content, category) VALUES (?, ?, ?)",
        (name, content, category)
    )
    db.commit()
    return jsonify({'message': 'Plantilla creada'}), 201

@app.route('/api/templates/<int:template_id>', methods=['PUT'])
@login_required
def update_template(template_id):
    data = request.get_json()
    db = get_db()
    template = db.execute("SELECT * FROM templates WHERE id=?", (template_id,)).fetchone()
    if not template:
        return jsonify({'error': 'Plantilla no encontrada'}), 404
    name = data.get('name', template['name'])
    content = data.get('content', template['content'])
    category = data.get('category', template['category'])
    db.execute(
        "UPDATE templates SET name=?, content=?, category=?, updated_at=datetime('now') WHERE id=?",
        (name, content, category, template_id)
    )
    db.commit()
    return jsonify({'message': 'Plantilla actualizada'})

@app.route('/api/templates/<int:template_id>', methods=['DELETE'])
@login_required
def delete_template(template_id):
    db = get_db()
    template = db.execute("SELECT * FROM templates WHERE id=?", (template_id,)).fetchone()
    if not template:
        return jsonify({'error': 'Plantilla no encontrada'}), 404
    db.execute("DELETE FROM templates WHERE id=?", (template_id,))
    db.commit()
    return jsonify({'message': 'Plantilla eliminada'})

# ============================================================
# SMS API
# ============================================================

@app.route('/api/sms/send', methods=['POST'])
@login_required
def send_sms():
    data = request.get_json()
    phones = data.get('phones', [])

    # NOTE: Los limites de envio (limite diario por miembro y tope de 500
    # numeros por envio) estan desactivados en produccion. El campo
    # team_config.daily_sms_limit y su pantalla de configuracion se conservan
    # para poder reactivarlos en el futuro sin perder datos.
    content = data.get('content', '').strip()
    contact_names = data.get('contact_names', {})
    if not phones or not content:
        return jsonify({'error': 'Numero(s) y contenido son requeridos'}), 400
    db = get_db()
    api_configured = is_sms_api_configured(g.user['id'])
    sms_config = get_team_sms_config(g.user['id'])
    contact_cache = build_contact_template_cache(db, phones)
    records = []
    errors = []

    if api_configured and len(phones) == 1:
        # Single SMS - use /sms/send endpoint
        raw_phone = phones[0].strip()
        phone = normalize_phone(raw_phone)
        name = contact_names.get(raw_phone, '') or contact_names.get(phone, '')
        msg = apply_template_vars(content, raw_phone, contact_names, contact_cache)
        result = sms_api_send_single(phone, msg)
        api_code = result.get('code', -1)
        api_msg = result.get('msg', '')
        if api_code == 0:
            msgid = ''
            if result.get('data'):
                msgid = result['data'].get('msgid', '')
            db.execute(
                "INSERT INTO sms_records (phone, contact_name, content, status, msgid, api_code, api_msg, sent_at, created_by) VALUES (?, ?, ?, 'sent', ?, ?, ?, datetime('now'), ?)",
                (phone, name, msg, msgid, api_code, api_msg, g.user['id'])
            )
            records.append({'phone': phone, 'status': 'sent', 'msgid': msgid})
        else:
            status_text = SMS_STATUS_CODES.get(api_code, api_msg)
            db.execute(
                "INSERT INTO sms_records (phone, contact_name, content, status, api_code, api_msg, sent_at, created_by) VALUES (?, ?, ?, 'failed', ?, ?, datetime('now'), ?)",
                (phone, name, msg, api_code, f"Code {api_code}: {status_text}", g.user['id'])
            )
            records.append({'phone': phone, 'status': 'failed', 'error': status_text})
            errors.append(f"{phone}: {status_text}")

    elif api_configured and len(phones) > 1:
        # Multiple SMS - use /sms/rsend endpoint (max 200 per batch)
        phone_content_pairs = []
        phone_name_map = {}
        for raw in phones:
            raw = (raw or '').strip()
            if not raw:
                continue
            phone = normalize_phone(raw)
            name = contact_names.get(raw, '') or contact_names.get(phone, '')
            msg = apply_template_vars(content, raw, contact_names, contact_cache)
            phone_content_pairs.append((phone, msg))
            phone_name_map[phone] = (name, msg)

        # Split into batches of 200
        for batch_start in range(0, len(phone_content_pairs), 200):
            batch = phone_content_pairs[batch_start:batch_start + 200]
            result = sms_api_send_batch(batch)
            api_code = result.get('code', -1)
            api_msg = result.get('msg', '')

            if api_code == 0 and result.get('data'):
                # Map results back to phones. The API returns das in 0052... form,
                # so build a lookup keyed by both das and the +52... normalized form.
                result_map = {}
                for item in result['data']:
                    das_val = str(item.get('das', '') or '').strip()
                    if das_val:
                        plus_form = ('+' + das_val[2:]) if das_val.startswith('00') else ('+' + das_val)
                        result_map[das_val] = item
                        result_map[plus_form] = item

                for phone, msg in batch:
                    name = phone_name_map[phone][0]
                    item = result_map.get(phone) or result_map.get(phone_to_das(phone)) or {}
                    item_code = item.get('state', 0)
                    msgid = str(item.get('msgid', '') or '')
                    if item_code == 0:
                        status = 'sent'
                    else:
                        status = 'failed'
                    db.execute(
                        "INSERT INTO sms_records (phone, contact_name, content, status, msgid, api_code, api_msg, sent_at, created_by) VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'), ?)",
                        (phone, name, msg, status, msgid, item_code, api_msg, g.user['id'])
                    )
                    records.append({'phone': phone, 'status': status, 'msgid': msgid})
                    if status == 'failed':
                        errors.append(f"{phone}: {SMS_STATUS_CODES.get(item_code, 'Error desconocido')}")
            else:
                # Entire batch failed
                status_text = SMS_STATUS_CODES.get(api_code, api_msg)
                for phone, msg in batch:
                    name = phone_name_map[phone][0]
                    db.execute(
                        "INSERT INTO sms_records (phone, contact_name, content, status, api_code, api_msg, sent_at, created_by) VALUES (?, ?, ?, 'failed', ?, ?, datetime('now'), ?)",
                        (phone, name, msg, api_code, f"Code {api_code}: {status_text}", g.user['id'])
                    )
                    records.append({'phone': phone, 'status': 'failed', 'error': status_text})
                    errors.append(f"{phone}: {status_text}")
    else:
        # API not configured - simulate sending (for testing)
        for phone in phones:
            phone = phone.strip()
            if not phone:
                continue
            name = contact_names.get(phone, '')
            msg = apply_template_vars(content, phone, contact_names, contact_cache)
            db.execute(
                "INSERT INTO sms_records (phone, contact_name, content, status, api_msg, sent_at, created_by) VALUES (?, ?, ?, 'sent', ?, datetime('now'), ?)",
                (phone, name, msg, 'API no configurada - envio simulado', g.user['id'])
            )
            records.append({'phone': phone, 'status': 'sent', 'simulated': True})

    db.commit()
    # Log the action
    log_status = 'success' if not errors else ('partial' if records else 'error')
    db.execute(
        "INSERT INTO send_logs (action, details, status) VALUES (?, ?, ?)",
        ('send', json.dumps({'phones_count': len(records), 'user': g.user['username'], 'api_configured': api_configured, 'errors': errors[:5]}), log_status)
    )
    db.commit()

    sent_count = sum(1 for r in records if r['status'] == 'sent')
    failed_count = sum(1 for r in records if r['status'] == 'failed')
    result_msg = f'{sent_count} mensaje(s) enviado(s)'
    if failed_count > 0:
        result_msg += f', {failed_count} fallido(s)'
    if not api_configured:
        result_msg += ' (modo simulacion - API no configurada)'

    return jsonify({'message': result_msg, 'records': records, 'errors': errors[:10]})

@app.route('/api/sms/schedule', methods=['POST'])
@login_required
def schedule_sms():
    data = request.get_json()
    phones = data.get('phones', [])
    content = data.get('content', '').strip()
    scheduled_at = data.get('scheduled_at', '').strip()
    contact_names = data.get('contact_names', {})
    if not phones or not content or not scheduled_at:
        return jsonify({'error': 'Numero(s), contenido y fecha son requeridos'}), 400
    db = get_db()
    contact_cache = build_contact_template_cache(db, phones)
    count = 0
    for phone in phones:
        phone = phone.strip()
        if not phone:
            continue
        name = contact_names.get(phone, '')
        msg = apply_template_vars(content, phone, contact_names, contact_cache)
        db.execute(
            "INSERT INTO sms_records (phone, contact_name, content, status, scheduled_at, created_by) VALUES (?, ?, ?, 'scheduled', ?, ?)",
            (phone, name, msg, scheduled_at, g.user['id'])
        )
        count += 1
    db.commit()
    return jsonify({'message': f'{count} mensaje(s) programado(s)'}), 201

@app.route('/api/sms/records', methods=['GET'])
@login_required
def list_sms_records():
    db = get_db()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    search = request.args.get('search', '').strip()
    offset = (page - 1) * per_page
    query = ("SELECT r.*, u.username AS sender_username, u.full_name AS sender_full_name, "
             "u.role AS sender_role FROM sms_records r "
             "LEFT JOIN users u ON u.id = r.created_by WHERE 1=1")
    count_query = "SELECT COUNT(*) as total FROM sms_records r WHERE 1=1"
    params = []
    # Role-based scope filtering
    if g.user['role'] == 'team_member':
        query += " AND r.created_by = ?"
        count_query += " AND created_by = ?"
        params.append(g.user['id'])
    elif g.user['role'] == 'team_admin':
        query += " AND r.created_by IN (SELECT id FROM users WHERE id=? OR team_creator_id=?)"
        count_query += " AND created_by IN (SELECT id FROM users WHERE id=? OR team_creator_id=?)"
        params.extend([g.user['id'], g.user['id']])
    # admin sees all
    if status:
        query += " AND r.status = ?"
        count_query += " AND status = ?"
        params.append(status)
    if date_from:
        query += " AND r.created_at >= ?"
        count_query += " AND created_at >= ?"
        params.append(date_from)
    if date_to:
        query += " AND r.created_at <= ?"
        count_query += " AND created_at <= ?"
        params.append(date_to + ' 23:59:59')
    if search:
        query += " AND (r.phone LIKE ? OR r.contact_name LIKE ? OR r.content LIKE ? OR u.username LIKE ?)"
        count_query += " AND (phone LIKE ? OR contact_name LIKE ? OR content LIKE ?)"
        count_params = params + [f'%{search}%'] * 3
        params.extend([f'%{search}%'] * 4)
    else:
        count_params = list(params)
    total = db.execute(count_query, count_params).fetchone()['total']
    query += " ORDER BY r.created_at DESC LIMIT ? OFFSET ?"
    params.extend([per_page, offset])
    records = db.execute(query, params).fetchall()
    return jsonify({
        'records': [dict(r) for r in records],
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page
    })

@app.route('/api/sms/statistics', methods=['GET'])
@login_required
def sms_statistics():
    db = get_db()
    # Build scope filter based on role
    scope_filter = ''
    scope_params = []
    if g.user['role'] == 'team_member':
        scope_filter = ' AND created_by = ?'
        scope_params = [g.user['id']]
    elif g.user['role'] == 'team_admin':
        scope_filter = ' AND created_by IN (SELECT id FROM users WHERE id=? OR team_creator_id=?)'
        scope_params = [g.user['id'], g.user['id']]
    today = datetime.now().strftime('%Y-%m-%d')
    today_sent = db.execute(
        f"SELECT COUNT(*) as count FROM sms_records WHERE status='sent' AND date(sent_at)=? {scope_filter}",
        [today] + scope_params
    ).fetchone()['count']
    today_total = db.execute(
        f"SELECT COUNT(*) as count FROM sms_records WHERE date(created_at)=? {scope_filter}",
        [today] + scope_params
    ).fetchone()['count']
    total_sent = db.execute(
        f"SELECT COUNT(*) as count FROM sms_records WHERE status='sent' {scope_filter}",
        scope_params
    ).fetchone()['count']
    total_failed = db.execute(
        f"SELECT COUNT(*) as count FROM sms_records WHERE status='failed' {scope_filter}",
        scope_params
    ).fetchone()['count']
    total_pending = db.execute(
        f"SELECT COUNT(*) as count FROM sms_records WHERE status IN ('pending', 'scheduled') {scope_filter}",
        scope_params
    ).fetchone()['count']
    total_all = db.execute(
        f"SELECT COUNT(*) as count FROM sms_records WHERE 1=1 {scope_filter}",
        scope_params
    ).fetchone()['count']
    # Last 7 days stats
    last_7_days = []
    for i in range(6, -1, -1):
        day = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        count = db.execute(
            f"SELECT COUNT(*) as count FROM sms_records WHERE status='sent' AND date(sent_at)=? {scope_filter}",
            [day] + scope_params
        ).fetchone()['count']
        last_7_days.append({'date': day, 'count': count})
    success_rate = (total_sent / total_all * 100) if total_all > 0 else 0
    return jsonify({
        'today_sent': today_sent,
        'today_total': today_total,
        'total_sent': total_sent,
        'total_failed': total_failed,
        'total_pending': total_pending,
        'total_all': total_all,
        'success_rate': round(success_rate, 1),
        'last_7_days': last_7_days,
        'total_contacts': db.execute("SELECT COUNT(*) as count FROM contacts").fetchone()['count'],
        'total_templates': db.execute("SELECT COUNT(*) as count FROM templates").fetchone()['count']
    })

# ============================================================
# SMS Config API (admin)
# ============================================================

@app.route('/api/config/sms', methods=['GET'])
@admin_required
def get_sms_configs():
    db = get_db()
    configs = db.execute("SELECT * FROM sms_api_configs ORDER BY id").fetchall()
    config_list = []
    for c in configs:
        config_list.append({
            'id': c['id'],
            'name': c['name'] or '',
            'country': c['country'] or '',
            'domain': c['domain'] or '',
            'spid': c['spid'] or '',
            'api_pwd': c['api_pwd'] or '',
            'sender_name': c['sender_name'] or '',
            'is_active': bool(c['is_active']),
            'updated_at': c['updated_at']
        })
    return jsonify({'configs': config_list})

@app.route('/api/config/sms', methods=['POST'])
@admin_required
def create_sms_config():
    data = request.get_json()
    db = get_db()
    name = data.get('name', '').strip()
    country = data.get('country', '').strip()
    domain = data.get('domain', '').strip()
    spid = data.get('spid', '').strip()
    api_pwd = data.get('api_pwd', '').strip()
    sender_name = data.get('sender_name', '').strip()
    if not name:
        return jsonify({'error': 'Nombre es requerido'}), 400
    try:
        cursor = db.execute(
            "INSERT INTO sms_api_configs (name, country, domain, spid, api_pwd, sender_name) VALUES (?, ?, ?, ?, ?, ?)",
            (name, country, domain, spid, api_pwd, sender_name)
        )
        db.commit()
        return jsonify({'message': 'Configuracion creada', 'id': cursor.lastrowid})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/config/sms/<int:config_id>', methods=['PUT'])
@admin_required
def update_sms_config(config_id):
    data = request.get_json()
    db = get_db()
    name = data.get('name', '').strip()
    country = data.get('country', '').strip()
    domain = data.get('domain', '').strip()
    spid = data.get('spid', '').strip()
    api_pwd = data.get('api_pwd', '').strip()
    sender_name = data.get('sender_name', '').strip()
    is_active = data.get('is_active', True)
    db.execute(
        "UPDATE sms_api_configs SET name=?, country=?, domain=?, spid=?, api_pwd=?, sender_name=?, is_active=?, updated_at=datetime('now') WHERE id=?",
        (name, country, domain, spid, api_pwd, sender_name, 1 if is_active else 0, config_id)
    )
    db.commit()
    return jsonify({'message': 'Configuracion actualizada'})

@app.route('/api/config/sms/<int:config_id>', methods=['DELETE'])
@admin_required
def delete_sms_config(config_id):
    db = get_db()
    db.execute("DELETE FROM sms_api_configs WHERE id=?", (config_id,))
    db.commit()
    return jsonify({'message': 'Configuracion eliminada'})

@app.route('/api/config/sms/test', methods=['POST'])
@admin_required
def test_sms_config():
    data = request.get_json()
    config_id = data.get('config_id')
    db = get_db()
    config = get_sms_api_config(config_id)
    if not config or not (config.get('domain') and config.get('spid') and config.get('api_pwd')):
        return jsonify({'error': 'Configure el Dominio, SPID y Contrasena API primero'}), 400

    # Test connection by calling charset check with a simple text
    result = sms_api_check_charset("Test", config)
    api_code = result.get('code', -1)

    if api_code == 0:
        charset_info = result.get('data', {})
        db.execute(
            "INSERT INTO send_logs (action, details, status) VALUES (?, ?, ?)",
            ('test_connection', json.dumps({
                'domain': config['domain'],
                'spid': config['spid'],
                'charset': charset_info.get('charset', 'N/A'),
                'result': 'success'
            }), 'success')
        )
        db.commit()
        return jsonify({
            'message': f'Conexion exitosa. Codificacion detectada: {charset_info.get("charset", "N/A")}. '
                       f'Longitud por SMS: {charset_info.get("single", "N/A")} caracteres.'
        })
    else:
        error_msg = SMS_STATUS_CODES.get(api_code, result.get('msg', 'Error desconocido'))
        db.execute(
            "INSERT INTO send_logs (action, details, status) VALUES (?, ?, ?)",
            ('test_connection', json.dumps({
                'domain': config['domain'],
                'spid': config['spid'],
                'error_code': api_code,
                'error_msg': error_msg,
                'result': 'failed'
            }), 'error')
        )
        db.commit()
        return jsonify({'error': f'Error de conexion (codigo {api_code}): {error_msg}'}), 400

# ---------- Configuracion de facturacion ----------
@app.route('/api/settings/billing', methods=['GET'])
@login_required
def get_billing_settings():
    """Precio por SMS (configurado por el admin). Cualquier usuario puede leerlo."""
    db = get_db()
    return jsonify({'sms_unit_price': get_sms_unit_price(db)})

@app.route('/api/settings/billing', methods=['PUT'])
@admin_required
def update_billing_settings():
    """El administrador define el coste por SMS enviado (se cobra por intento)."""
    db = get_db()
    data = request.get_json() or {}
    raw = data.get('sms_unit_price')
    try:
        price = float(raw)
    except (TypeError, ValueError):
        return jsonify({'error': 'El precio por SMS debe ser un numero valido'}), 400
    if price < 0:
        return jsonify({'error': 'El precio por SMS no puede ser negativo'}), 400
    price = round(price, 6)
    set_sms_unit_price(price, db)
    db.execute(
        "INSERT INTO send_logs (action, details, status) VALUES (?, ?, ?)",
        ('billing_update', json.dumps({'sms_unit_price': price}, ensure_ascii=False), 'success')
    )
    db.commit()
    return jsonify({'message': 'Precio por SMS actualizado', 'sms_unit_price': price})

@app.route('/api/admin/user-usage', methods=['GET'])
@manager_required
def get_user_usage():
    """Get per-user SMS Usage statistics. System admin sees all; team admin sees own team."""
    db = get_db()
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    # Build date filter
    date_filter = ''
    params = []
    if date_from:
        date_filter += ' AND date(r.created_at) >= ?'
        params.append(date_from)
    if date_to:
        date_filter += ' AND date(r.created_at) <= ?'
        params.append(date_to)
    # Build user filter
    user_filter = ''
    if g.user['role'] == 'team_admin':
        user_filter = ' AND (u.id = ? OR u.team_creator_id = ?)'
        params = [g.user['id'], g.user['id']] + params
    query = f"""
        SELECT
            u.id as user_id,
            u.username,
            u.full_name,
            u.role,
            u.is_active,
            u.team_creator_id,
            COALESCE(SUM(CASE WHEN r.status='sent' THEN 1 ELSE 0 END), 0) as sent,
            COALESCE(SUM(CASE WHEN r.status='failed' THEN 1 ELSE 0 END), 0) as failed,
            COALESCE(SUM(CASE WHEN r.status IN ('pending','scheduled') THEN 1 ELSE 0 END), 0) as pending,
            COALESCE(COUNT(r.id), 0) as total,
            MAX(r.created_at) as last_activity
        FROM users u
        LEFT JOIN sms_records r ON u.id = r.created_by {date_filter}
        WHERE 1=1 {user_filter}
        GROUP BY u.id
        ORDER BY total DESC
    """
    # Build user lookup for team affiliation
    user_map = {}
    all_users = db.execute('SELECT id, username, full_name FROM users').fetchall()
    for u in all_users:
        user_map[u['id']] = {'username': u['username'], 'full_name': u['full_name']}

    rows = db.execute(query, params).fetchall()
    users = []
    for row in rows:
        total = row['total']
        sent = row['sent']
        rate = round((sent / total * 100), 1) if total > 0 else 0
        # Team affiliation
        team_affiliation = '-'
        if row['role'] == 'team_member' and row['team_creator_id']:
            creator = user_map.get(row['team_creator_id'])
            if creator:
                team_affiliation = creator['full_name'] or creator['username']
        users.append({
            'user_id': row['user_id'],
            'username': row['username'],
            'full_name': row['full_name'],
            'role': row['role'],
            'role_label': ROLE_LABELS.get(row['role'], row['role']),
            'is_active': bool(row['is_active']),
            'team_affiliation': team_affiliation,
            'sent': sent,
            'failed': row['failed'],
            'pending': row['pending'],
            'total': total,
            'success_rate': rate,
            'last_activity': row['last_activity'] or ''
        })
    # Overall summary - only use date params (not user_filter params)
    date_params = []
    if date_from:
        date_params.append(date_from)
    if date_to:
        date_params.append(date_to)
    summary_filter = ''
    summary_params = list(date_params)
    if g.user['role'] == 'team_admin':
        summary_filter = ' AND created_by IN (SELECT id FROM users WHERE id=? OR team_creator_id=?)'
        summary_params.extend([g.user['id'], g.user['id']])
    summary = db.execute(f"""
        SELECT
            COUNT(*) as total_users,
            COALESCE(SUM(CASE WHEN status='sent' THEN 1 ELSE 0 END), 0) as total_sent,
            COALESCE(SUM(CASE WHEN status='failed' THEN 1 ELSE 0 END), 0) as total_failed,
            COALESCE(COUNT(id), 0) as total_all
        FROM sms_records WHERE 1=1 {date_filter} {summary_filter}
    """, summary_params).fetchone()

    # Team summary: aggregate by team (team_admin's team)
    team_summary = []
    if g.user['role'] == 'admin':
        # For system admin: show all teams
        team_rows = db.execute(f"""
            SELECT
                ta.id as team_admin_id,
                ta.username as team_admin_username,
                ta.full_name as team_admin_name,
                COUNT(DISTINCT m.id) + 1 as member_count,
                COALESCE(SUM(CASE WHEN r.status='sent' THEN 1 ELSE 0 END), 0) as team_sent,
                COALESCE(SUM(CASE WHEN r.status='failed' THEN 1 ELSE 0 END), 0) as team_failed,
                COALESCE(COUNT(r.id), 0) as team_total
            FROM users ta
            LEFT JOIN users m ON m.team_creator_id = ta.id
            LEFT JOIN sms_records r ON (r.created_by = ta.id OR r.created_by = m.id) {('AND date(r.created_at) >= ? AND date(r.created_at) <= ?' if date_from and date_to else '')}
            WHERE ta.role = 'team_admin'
            GROUP BY ta.id
            ORDER BY team_total DESC
        """, ([date_from, date_to] if date_from and date_to else [])).fetchall()
        for tr in team_rows:
            team_total = tr['team_total'] or 0
            team_sent = tr['team_sent'] or 0
            team_failed = tr['team_failed'] or 0
            team_rate = round((team_sent / team_total * 100), 1) if team_total > 0 else 0
            team_summary.append({
                'team_name': tr['team_admin_name'] or tr['team_admin_username'],
                'team_admin': tr['team_admin_username'],
                'member_count': tr['member_count'],
                'total': team_total,
                'sent': team_sent,
                'failed': team_failed,
                'rate': team_rate
            })
    elif g.user['role'] == 'team_admin':
        # For team admin: show own team only
        member_ids = [u['id'] for u in db.execute(
            "SELECT id FROM users WHERE team_creator_id = ?", (g.user['id'],)
        ).fetchall()]
        member_ids.append(g.user['id'])
        placeholders = ','.join('?' * len(member_ids))
        team_row = db.execute(f"""
            SELECT
                COUNT(DISTINCT u.id) as member_count,
                COALESCE(SUM(CASE WHEN r.status='sent' THEN 1 ELSE 0 END), 0) as team_sent,
                COALESCE(SUM(CASE WHEN r.status='failed' THEN 1 ELSE 0 END), 0) as team_failed,
                COALESCE(COUNT(r.id), 0) as team_total
            FROM users u
            LEFT JOIN sms_records r ON r.created_by = u.id {('AND date(r.created_at) >= ? AND date(r.created_at) <= ?' if date_from and date_to else '')}
            WHERE u.id IN ({placeholders})
        """, (member_ids + ([date_from, date_to] if date_from and date_to else []))).fetchone()
        team_total = team_row['team_total'] or 0
        team_sent = team_row['team_sent'] or 0
        team_failed = team_row['team_failed'] or 0
        team_rate = round((team_sent / team_total * 100), 1) if team_total > 0 else 0
        team_summary.append({
            'team_name': g.user['full_name'] or g.user['username'],
            'team_admin': g.user['username'],
            'member_count': team_row['member_count'],
            'total': team_total,
            'sent': team_sent,
            'failed': team_failed,
            'rate': team_rate
        })

    return jsonify({
        'users': users,
        'summary': {
            'total_users': summary['total_users'] if summary else 0,
            'total_sent': summary['total_sent'] if summary else 0,
            'total_failed': summary['total_failed'] if summary else 0,
            'total_all': summary['total_all'] if summary else 0
        },
        'team_summary': team_summary
    })

@app.route('/api/admin/unified-stats', methods=['GET'])
@login_required
def get_unified_stats():
    """Get unified statistics with 3 panels: my account, my team, all teams"""
    db = get_db()
    user_id = session['user_id']
    role = session['role']
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    filter_user_id = request.args.get('user_id', '')  # Account filter

    # Build date filter. One version for single-table queries (no alias),
    # one for JOIN queries using the sms_records alias `r`.
    date_filter = ''
    date_filter_r = ''
    date_params = []
    if date_from:
        date_filter += ' AND date(created_at) >= ?'
        date_filter_r += ' AND date(r.created_at) >= ?'
        date_params.append(date_from)
    if date_to:
        date_filter += ' AND date(created_at) <= ?'
        date_filter_r += ' AND date(r.created_at) <= ?'
        date_params.append(date_to)

    # 1. My Account stats (with optional account filter)
    account_user_id = int(filter_user_id) if filter_user_id and filter_user_id.isdigit() else user_id
    my_account = db.execute(f"""
        SELECT
            COALESCE(SUM(CASE WHEN status='sent' THEN 1 ELSE 0 END), 0) as sent,
            COALESCE(SUM(CASE WHEN status='failed' THEN 1 ELSE 0 END), 0) as failed,
            COALESCE(SUM(CASE WHEN status IN ('pending','scheduled') THEN 1 ELSE 0 END), 0) as pending,
            COALESCE(COUNT(id), 0) as total
        FROM sms_records WHERE created_by = ? {date_filter}
    """, [account_user_id] + date_params).fetchone()

    my_account_data = {
        'total': my_account['total'] if my_account else 0,
        'sent': my_account['sent'] if my_account else 0,
        'failed': my_account['failed'] if my_account else 0,
        'pending': my_account['pending'] if my_account else 0,
        'rate': round((my_account['sent'] / my_account['total'] * 100), 1) if my_account and my_account['total'] > 0 else 0
    }

    # 2. My Team stats
    my_team_data = None
    if role in ('admin', 'team_admin'):
        if role == 'team_admin':
            # Get team member IDs
            member_ids = [u['id'] for u in db.execute(
                "SELECT id FROM users WHERE team_creator_id = ?", (user_id,)
            ).fetchall()]
            member_ids.append(user_id)
            placeholders = ','.join('?' * len(member_ids))
            team_stats = db.execute(f"""
                SELECT
                    COUNT(DISTINCT u.id) as member_count,
                    COALESCE(SUM(CASE WHEN r.status='sent' THEN 1 ELSE 0 END), 0) as sent,
                    COALESCE(SUM(CASE WHEN r.status='failed' THEN 1 ELSE 0 END), 0) as failed,
                    COALESCE(SUM(CASE WHEN r.status IN ('pending','scheduled') THEN 1 ELSE 0 END), 0) as pending,
                    COALESCE(COUNT(r.id), 0) as total
                FROM users u
                LEFT JOIN sms_records r ON r.created_by = u.id {date_filter_r}
                WHERE u.id IN ({placeholders})
            """, member_ids + date_params).fetchone()
        else:
            # Admin: get all teams summary
            team_stats = db.execute(f"""
                SELECT
                    COUNT(DISTINCT u.id) as member_count,
                    COALESCE(SUM(CASE WHEN r.status='sent' THEN 1 ELSE 0 END), 0) as sent,
                    COALESCE(SUM(CASE WHEN r.status='failed' THEN 1 ELSE 0 END), 0) as failed,
                    COALESCE(SUM(CASE WHEN r.status IN ('pending','scheduled') THEN 1 ELSE 0 END), 0) as pending,
                    COALESCE(COUNT(r.id), 0) as total
                FROM users u
                LEFT JOIN sms_records r ON r.created_by = u.id {date_filter_r}
                WHERE u.role IN ('team_admin', 'team_member')
            """, date_params).fetchone()

        if team_stats:
            # Get team name (admin's username)
            team_name = 'Mi Equipo'
            daily_limit = 0
            if role == 'team_admin':
                admin_row = db.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
                if admin_row:
                    team_name = f"Equipo de {admin_row['username']}"
                # Get daily limit from team_config
                limit_row = db.execute("SELECT daily_sms_limit FROM team_config WHERE team_admin_id = ?", (user_id,)).fetchone()
                if limit_row and limit_row['daily_sms_limit']:
                    daily_limit = limit_row['daily_sms_limit']
            elif role == 'admin':
                team_name = 'Todos los Equipos'

            # Today's SMS count
            today_filter_simple = "AND date(created_at) = date('now')" if db.db_type != 'postgres' else "AND date(created_at) = CURRENT_DATE"
            if role == 'team_admin':
                today_row = db.execute(f"""
                    SELECT COUNT(*) as cnt FROM sms_records
                    WHERE created_by IN ({placeholders}) {today_filter_simple}
                """, member_ids).fetchone()
            else:
                today_row = db.execute(f"""
                    SELECT COUNT(*) as cnt FROM sms_records
                    WHERE created_by IN (SELECT id FROM users WHERE role IN ('team_admin','team_member'))
                    {today_filter_simple}
                """).fetchone()
            today_count = today_row['cnt'] if today_row else 0

            # Last activity
            if role == 'team_admin':
                last_row = db.execute(f"""
                    SELECT MAX(created_at) as last_at FROM sms_records
                    WHERE created_by IN ({placeholders})
                """, member_ids).fetchone()
            else:
                last_row = db.execute("""
                    SELECT MAX(created_at) as last_at FROM sms_records
                    WHERE created_by IN (SELECT id FROM users WHERE role IN ('team_admin','team_member'))
                """).fetchone()
            last_activity = last_row['last_at'] if last_row and last_row['last_at'] else None

            my_team_data = {
                'team_name': team_name,
                'member_count': team_stats['member_count'],
                'total': team_stats['total'],
                'sent': team_stats['sent'],
                'failed': team_stats['failed'],
                'pending': team_stats['pending'],
                'today': today_count,
                'daily_limit': daily_limit,
                'last_activity': last_activity,
                'rate': round((team_stats['sent'] / team_stats['total'] * 100), 1) if team_stats['total'] > 0 else 0
            }
    # 3. All Teams stats (admin only).
    #
    # The page lists one row per "team unit": every team_admin becomes a team
    # (the row aggregates that admin plus all members they created). The system
    # admin itself is NOT a separate team -- any SMS the admin sends is counted
    # under their own "Administrador" row in the detail table, so the summary
    # cards always equal the sum of all rows below (including the admin row).
    all_teams_data = None
    all_teams_list = []
    if role == 'admin':
        unit_price = get_sms_unit_price(db)

        def _build_unit_row(name, label, user_ids, unit_role='team_admin'):
            """Build one summary/detail row for a set of user ids."""
            if not user_ids:
                return {
                    'unit_role': unit_role,
                    'team_name': name, 'admin_name': name, 'team_admin': label,
                    'admin_username': label, 'team_admin_username': label,
                    'member_count': 0, 'total': 0, 'sent': 0, 'failed': 0,
                    'today': 0, 'rate': 0,
                    'total_cost': 0, 'sent_cost': 0, 'failed_cost': 0,
                    'cost_total': 0, 'cost_sent': 0, 'cost_failed': 0,
                }
            placeholders = ','.join(['?'] * len(user_ids))
            row = db.execute(f"""
                SELECT
                    COALESCE(SUM(CASE WHEN r.status='sent'   THEN 1 ELSE 0 END), 0) AS sent,
                    COALESCE(SUM(CASE WHEN r.status='failed' THEN 1 ELSE 0 END), 0) AS failed,
                    COALESCE(COUNT(r.id), 0) AS total
                FROM users u
                LEFT JOIN sms_records r ON r.created_by = u.id {date_filter_r}
                WHERE u.id IN ({placeholders})
            """, list(date_params) + user_ids).fetchone()
            t_total = row['total'] or 0
            t_sent = row['sent'] or 0
            t_failed = row['failed'] or 0
            today_row = db.execute(f"""
                SELECT COUNT(*) AS cnt FROM sms_records
                WHERE date(created_at) = date('now')
                  AND created_by IN ({placeholders})
            """, user_ids).fetchone()
            t_today = (today_row['cnt'] if today_row else 0) or 0
            t_rate = round((t_sent / t_total * 100), 1) if t_total > 0 else 0
            t_cost = round(t_total * unit_price, 4)
            s_cost = round(t_sent * unit_price, 4)
            f_cost = round(t_failed * unit_price, 4)
            return {
                'unit_role': unit_role,
                'team_name': name,
                'admin_name': name,
                'admin_username': label,
                'team_admin_username': label,
                'team_admin': label,
                'member_count': len(user_ids),
                'total': t_total,
                'sent': t_sent,
                'failed': t_failed,
                'today': t_today,
                'rate': t_rate,
                'total_cost': t_cost,
                'sent_cost': s_cost,
                'failed_cost': f_cost,
                'cost_total': t_cost,
                'cost_sent': s_cost,
                'cost_failed': f_cost,
            }

        # (a) One row per team_admin (admin + their members).
        team_admins = db.execute(
            "SELECT id, username, full_name FROM users WHERE role='team_admin' ORDER BY id"
        ).fetchall()
        for ta in team_admins:
            member_ids = [r['id'] for r in db.execute(
                "SELECT id FROM users WHERE id = ? OR team_creator_id = ?",
                (ta['id'], ta['id'])
            ).fetchall()]
            all_teams_list.append(_build_unit_row(
                ta['full_name'] or ta['username'], ta['username'], member_ids
            ))

        # (b) System admin as its own team unit (only the admin itself).
        admin_row = db.execute(
            "SELECT id, username, full_name FROM users WHERE role='admin' ORDER BY id LIMIT 1"
        ).fetchone()
        if admin_row:
            all_teams_list.append(_build_unit_row(
                admin_row['full_name'] or 'Administrador',
                admin_row['username'], [admin_row['id']]
            ))

        # Sort the combined list by total SMS desc so the biggest team is on top.
        all_teams_list.sort(key=lambda x: x['total'], reverse=True)

        # Summary cards = sum of every row above (teams + admin).
        all_teams_data = {
            'member_count': sum(t['member_count'] for t in all_teams_list),
            'total': sum(t['total'] for t in all_teams_list),
            'sent': sum(t['sent'] for t in all_teams_list),
            'failed': sum(t['failed'] for t in all_teams_list),
            'today': sum(t['today'] for t in all_teams_list),
            'unit_price': unit_price,
            'total_cost': round(sum(t['total_cost'] for t in all_teams_list), 4),
        }
        all_teams_data['rate'] = round(
            (all_teams_data['sent'] / all_teams_data['total'] * 100), 1
        ) if all_teams_data['total'] > 0 else 0
        # The number of "teams" shown in the badge is the number of team_admin
        # groups (the admin row is not counted as a team).
        all_teams_data['team_count'] = len(team_admins)

    # Get users list for filter dropdown
    users_list = []
    if role == 'admin':
        users_list = db.execute("SELECT id, username, full_name FROM users WHERE role IN ('team_admin','team_member') ORDER BY username").fetchall()
    elif role == 'team_admin':
        users_list = db.execute("SELECT id, username, full_name FROM users WHERE id = ? OR team_creator_id = ? ORDER BY username", (user_id, user_id)).fetchall()
    else:
        users_list = [{'id': user_id, 'username': session.get('username', ''), 'full_name': ''}]

    # Convert to dicts
    users_dicts = [{'id': u['id'], 'username': u['username'], 'full_name': u['full_name']} for u in users_list]

    # Precio unitario configurado por el administrador (coste por SMS enviado).
    unit_price = get_sms_unit_price(db)

    # Costes a nivel de cuenta personal.
    my_account_data['total_cost'] = calc_cost(my_account_data.get('total'), unit_price)
    my_account_data['sent_cost'] = calc_cost(my_account_data.get('sent'), unit_price)
    my_account_data['failed_cost'] = calc_cost(my_account_data.get('failed'), unit_price)

    # Costes del equipo del team_admin.
    if my_team_data:
        my_team_data['total_cost'] = calc_cost(my_team_data.get('total'), unit_price)
        my_team_data['sent_cost'] = calc_cost(my_team_data.get('sent'), unit_price)
        my_team_data['failed_cost'] = calc_cost(my_team_data.get('failed'), unit_price)

    # Cost fields are computed inside _build_unit_row for all_teams_list and
    # all_teams_data already includes the summed total_cost. Nothing extra to do.

    return jsonify({
        'my_account': my_account_data,
        'my_team': my_team_data,
        'all_teams': all_teams_data,
        'all_teams_list': all_teams_list,
        'users': users_dicts,
        'unit_price': unit_price
    })

@app.route('/api/admin/export-teams', methods=['GET'])
@login_required
def export_teams_stats():
    """Exporta la estadistica por equipos a Excel (solo administrador)."""
    if session.get('role') != 'admin':
        return jsonify({'error': 'No autorizado'}), 403

    db = get_db()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    is_pg = db.db_type == 'postgres'

    date_filter_r = ''
    date_params = []
    if date_from:
        date_filter_r += " AND date(r.created_at) >= ?"
        date_params.append(date_from)
    if date_to:
        date_filter_r += " AND date(r.created_at) <= ?"
        date_params.append(date_to)

    unit_price = get_sms_unit_price(db)
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'openpyxl no esta instalado en el servidor'}), 500

    # Desglose por equipo (team_admin groups + admin as its own unit).
    team_rows = db.execute(f"""
        SELECT
            u.id as unit_id,
            u.username as unit_username,
            u.full_name as unit_name,
            u.role as unit_role,
            COUNT(DISTINCT m.id) + 1 as member_count,
            COALESCE(SUM(CASE WHEN r.status='sent' THEN 1 ELSE 0 END), 0) as sent,
            COALESCE(SUM(CASE WHEN r.status='failed' THEN 1 ELSE 0 END), 0) as failed,
            COALESCE(COUNT(r.id), 0) as total
        FROM users u
        LEFT JOIN users m ON (u.role='team_admin' AND m.team_creator_id = u.id)
        LEFT JOIN sms_records r ON (r.created_by = u.id OR r.created_by = m.id) {date_filter_r}
        WHERE u.role IN ('team_admin','admin')
        GROUP BY u.id
        ORDER BY total DESC
    """, date_params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipos"

    headers = [
        "Equipo", "Administrador", "Miembros", "Total SMS",
        "Enviados", "Fallidos", "Costo", "Hoy", "Exito %",
        "Costo enviados", "Costo fallidos", "Precio unitario"
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    today_expr = "date('now')" if not is_pg else "CURRENT_DATE"

    for tr in team_rows:
        team_total = tr['total'] or 0
        team_sent = tr['sent'] or 0
        team_failed = tr['failed'] or 0
        team_rate = round((team_sent / team_total * 100), 1) if team_total > 0 else 0
        unit_id = tr['unit_id']
        today_row = db.execute(f"""
            SELECT COUNT(*) as cnt FROM sms_records
            WHERE created_by IN (
                SELECT id FROM users WHERE id = ?
                UNION
                SELECT id FROM users WHERE team_creator_id = ?
            )
            AND date(created_at) = {today_expr}
        """, (unit_id, unit_id)).fetchone()
        team_today = today_row['cnt'] if today_row else 0
        if tr['unit_role'] == 'admin':
            display_name = "Administrador"
            label = "Cuenta admin"
        else:
            display_name = tr['unit_name'] or tr['unit_username']
            label = tr['unit_username']
        ws.append([
            display_name, label, tr['member_count'], team_total,
            team_sent, team_failed, calc_cost(team_total, unit_price),
            team_today, f"{team_rate}%",
            calc_cost(team_sent, unit_price),
            calc_cost(team_failed, unit_price), unit_price
        ])

    widths = [22, 20, 12, 12, 12, 12, 12, 10, 10, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    period = ""
    if date_from or date_to:
        period = f"_{date_from or 'inicio'}_{date_to or 'fin'}"
    filename = f"estadisticas_equipos{period}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/admin/team-daily-stats', methods=['GET'])
@login_required
def get_team_daily_stats():
    """Get daily SMS statistics for the team (last 30 days)"""
    db = get_db()
    user_id = session['user_id']
    role = session['role']
    user = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()

    if role == 'admin':
        date_filter = ""
        params = ()
    elif role == 'team_admin':
        member_ids = [u['id'] for u in db.execute(
            "SELECT id FROM users WHERE team_creator_id = ?", (user['id'],)
        ).fetchall()]
        member_ids.append(user['id'])
        placeholders = ','.join('?' * len(member_ids))
        date_filter = f" AND created_by IN ({placeholders})"
        params = tuple(member_ids)
    else:
        date_filter = " AND created_by = ?"
        params = (user['id'],)

    # Get last 30 days daily stats
    rows = db.execute(f"""
        SELECT
            DATE(created_at) as day,
            COUNT(*) as total,
            SUM(CASE WHEN status='sent' THEN 1 ELSE 0 END) as sent,
            SUM(CASE WHEN status='failed' THEN 1 ELSE 0 END) as failed
        FROM sms_records
        WHERE created_at >= date('now', '-30 days') {date_filter}
        GROUP BY DATE(created_at)
        ORDER BY day ASC
    """, params).fetchall()

    # Fill in missing days with zeros
    from datetime import datetime, timedelta
    stats = {}
    for row in rows:
        stats[row['day']] = {'total': row['total'], 'sent': row['sent'], 'failed': row['failed']}

    result = []
    for i in range(29, -1, -1):
        day = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        result.append({
            'day': day,
            'total': stats.get(day, {}).get('total', 0),
            'sent': stats.get(day, {}).get('sent', 0),
            'failed': stats.get(day, {}).get('failed', 0)
        })

    # Compute summary totals
    total_all = sum(r['total'] for r in result)
    sent_all = sum(r['sent'] for r in result)
    today_str = datetime.now().strftime('%Y-%m-%d')
    today_all = stats.get(today_str, {}).get('total', 0)

    return jsonify({'daily_stats': result, 'total': total_all, 'today': today_all, 'sent': sent_all})

# ==================== Estadisticas por Equipo (Admin) ====================

@app.route('/api/admin/team-stats', methods=['GET'])
@login_required
def get_team_stats():
    """Get aggregated SMS statistics grouped by team (for system admin only)"""
    if session['role'] != 'admin':
        return jsonify({'error': 'Acceso denegado'}), 403

    db = get_db()

    # Get all team admins with their team info
    team_admins = db.execute("""
        SELECT u.id, u.username, u.full_name, u.is_active, u.created_at,
               (SELECT COUNT(*) FROM users WHERE team_creator_id = u.id) as member_count,
               (SELECT daily_sms_limit FROM team_config WHERE team_admin_id = u.id) as daily_limit
        FROM users u
        WHERE u.role = 'team_admin'
        ORDER BY u.username ASC
    """).fetchall()

    teams = []
    for ta in team_admins:
        ta_id = ta['id']
        # Get member IDs (team admin + their members)
        member_ids = [u['id'] for u in db.execute(
            "SELECT id FROM users WHERE team_creator_id = ? UNION SELECT ? as id",
            (ta_id, ta_id)
        ).fetchall()]
        placeholders = ','.join('?' * len(member_ids))

        # Total SMS stats for this team
        total_row = db.execute(f"""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN status='sent' THEN 1 ELSE 0 END) as sent,
                SUM(CASE WHEN status='failed' THEN 1 ELSE 0 END) as failed,
                SUM(CASE WHEN status='pending' THEN 1 ELSE 0 END) as pending
            FROM sms_records
            WHERE created_by IN ({placeholders})
        """, member_ids).fetchone()

        # Today's SMS count
        today_row = db.execute(f"""
            SELECT COUNT(*) as cnt
            FROM sms_records
            WHERE created_by IN ({placeholders})
              AND DATE(created_at) = DATE('now')
        """, member_ids).fetchone()

        # Last 30 days daily breakdown for chart
        daily_rows = db.execute(f"""
            SELECT
                DATE(created_at) as day,
                COUNT(*) as total,
                SUM(CASE WHEN status='sent' THEN 1 ELSE 0 END) as sent
            FROM sms_records
            WHERE created_by IN ({placeholders})
              AND created_at >= date('now', '-30 days')
            GROUP BY DATE(created_at)
            ORDER BY day ASC
        """, member_ids).fetchall()

        daily_stats = {}
        for dr in daily_rows:
            daily_stats[dr['day']] = {'total': dr['total'], 'sent': dr['sent']}

        from datetime import datetime, timedelta
        chart_data = []
        for i in range(29, -1, -1):
            day = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
            chart_data.append({
                'day': day,
                'total': daily_stats.get(day, {}).get('total', 0),
                'sent': daily_stats.get(day, {}).get('sent', 0)
            })

        total_sms = total_row['total'] if total_row and total_row['total'] else 0
        sent_sms = total_row['sent'] if total_row and total_row['sent'] else 0
        success_rate = round(sent_sms / total_sms * 100, 1) if total_sms > 0 else 0

        teams.append({
            'team_admin_id': ta_id,
            'team_name': ta['full_name'] or ta['username'],
            'username': ta['username'],
            'is_active': ta['is_active'],
            'member_count': ta['member_count'],
            'daily_limit': ta['daily_limit'] if ta['daily_limit'] else 0,
            'total_sms': total_sms,
            'sent_sms': sent_sms,
            'failed_sms': total_row['failed'] if total_row and total_row['failed'] else 0,
            'pending_sms': total_row['pending'] if total_row and total_row['pending'] else 0,
            'today_sms': today_row['cnt'] if today_row else 0,
            'success_rate': success_rate,
            'daily_chart': chart_data
        })

    # Grand totals
    grand_total = sum(t['total_sms'] for t in teams)
    grand_sent = sum(t['sent_sms'] for t in teams)
    grand_today = sum(t['today_sms'] for t in teams)
    grand_members = sum(t['member_count'] for t in teams) + len(teams)

    return jsonify({
        'teams': teams,
        'summary': {
            'total_teams': len(teams),
            'total_members': grand_members,
            'total_sms': grand_total,
            'today_sms': grand_today,
            'sent_sms': grand_sent,
            'success_rate': round(grand_sent / grand_total * 100, 1) if grand_total > 0 else 0
        }
    })

# ==================== Configuracion de limite diario ====================

@app.route('/api/config/daily-limit', methods=['GET'])
@manager_required
def get_daily_limit():
    """Obtener el limite diario de SMS del equipo actual"""
    db = get_db()
    user = g.user
    if user['role'] == 'admin':
        # Admin: ver limite global (primer registro sin team_admin_id especifico)
        row = db.execute(
            "SELECT daily_sms_limit FROM team_config WHERE team_admin_id=0"
        ).fetchone()
        return jsonify({'daily_limit': row['daily_sms_limit'] if row else 100})
    else:
        # Team admin: ver limite de su equipo
        row = db.execute(
            "SELECT daily_sms_limit FROM team_config WHERE team_admin_id=?",
            (user['id'],)
        ).fetchone()
        return jsonify({'daily_limit': row['daily_sms_limit'] if row else 100})

@app.route('/api/config/daily-limit', methods=['POST'])
@manager_required
def set_daily_limit():
    """Establecer el limite diario de SMS"""
    db = get_db()
    user = g.user
    data = request.get_json()
    limit = data.get('limit')

    if limit is None or not isinstance(limit, int) or limit < 0:
        return jsonify({'error': 'Proporcione un valor valido (entero no negativo)'}), 400

    if user['role'] == 'admin':
        team_admin_id = 0  # global config
    else:
        team_admin_id = user['id']

    existing = db.execute(
        "SELECT id FROM team_config WHERE team_admin_id=?",
        (team_admin_id,)
    ).fetchone()
    if existing:
        db.execute(
            "UPDATE team_config SET daily_sms_limit=? WHERE team_admin_id=?",
            (limit, team_admin_id)
        )
    else:
        db.execute(
            "INSERT INTO team_config (team_admin_id, daily_sms_limit) VALUES (?, ?)",
            (team_admin_id, limit)
        )
    db.commit()
    return jsonify({'message': 'Limite diario actualizado', 'daily_limit': limit})

@app.route('/api/config/team-api-config', methods=['GET'])
@admin_required
def get_team_api_configs():
    """Get all teams with their API config assignments."""
    db = get_db()
    teams = db.execute('''
        SELECT tc.id, tc.team_admin_id, tc.api_config_id, tc.daily_sms_limit,
               u.username as team_admin_name, u.full_name as team_admin_full_name,
               sac.name as api_config_name, sac.country as api_config_country
        FROM team_config tc
        JOIN users u ON tc.team_admin_id = u.id
        LEFT JOIN sms_api_configs sac ON tc.api_config_id = sac.id
        ORDER BY u.username
    ''').fetchall()
    team_list = []
    for t in teams:
        team_list.append({
            'id': t['id'],
            'team_admin_id': t['team_admin_id'],
            'team_admin_name': t['team_admin_name'],
            'team_admin_full_name': t['team_admin_full_name'] or '',
            'api_config_id': t['api_config_id'],
            'api_config_name': t['api_config_name'] or 'Sin asignar',
            'api_config_country': t['api_config_country'] or '',
            'daily_sms_limit': t['daily_sms_limit']
        })
    configs = db.execute("SELECT id, name, country FROM sms_api_configs WHERE is_active=1 ORDER BY name").fetchall()
    config_list = [{'id': c['id'], 'name': c['name'], 'country': c['country']} for c in configs]
    return jsonify({'teams': team_list, 'configs': config_list})

@app.route('/api/config/team-api-config', methods=['PUT'])
@admin_required
def update_team_api_config():
    """Update team's API config assignment."""
    db = get_db()
    data = request.get_json()
    team_admin_id = data.get('team_admin_id')
    api_config_id = data.get('api_config_id')
    if not team_admin_id:
        return jsonify({'error': 'team_admin_id es requerido'}), 400
    existing = db.execute(
        "SELECT id FROM team_config WHERE team_admin_id=?",
        (team_admin_id,)
    ).fetchone()
    if existing:
        db.execute(
            "UPDATE team_config SET api_config_id=? WHERE team_admin_id=?",
            (api_config_id if api_config_id else None, team_admin_id)
        )
    else:
        db.execute(
            "INSERT INTO team_config (team_admin_id, api_config_id, daily_sms_limit) VALUES (?, ?, 100)",
            (team_admin_id, api_config_id if api_config_id else None)
        )
    db.commit()
    return jsonify({'message': 'Configuracion API actualizada'})

@app.route('/api/config/logs', methods=['GET'])
@admin_required
def get_send_logs():
    db = get_db()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    offset = (page - 1) * per_page
    total = db.execute("SELECT COUNT(*) as count FROM send_logs").fetchone()['count']
    logs = db.execute("SELECT * FROM send_logs ORDER BY created_at DESC LIMIT ? OFFSET ?", (per_page, offset)).fetchall()
    return jsonify({
        'logs': [dict(l) for l in logs],
        'total': total,
        'page': page,
        'per_page': per_page
    })

# ============================================================
# Process scheduled messages (simple cron-like)
# ============================================================

def process_scheduled_messages():
    """Process pending scheduled messages using the real SMS API."""
    db = get_db()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    scheduled = db.execute(
        "SELECT * FROM sms_records WHERE status='scheduled' AND scheduled_at <= ?",
        (now,)
    ).fetchall()

    if not scheduled:
        return 0

    api_configured = is_sms_api_configured()
    processed = 0

    if api_configured:
        # Group into batches for efficient sending
        phone_content_pairs = [(r['phone'], r['content']) for r in scheduled]

        for batch_start in range(0, len(phone_content_pairs), 200):
            batch = phone_content_pairs[batch_start:batch_start + 200]
            batch_records = scheduled[batch_start:batch_start + 200]
            result = sms_api_send_batch(batch)
            api_code = result.get('code', -1)

            if api_code == 0 and result.get('data'):
                result_map = {}
                for item in result['data']:
                    result_map[item.get('das', '')] = item

                for record in batch_records:
                    item = result_map.get(record['phone'], {})
                    item_state = item.get('state', 0)
                    msgid = item.get('msgid', '')
                    if item_state == 0:
                        db.execute(
                            "UPDATE sms_records SET status='sent', msgid=?, api_code=?, sent_at=datetime('now') WHERE id=?",
                            (msgid, api_code, record['id'])
                        )
                    else:
                        db.execute(
                            "UPDATE sms_records SET status='failed', api_code=?, api_msg=? WHERE id=?",
                            (item_state, SMS_STATUS_CODES.get(item_state, 'Error'), record['id'])
                        )
                    processed += 1
            else:
                # Entire batch failed
                for record in batch_records:
                    db.execute(
                        "UPDATE sms_records SET status='failed', api_code=?, api_msg=? WHERE id=?",
                        (api_code, SMS_STATUS_CODES.get(api_code, result.get('msg', 'Error')), record['id'])
                    )
                    processed += 1
    else:
        # API not configured - simulate
        for record in scheduled:
            db.execute(
                "UPDATE sms_records SET status='sent', api_msg='API no configurada - envio simulado', sent_at=datetime('now') WHERE id=?",
                (record['id'],)
            )
            processed += 1

    if processed:
        db.commit()
    return processed

@app.route('/api/sms/process-scheduled', methods=['POST'])
@login_required
def trigger_process_scheduled():
    count = process_scheduled_messages()
    return jsonify({'message': f'{count} mensaje(s) programado(s) procesado(s)'})

@app.route('/api/sms/query-status', methods=['POST'])
@login_required
def query_sms_status_endpoint():
    """Query delivery status for sent messages."""
    data = request.get_json()
    record_ids = data.get('record_ids', [])
    if not record_ids:
        return jsonify({'error': 'IDs de registros son requeridos'}), 400

    db = get_db()
    if not is_sms_api_configured():
        return jsonify({'error': 'API SMS no configurada'}), 400

    # Get msgids from records
    placeholders = ','.join(['?'] * len(record_ids))
    records = db.execute(
        f"SELECT id, msgid, phone FROM sms_records WHERE id IN ({placeholders}) AND msgid != ''",
        record_ids
    ).fetchall()

    if not records:
        return jsonify({'error': 'No se encontraron mensajes con ID de API'}), 404

    msgids = [r['msgid'] for r in records]
    result = sms_api_query_status(msgids)
    api_code = result.get('code', -1)

    if api_code == 0 and result.get('data'):
        # Update records with status
        state_map = {0: 'pending', 1: 'sent', 2: 'failed'}
        state_names = {0: 'Sin回执 (Enviado)', 1: 'Entregado', 2: 'Fallido'}
        updates = []
        for item in result['data']:
            msgid = item.get('msgid', '')
            state = item.get('state', 0)
            status = state_map.get(state, 'pending')
            db.execute(
                "UPDATE sms_records SET status=? WHERE msgid=?",
                (status, msgid)
            )
            updates.append({'msgid': msgid, 'state': state, 'state_name': state_names.get(state, 'Desconocido')})
        db.commit()
        return jsonify({'message': f'{len(updates)} estado(s) actualizado(s)', 'updates': updates})
    else:
        error_msg = SMS_STATUS_CODES.get(api_code, result.get('msg', 'Error'))
        return jsonify({'error': f'Error al consultar estado (codigo {api_code}): {error_msg}'}), 400

@app.route('/api/sms/check-charset', methods=['POST'])
@login_required
def check_sms_charset():
    """Check charset and billing info for SMS content."""
    data = request.get_json()
    content = data.get('content', '').strip()
    if not content:
        return jsonify({'error': 'Contenido es requerido'}), 400

    if not is_sms_api_configured():
        # Return local estimation
        # Spanish uses UCS2: 70 chars per SMS, 67 per long SMS part
        char_count = len(content)
        if char_count <= 70:
            parts = 1
            single = 70
        else:
            parts = (char_count + 66) // 67  # ceil division
            single = 67
        return jsonify({
            'charset': 'UCS2',
            'parts': parts,
            'single': single,
            'char_count': char_count,
            'detail': f'Contenido con {char_count} caracteres. Codificacion UCS2 (Espanol). Se factura como {parts} SMS.',
            'api_configured': False
        })

    result = sms_api_check_charset(content)
    api_code = result.get('code', -1)
    if api_code == 0:
        data = result.get('data', {})
        data['api_configured'] = True
        return jsonify(data)
    else:
        error_msg = SMS_STATUS_CODES.get(api_code, result.get('msg', 'Error'))
        return jsonify({'error': f'Error al verificar codificacion (codigo {api_code}): {error_msg}'}), 400

# ============================================================
# Daily auto-clear contacts (background thread)
# ============================================================
AUTO_CLEAR_LOCK_ID = 82347109


def _parse_hhmm(value):
    if not value or not isinstance(value, str):
        return None
    m = re.match(r'^(\d{1,2}):(\d{2})$', value.strip())
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    if 0 <= h <= 23 and 0 <= mi <= 59:
        return h, mi
    return None


def _setting_get(key, default=''):
    db = get_db()
    cur = db.execute("SELECT value FROM system_settings WHERE key=?", (key,))
    row = cur.fetchone()
    if not row:
        return default
    return row['value'] if isinstance(row, dict) else row[0]


def _setting_set(key, value):
    db = get_db()
    if get_db_type() == 'postgres':
        db.execute(
            """
            INSERT INTO system_settings (key, value, updated_at)
            VALUES (%s, %s, NOW())
            ON CONFLICT (key) DO UPDATE
            SET value=EXCLUDED.value, updated_at=NOW()
            """,
            (key, str(value)),
        )
    else:
        db.execute(
            """
            INSERT INTO system_settings (key, value, updated_at)
            VALUES (?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(key) DO UPDATE
            SET value=excluded.value, updated_at=CURRENT_TIMESTAMP
            """,
            (key, str(value)),
        )
    db.commit()


def get_auto_clear_config():
    return {
        'enabled': str(_setting_get('auto_clear_enabled', 'true')).lower() in ('1', 'true', 'yes', 'on'),
        'time': _setting_get('auto_clear_time', '03:00'),
        'last_run_at': _setting_get('auto_clear_last_run_at', ''),
        'last_run_count': int(_setting_get('auto_clear_last_run_count', '0') or 0),
        'last_run_status': _setting_get('auto_clear_last_run_status', ''),
    }


def set_auto_clear_config(enabled=None, time_str=None):
    if time_str is not None:
        parsed = _parse_hhmm(time_str)
        if parsed is None:
            raise ValueError('Hora invalida, use HH:MM')
        _setting_set('auto_clear_time', f'{parsed[0]:02d}:{parsed[1]:02d}')
    if enabled is not None:
        _setting_set('auto_clear_enabled', 'true' if enabled else 'false')


def run_auto_clear_contacts(triggered_by='scheduler'):
    """Delete contacts older than the retention window defined by the owning
    user's category (user_categories.retention_days). retention_days=0 or NULL
    means 'retain forever'. Contacts without a creator or whose creator has no
    category are retained. Groups are NEVER deleted (replaces the previous
    full-wipe behavior). SMS/voice records are always retained."""
    db = get_db()
    lock_acquired = False
    try:
        if get_db_type() == 'postgres':
            cur = db.execute("SELECT pg_try_advisory_lock(%s) AS got", (AUTO_CLEAR_LOCK_ID,))
            row = cur.fetchone()
            got = row['got'] if isinstance(row, dict) else row[0]
            lock_acquired = bool(got)
        else:
            db.execute("CREATE TABLE IF NOT EXISTS maintenance_locks (lock_name TEXT PRIMARY KEY, locked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
            db.execute("DELETE FROM maintenance_locks WHERE lock_name='auto_clear_contacts' AND locked_at < datetime('now', '-1 day')")
            try:
                db.execute("INSERT INTO maintenance_locks(lock_name) VALUES('auto_clear_contacts')")
                db.commit()
                lock_acquired = True
            except Exception:
                lock_acquired = False

        if not lock_acquired:
            return 0, 'Ya se esta ejecutando una limpieza'

        is_pg = get_db_type() == 'postgres'
        if is_pg:
            # Delete contacts whose owning user belongs to a category with a
            # finite retention window and whose created_at is older than that
            # window. NOW() - (days || ' days')::interval handles the cutoff.
            count_cur = db.execute(
                """
                SELECT COUNT(*) AS c FROM contacts ct
                WHERE EXISTS (
                    SELECT 1 FROM users u
                    JOIN user_categories cat ON u.category_id = cat.id
                    WHERE u.id = ct.created_by
                      AND cat.retention_days > 0
                      AND ct.created_at < NOW() - (cat.retention_days::text || ' days')::interval
                )
                """
            )
            count_row = count_cur.fetchone()
            count = int((count_row['c'] if isinstance(count_row, dict) else count_row[0]) if count_row else 0)
            db.execute(
                """
                DELETE FROM contacts ct
                WHERE EXISTS (
                    SELECT 1 FROM users u
                    JOIN user_categories cat ON u.category_id = cat.id
                    WHERE u.id = ct.created_by
                      AND cat.retention_days > 0
                      AND ct.created_at < NOW() - (cat.retention_days::text || ' days')::interval
                )
                """
            )
        else:
            # SQLite: datetime('now', '-N days') per row via the joined days.
            # Note: SQLite does NOT allow an alias on the DELETE target table.
            count_cur = db.execute(
                """
                SELECT COUNT(*) AS c FROM contacts
                WHERE EXISTS (
                    SELECT 1 FROM users u
                    JOIN user_categories cat ON u.category_id = cat.id
                    WHERE u.id = contacts.created_by
                      AND cat.retention_days > 0
                      AND contacts.created_at < datetime('now', '-' || cat.retention_days || ' days')
                )
                """
            )
            count_row = count_cur.fetchone()
            count = int((count_row['c'] if isinstance(count_row, dict) else count_row[0]) if count_row else 0)
            db.execute(
                """
                DELETE FROM contacts
                WHERE EXISTS (
                    SELECT 1 FROM users u
                    JOIN user_categories cat ON u.category_id = cat.id
                    WHERE u.id = contacts.created_by
                      AND cat.retention_days > 0
                      AND contacts.created_at < datetime('now', '-' || cat.retention_days || ' days')
                )
                """
            )
        db.commit()

        now = datetime.utcnow().isoformat()
        _setting_set('auto_clear_last_run_at', now)
        _setting_set('auto_clear_last_run_count', str(count))
        _setting_set('auto_clear_last_run_status', 'ok')

        details = f"Contactos expirados eliminados segun retencion por categoria: {count}; grupos conservados; disparado por: {triggered_by}"
        db.execute(
            "INSERT INTO send_logs (action, status, details) VALUES (?, ?, ?)",
            ('auto_clear_contacts', 'ok', details),
        )
        db.commit()
        app.logger.warning('auto_clear_contacts completed: %s', details)
        return count, details
    except Exception as exc:
        db.rollback()
        try:
            _setting_set('auto_clear_last_run_status', f'error: {str(exc)[:200]}')
            db.execute(
                "INSERT INTO send_logs (action, status, details) VALUES (?, ?, ?)",
                ('auto_clear_contacts', 'error', f'Error: {str(exc)[:400]}'),
            )
            db.commit()
        except Exception:
            pass
        app.logger.exception('auto_clear_contacts failed')
        raise
    finally:
        try:
            if get_db_type() == 'postgres' and lock_acquired:
                db.execute("SELECT pg_advisory_unlock(%s)", (AUTO_CLEAR_LOCK_ID,))
                db.commit()
            elif lock_acquired:
                db.execute("DELETE FROM maintenance_locks WHERE lock_name='auto_clear_contacts'")
                db.commit()
        except Exception:
            pass


def _auto_clear_loop():
    app.logger.info('[auto-clear] background loop started')
    time.sleep(15)
    last_run_date = None
    while True:
        try:
            with app.app_context():
                cfg = get_auto_clear_config()
                now = datetime.now()
                if cfg['enabled']:
                    h, m = _parse_hhmm(cfg['time']) or (3, 0)
                    today = now.date().isoformat()
                    if (now.hour > h or (now.hour == h and now.minute >= m)) and last_run_date != today:
                        try:
                            run_auto_clear_contacts(triggered_by='scheduler')
                        except Exception as exc:
                            app.logger.warning('[auto-clear] run failed: %s', exc)
                        last_run_date = today
        except Exception as exc:
            app.logger.warning('[auto-clear] loop error: %s', exc)
        time.sleep(30)


@app.route('/api/config/auto-clear', methods=['GET'])
@login_required
@admin_required
def api_auto_clear_config():
    return jsonify(get_auto_clear_config())


@app.route('/api/config/auto-clear', methods=['PUT'])
@login_required
@admin_required
def api_auto_clear_update():
    data = request.get_json() or {}
    enabled = data.get('enabled')
    time_str = data.get('time')
    if enabled is None and time_str is None:
        return jsonify({'error': 'No hay cambios'}), 400
    try:
        set_auto_clear_config(enabled=enabled, time_str=time_str)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    return jsonify(get_auto_clear_config())


@app.route('/api/config/auto-clear/run-now', methods=['POST'])
@login_required
@admin_required
def api_auto_clear_run_now():
    try:
        count, details = run_auto_clear_contacts(triggered_by=f'admin:{session.get("user_id")}')
        return jsonify({'success': True, 'deleted': count, 'details': details})
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


# ============================================================
# Voice Call (电呼) Provider Integration
# ============================================================

VOICE_STATUS_LABELS = {
    'pending': 'Pendiente',
    'initiated': 'Iniciada',
    'ringing': 'Llamando',
    'answered': 'Contestada',
    'completed': 'Completada',
    'failed': 'Fallida',
    'no-answer': 'Sin respuesta',
    'busy': 'Ocupado',
    'canceled': 'Cancelada',
}


def get_voice_config(country=None):
    """Return a voice config row as a dict.

    Voice configs are stored one row per country (like sms_api_configs). When
    a country is given (mx/co/pe) the matching row is returned; for an
    unknown/empty country the first active row is used as a generic default.
    Returns None when no row exists.
    """
    db = get_db()
    cc = normalize_country(country)
    row = None
    if cc:
        row = db.execute(
            "SELECT * FROM voice_configs WHERE country = ? ORDER BY id LIMIT 1", (cc,)
        ).fetchone()
    if not row:
        row = db.execute(
            "SELECT * FROM voice_configs WHERE is_active = 1 ORDER BY id LIMIT 1"
        ).fetchone()
    if not row:
        return None
    d = dict(row)
    return {
        'id': d.get('id'),
        'name': d.get('name') or '',
        'country': d.get('country') or cc or '',
        'provider': (d.get('provider') or 'simulation').strip().lower(),
        'api_domain': d.get('api_domain') or '',
        'voice_appid': d.get('voice_appid') or '',
        'voice_accesskey': d.get('voice_accesskey') or '',
        'from_number': d.get('from_number') or '',
        'voice_token': d.get('voice_token') or '',
        'voice_token_expiry': int(d.get('voice_token_expiry') or 0),
        'is_active': bool(d.get('is_active')),
        'updated_at': d.get('updated_at'),
        '_country': d.get('country') or cc or '',
    }


def list_voice_configs():
    """Return all voice config rows (one per country) for the admin UI."""
    db = get_db()
    rows = db.execute("SELECT * FROM voice_configs ORDER BY id").fetchall()
    out = []
    for d in rows:
        d = dict(d)
        out.append({
            'id': d.get('id'),
            'name': d.get('name') or '',
            'country': d.get('country') or '',
            'provider': (d.get('provider') or 'simulation').strip().lower(),
            'api_domain': d.get('api_domain') or '',
            'voice_appid': d.get('voice_appid') or '',
            'from_number': d.get('from_number') or '',
            'has_accesskey': bool(d.get('voice_accesskey')),
            'is_active': bool(d.get('is_active')),
            'configured': bool(
                (d.get('provider') or '').strip().lower() == 'infin8linx'
                and (d.get('api_domain') or '') and (d.get('voice_appid') or '')
                and (d.get('voice_accesskey') or '')
            ),
            'updated_at': d.get('updated_at'),
        })
    return out


def resolve_voice_config(country=None):
    """Return the voice config to use for a call from an agent of `country`.

    Mirrors the SMS resolution: pick the row for the agent's country; if the
    agent has no country, fall back to the first active config. The returned
    dict already carries the generic keys (api_domain/voice_appid/...) used by
    the infin8linx_* helpers.
    """
    return get_voice_config(country)


def is_voice_configured(country=None):
    cfg = resolve_voice_config(country)
    if not cfg:
        return False
    if cfg['provider'] in ('', 'simulation', 'simulacion', 'none'):
        return False
    if cfg['provider'] == 'infin8linx':
        return bool(cfg['api_domain'] and cfg['voice_appid'] and cfg['voice_accesskey'])
    return False


def infin8linx_parse_response(resp):
    """Parse a standard infin8linx JSON envelope. Returns (ok, data, error_msg)."""
    if resp.status_code >= 300:
        return False, None, 'El servidor respondio HTTP %s (revisar URL/puerto)' % resp.status_code
    try:
        data = resp.json()
    except Exception:
        return False, None, 'Respuesta no JSON del servidor (revisar que la URL apunte al API de voz)'
    if int(data.get('ret', 0)) != 200:
        return False, data, str(data.get('msg') or 'Error ret=%s' % data.get('ret'))[:300]
    body = data.get('data') or {}
    if int(body.get('status', 1)) != 0:
        errs = body.get('errors') or {}
        return False, data, str(body.get('desc') or errs.get('codemsg') or 'Error en la operacion')[:300]
    return True, data, ''


def _twilio_tts_xml(script_text, lang='es-MX', voice='alice'):
    """Build TwiML <Say> response with the call script."""
    safe = (script_text or '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Response><Say language="' + lang + '" voice="' + voice + '">'
        + safe + '</Say></Response>'
    )


# in-memory cache for Infinity auth tokens (one per gunicorn worker, per config row)
# Keyed by voice_configs.id so each country's token is independent.
_INFIN_TOKEN_CACHE = {}


def _token_cache_key(config):
    return 'vcfg:%s' % ((config or {}).get('id') or (config or {}).get('_country') or 'default')


def _token_db_columns(config):
    """Voice configs store the token directly on the config row."""
    return 'voice_token', 'voice_token_expiry'


def _voice_api_url(config, scheme=None):
    """Build the Infinity base URL. Honors a detected/stored scheme
    (config['voice_scheme'] = 'http'|'https') so the test connection can
    persist whichever protocol the provider actually answers on.
    """
    raw = ((config or {}).get('api_domain') or '').strip()
    if not raw:
        return ''
    raw = re.sub(r'^https?://', '', raw)
    sch = scheme or (config or {}).get('voice_scheme') or 'https'
    return (sch or 'https').lower() + '://' + raw.rstrip('/')


def _voice_probe_host(config, appid, accesskey, timeout=10):
    """Try the Infinity Login endpoint over https then http.

    Returns (working_scheme, response_or_none, error_message). Used by the
    connection test to distinguish a wrong protocol (http vs https) from a
    network/firewall block.
    """
    raw = ((config or {}).get('api_domain') or '').strip()
    if not raw:
        return None, None, 'URL de la API vacia'
    host = re.sub(r'^https?://', '', raw).rstrip('/')
    last_err = ''
    for scheme in ('https', 'http'):
        url = scheme + '://' + host
        try:
            resp = http_requests.post(url, data={
                'service': 'App.Sip_Auth.Login',
                'appid': appid,
                'accesskey': accesskey,
            }, timeout=timeout)
            return scheme, resp, ''
        except http_requests.exceptions.ConnectTimeout:
            last_err = 'Tiempo de espera agotado al conectar (%s://%s)' % (scheme, host)
        except http_requests.exceptions.ReadTimeout:
            # Connected but the server did not answer in time; treat as reachable
            return scheme, None, 'El servidor respondio lento (%s://%s)' % (scheme, host)
        except http_requests.exceptions.SSLError as e:
            last_err = 'Error SSL/TLS en %s://%s: %s' % (scheme, host, str(e)[:120])
        except http_requests.exceptions.ConnectionError as e:
            last_err = 'No se pudo conectar a %s://%s: %s' % (scheme, host, str(e)[:120])
        except Exception as e:  # noqa: BLE001
            last_err = '%s en %s://%s' % (type(e).__name__, scheme, host)
    return None, None, (last_err or 'No se pudo conectar al servidor')


def _parse_infin_login_response(resp):
    """Parse an App.Sip_Auth.Login response. Returns (token, error)."""
    try:
        data = resp.json()
    except Exception:
        return '', 'Respuesta no JSON del servidor (revisar que la URL apunte al API de voz)'
    if resp.status_code >= 300 or int(data.get('ret', 0)) != 200:
        return '', str(data.get('msg') or ('HTTP %s - revisar URL/puerto' % resp.status_code))[:300]
    body = data.get('data') or {}
    if int(body.get('status', 1)) != 0:
        errs = body.get('errors') or {}
        return '', str(body.get('desc') or errs.get('codemsg') or 'No se pudo obtener el token')[:300]
    result = body.get('result') or {}
    token = (result.get('token') or '').strip()
    if not token:
        return '', 'El servidor no devolvio token'
    return token, ''


def _persist_voice_scheme_and_token(config, scheme, token, expiry):
    _INFIN_TOKEN_CACHE[_token_cache_key(config)] = {'token': token, 'expiry': expiry}
    try:
        db = get_db()
        tok_col, exp_col = _token_db_columns(config)
        db.execute(
            "UPDATE voice_configs SET voice_scheme=?, %s=?, %s=? WHERE id=?" % (tok_col, exp_col),
            (scheme, token, expiry, config.get('id')))
        db.commit()
    except Exception:
        pass


def infin8linx_get_token(config, force=False):
    """Obtain an auth token from Infinity (service App.Sip_Auth.Login).

    Tokens last 12h per docs; we refresh 10 minutes early. The token is also
    persisted on the voice_configs row so multiple workers share it.
    """
    import time as _time
    now = int(time.time())
    cache_key = _token_cache_key(config)
    if not force:
        # Prefer the value freshly read from DB (shared across workers)
        db_token = (config or {}).get('voice_token') or ''
        db_expiry = int((config or {}).get('voice_token_expiry') or 0)
        if db_token and db_expiry - 600 > now:
            return db_token, None
        entry = _INFIN_TOKEN_CACHE.get(cache_key) or {}
        cached = entry.get('token') or ''
        if cached and int(entry.get('expiry') or 0) - 600 > now:
            return cached, None
    api_url = _voice_api_url(config or {})
    appid = (config or {}).get('voice_appid') or ''
    accesskey = (config or {}).get('voice_accesskey') or ''
    if not (api_url and appid and accesskey):
        return '', 'Faltan datos de acceso (URL, AppID o AccessKey)'

    # First attempt with the stored (or default https) scheme.
    try:
        resp = http_requests.post(api_url, data={
            'service': 'App.Sip_Auth.Login',
            'appid': appid,
            'accesskey': accesskey,
        }, timeout=15)
        token, err = _parse_infin_login_response(resp)
        if token:
            expiry = now + 12 * 3600
            _persist_voice_scheme_and_token(
                config, (config or {}).get('voice_scheme') or 'https', token, expiry)
            return token, None
        # Business-level error (bad credentials etc.) — re-probing schemes
        # won't help; surface the provider message.
        if err and not err.startswith('Respuesta no JSON'):
            return '', err
    except Exception as e:
        msg = str(e)
        # If the stored scheme cannot even connect / fails SSL (e.g. config
        # saved as https but the gateway speaks plain http like port 4434),
        # fall through to probing both schemes and persist the one that works.
        if not ('timed out' in msg.lower() or 'NewConnectionError' in msg
                or 'Failed to establish a new connection' in msg
                or 'SSLError' in type(e).__name__ or 'wrong version number' in msg.lower()
                or 'ConnectionError' in type(e).__name__):
            return '', 'Error de conexion: ' + msg[:200]

    # Auto-detect http/https and persist the working scheme (self-healing).
    scheme, resp, err = _voice_probe_host(config or {}, appid, accesskey, timeout=10)
    if not scheme or resp is None:
        return '', err or 'No se pudo conectar con la URL de la API (revisar IP/puerto/firewall)'
    token, perr = _parse_infin_login_response(resp)
    if not token:
        return '', perr or err or 'No se pudo obtener el token'
    expiry = now + 12 * 3600
    _persist_voice_scheme_and_token(config, scheme, token, expiry)
    return token, None


def _parse_extension_pool(raw):
    """Parse a comma/whitespace separated extension pool into a clean list."""
    if not raw:
        return []
    parts = re.split(r'[,\s;|/]+', str(raw))
    return [p.strip() for p in parts if p.strip()]


def infin8linx_pick_extension(config):
    """Pick a random extension from the configured pool.

    voice_extnumber may hold a single extension or a pool separated by
    commas/spaces/semicolons. Returns the chosen extension or '' if empty.
    """
    import random as _random
    pool = _parse_extension_pool((config or {}).get('voice_extnumber'))
    if not pool:
        return ''
    return _random.choice(pool)


def infin8linx_make_call(phone, config, extnumber=None, forced_ext='', customuuid=''):
    """Send MakeCall command to infin8linx. Returns (ok, call_sid, error_msg, extnumber).

    infin8linx MakeCall triggers a click-to-call between a SIP extension and
    the destination number. It returns only a command-ack (no call id); status
    is later obtained from the CDR/callback interface. We synthesise a local
    reference so the call can be tracked in voice_records.

    Selection order for the extension:
      1. forced_ext: fixed extension of the calling agent (when assigned)
      2. extnumber argument (explicit)
      3. one randomly chosen from the configured pool (voice_extnumber)
    """
    # Resolve the extension FIRST so that, even if the API/token call fails,
    # the record shows which extension would have been used.
    resolved_ext = ''
    if forced_ext and forced_ext.strip():
        resolved_ext = forced_ext.strip()
    elif extnumber:
        resolved_ext = str(extnumber).strip()
    else:
        resolved_ext = infin8linx_pick_extension(config)
    if not resolved_ext:
        return False, '', 'Falta el numero de extension (extnumber)', ''
    token, err = infin8linx_get_token(config)
    if err:
        return False, '', err, resolved_ext
    # Re-read the row so we use the scheme/extension the token call may have
    # just auto-detected and persisted (e.g. plain http on a custom port).
    try:
        db = get_db()
        fresh = db.execute(
            'SELECT * FROM voice_configs WHERE id=?', (config.get('id'),)).fetchone()
        if fresh:
            config = dict(fresh)
    except Exception:
        pass
    extnumber = resolved_ext
    # Infinity expects plain digits (no '+', spaces or separators) for both the
    # destination and the caller id. "+52722..." -> "52722...".
    def _digits(value):
        return ''.join(ch for ch in str(value or '') if ch.isdigit())

    dest_digits = _digits(phone)
    if not dest_digits:
        return False, '', f'Numero de destino invalido: {phone}', extnumber
    disnumber = (config or {}).get('from_number') or ''
    dis_digits = _digits(disnumber)
    payload = {
        'service': 'App.Sip_Call.MakeCall',
        'token': token,
        'extnumber': extnumber,
        'destnumber': dest_digits,
    }
    if dis_digits:
        payload['disnumber'] = dis_digits
    cid = str(customuuid or '').strip()
    if cid:
        payload['customuuid'] = cid

    def _post_makecall(scheme):
        url = _voice_api_url({**(config or {}), 'voice_scheme': scheme})
        return http_requests.post(url, data=payload, timeout=20)

    schemes_to_try = []
    stored_scheme = (config or {}).get('voice_scheme') or 'https'
    schemes_to_try.append(stored_scheme)
    for alt in ('http', 'https'):
        if alt not in schemes_to_try:
            schemes_to_try.append(alt)

    last_err = ''
    for idx, scheme in enumerate(schemes_to_try):
        try:
            resp = _post_makecall(scheme)
            try:
                data = resp.json()
            except Exception:
                # Endpoint replied but not JSON — could be wrong scheme/port;
                # try the alternate scheme before giving up.
                last_err = 'Respuesta no JSON del servidor (revisar URL/puerto)'
                continue
            if int(data.get('ret', 0)) != 200:
                msg = str(data.get('msg') or ('HTTP %s' % resp.status_code))
                # token expired/invalid -> force refresh once (same scheme)
                if idx == 0 and ('600' in str(data.get('ret')) or 'token' in msg.lower()):
                    token2, err2 = infin8linx_get_token(config, force=True)
                    if token2 and not err2:
                        payload['token'] = token2
                        resp = _post_makecall(scheme)
                        data = resp.json()
                        if int(data.get('ret', 0)) == 200:
                            body = data.get('data') or {}
                            if int(body.get('status', 1)) == 0:
                                ref = 'INF' + hashlib.sha1((phone + str(time.time())).encode()).hexdigest()[:16].upper()
                                return True, ref, '', extnumber
                            return False, '', str(body.get('desc') or 'Error al iniciar llamada')[:300], extnumber
                return False, '', msg[:300], extnumber
            body = data.get('data') or {}
            if int(body.get('status', 1)) != 0:
                errs = body.get('errors') or {}
                return False, '', str(body.get('desc') or errs.get('codemsg') or 'Error al iniciar llamada')[:300], extnumber
            # Persist the scheme that worked for future calls.
            if scheme != stored_scheme:
                try:
                    db = get_db()
                    db.execute("UPDATE voice_configs SET voice_scheme=? WHERE id=?",
                               (scheme, config.get('id')))
                    db.commit()
                except Exception:
                    pass
            ref = 'INF' + hashlib.sha1((phone + str(time.time())).encode()).hexdigest()[:16].upper()
            return True, ref, '', extnumber
        except Exception as e:
            msg = str(e)
            if 'NewConnectionError' in msg or 'Failed to establish a new connection' in msg:
                last_err = 'No se pudo conectar con la URL de la API (%s) (revisar IP/puerto)' % scheme
            elif 'timed out' in msg.lower():
                last_err = 'Tiempo de espera agotado al conectar por %s' % scheme
            elif 'wrong version number' in msg.lower() or 'SSLError' in type(e).__name__:
                last_err = 'El puerto habla HTTP plano, no HTTPS (%s)' % scheme
            else:
                last_err = 'Error de conexion (%s): %s' % (scheme, msg[:160])
            continue
    return False, '', last_err or 'No se pudo conectar con la URL de la API', extnumber


def infin8linx_hangup(config, extnumber):
    """Hang up the active call on an Infinity extension.

    Uses App.Sip_Call.HangupCall per the Infinity voice API docs.
    Reuses the cached 12h token and the same https/http auto-detection
    (and persistence) as MakeCall.
    Returns (ok: bool, error_msg: str).
    """
    ext = re.sub(r'[^0-9]', '', str(extnumber or ''))
    if not ext:
        return False, 'No hay extension para colgar'

    token, err = infin8linx_get_token(config)
    if err or not token:
        return False, err or 'No se pudo obtener el token'

    base = _voice_api_url(config)
    stored_scheme = (config.get('voice_scheme') or urlparse(base).scheme or 'https').lower()
    if stored_scheme not in ('https', 'http'):
        stored_scheme = 'https'
    schemes = [stored_scheme] + ([s for s in ('https', 'http') if s != stored_scheme])
    last_err = ''

    def _post(scheme):
        url = base
        if base.startswith('https://') or base.startswith('http://'):
            url = scheme + '://' + base.split('://', 1)[1]
        return requests.post(url, data={
            'service': 'App.Sip_Call.HangupCall',
            'token': token,
            'extnumber': ext,
        }, timeout=15)

    for idx, scheme in enumerate(schemes):
        try:
            resp = _post(scheme)
            if resp.status_code != 200:
                last_err = 'HTTP %s' % resp.status_code
                continue
            try:
                data = resp.json()
            except Exception:
                last_err = 'Respuesta no JSON del servidor (revisar URL/puerto)'
                continue
            if int(data.get('ret', 0)) != 200:
                msg = str(data.get('msg') or ('HTTP %s' % resp.status_code))
                if idx == 0 and ('600' in str(data.get('ret')) or 'token' in msg.lower()):
                    token2, err2 = infin8linx_get_token(config, force=True)
                    if token2 and not err2:
                        resp2 = _post(scheme)
                        try:
                            data2 = resp2.json()
                        except Exception:
                            data2 = {}
                        if int(data2.get('ret', 0)) == 200:
                            body2 = data2.get('data') or {}
                            if int(body2.get('status', 1)) == 0:
                                return True, ''
                            return False, str(body2.get('desc') or 'No se pudo colgar')[:300]
                return False, msg[:300]
            body = data.get('data') or {}
            if int(body.get('status', 1)) != 0:
                return False, str(body.get('desc') or 'No se pudo colgar')[:300]
            if scheme != stored_scheme:
                try:
                    db = get_db()
                    db.execute("UPDATE voice_configs SET voice_scheme=? WHERE id=?",
                               (scheme, config.get('id')))
                    db.commit()
                except Exception:
                    pass
            return True, ''
        except Exception as e:
            msg = str(e)
            if 'NewConnectionError' in msg or 'Failed to establish a new connection' in msg:
                last_err = 'No se pudo conectar con la URL de la API (%s)' % scheme
            elif 'timed out' in msg.lower():
                last_err = 'Tiempo de espera agotado al conectar por %s' % scheme
            elif 'wrong version number' in msg.lower() or 'SSLError' in type(e).__name__:
                last_err = 'El puerto habla HTTP plano, no HTTPS (%s)' % scheme
            else:
                last_err = 'Error de conexion (%s): %s' % (scheme, msg[:160])
            continue
    return False, last_err or 'No se pudo conectar con la URL de la API'


def infin8linx_get_record_url(config, record_file):
    """Fetch a short-lived download URL for a call recording.

    Calls App.Sip_Cdr.GetRecodeFile with the filename received in the CDR
    callback. Reuses token caching and http/https auto-detection.
    Returns (url, error_msg).
    """
    fname = str(record_file or '').strip()
    if not fname:
        return '', 'Nombre de grabacion vacio'
    token, err = infin8linx_get_token(config)
    if err or not token:
        return '', err or 'No se pudo obtener el token'
    base = _voice_api_url(config)
    stored_scheme = (config.get('voice_scheme') or urlparse(base).scheme or 'https').lower()
    if stored_scheme not in ('https', 'http'):
        stored_scheme = 'https'
    schemes = [stored_scheme] + [s for s in ('https', 'http') if s != stored_scheme]
    last_err = ''

    def _post(scheme):
        url = scheme + '://' + base.split('://', 1)[1] if '://' in base else base
        return http_requests.post(url, data={
            'service': 'App.Sip_Cdr.GetRecodeFile',
            'token': token,
            'filename': fname,
        }, timeout=15)

    for idx, scheme in enumerate(schemes):
        try:
            resp = _post(scheme)
            if resp.status_code != 200:
                last_err = 'HTTP %s' % resp.status_code
                continue
            try:
                data = resp.json()
            except Exception:
                last_err = 'Respuesta no JSON del servidor'
                continue
            if int(data.get('ret', 0)) != 200:
                msg = str(data.get('msg') or 'Error')
                if idx == 0 and ('600' in str(data.get('ret')) or 'token' in msg.lower()):
                    token2, err2 = infin8linx_get_token(config, force=True)
                    if token2 and not err2:
                        resp2 = _post(scheme)
                        try:
                            data = resp2.json()
                        except Exception:
                            data = {}
                        if int(data.get('ret', 0)) == 200:
                            result = data.get('data', {}).get('result') or {}
                            down = result.get('downurl') or ''
                            if down:
                                return down, ''
                return '', msg[:300]
            result = data.get('data', {}).get('result') or {}
            down = result.get('downurl') or ''
            if not down:
                return '', 'El proveedor no devolvio URL de grabacion'
            return down, ''
        except Exception as e:
            msg = str(e)
            if 'wrong version number' in msg.lower() or 'SSLError' in type(e).__name__:
                last_err = 'El puerto habla HTTP plano, no HTTPS (%s)' % scheme
            elif 'timed out' in msg.lower():
                last_err = 'Tiempo de espera agotado (%s)' % scheme
            else:
                last_err = 'Error de conexion (%s): %s' % (scheme, msg[:160])
            continue
    return '', last_err or 'No se pudo obtener la grabacion'


def voice_place_call(phone, script, config=None, contact_name='', forced_ext='', customuuid=''):
    """Place an outbound voice call that plays TTS script.

    Returns dict: {ok, call_sid, status, error_msg, price, duration, extnumber, customuuid}
    Provider 'simulation' does not hit any external API.
    forced_ext: si se indica (extension fija del agente/usuario que llama),
    infin8linx la utilizara en lugar de elegir una al azar del pool.
    customuuid: identificador propio que se envia a Infinity y regresa en el
    CDR/callback, permitiendo asociar la llamada con el registro local.
    """
    if config is None:
        config = get_voice_config()
    provider = (config or {}).get('provider', 'simulation') or 'simulation'
    normalized = normalize_phone(phone)
    forced_ext = (forced_ext or '').strip()
    if not customuuid:
        customuuid = 'VC' + uuid.uuid4().hex[:24]
    result = {
        'ok': False, 'call_sid': '', 'status': 'failed',
        'error_msg': '', 'price': 0.0, 'duration': 0, 'extnumber': forced_ext,
        'customuuid': customuuid,
    }

    if provider in ('', 'simulation', 'simulacion', 'none'):
        # Simulation: pseudo-random deterministic outcome so demos are reproducible
        digest = hashlib.sha256((normalized + script).encode()).hexdigest()
        bucket = int(digest[:6], 16) % 100
        if bucket < 10:
            result.update(status='no-answer', error_msg='Sin respuesta (simulacion)')
        elif bucket < 15:
            result.update(status='busy', error_msg='Ocupado (simulacion)')
        elif bucket < 20:
            result.update(status='failed', error_msg='Error de red (simulacion)')
        else:
            dur = 15 + (int(digest[6:10], 16) % 90)
            result.update(ok=True, call_sid='SIM' + digest[:14].upper(),
                          status='completed', duration=dur)
        return result

    if provider == 'twilio':
        try:
            from twilio.rest import Client
            from twilio.base.exceptions import TwilioRestException
            client = Client(config['account_sid'], config['auth_token'])
            twiml = _twilio_tts_xml(script)
            call = client.calls.create(
                to=normalized,
                from_=config['from_number'],
                twiml=twiml,
                status_callback_method='GET',
            )
            sid = getattr(call, 'sid', '') or ''
            status = (getattr(call, 'status', '') or 'initiated').lower()
            mapped = {
                'queued': 'initiated', 'initiated': 'initiated', 'ringing': 'ringing',
                'in-progress': 'answered', 'answered': 'answered',
                'completed': 'completed', 'busy': 'busy', 'failed': 'failed',
                'no-answer': 'no-answer', 'canceled': 'canceled',
            }.get(status, 'initiated')
            result.update(ok=mapped in ('initiated', 'ringing', 'answered', 'completed'),
                          call_sid=sid, status=mapped)
            if not result['ok']:
                result['error_msg'] = 'Estado Twilio: ' + status
            return result
        except ImportError:
            result['error_msg'] = 'Paquete twilio no instalado en el servidor'
            return result
        except Exception as e:
            # TwilioRestException has .msg / .code
            msg = getattr(e, 'msg', None) or str(e)
            result['error_msg'] = msg[:500]
            return result

    if provider == 'custom':
        # Generic JSON voice gateway. POST {to, from, text, callback}.
        try:
            payload = {
                'to': normalized,
                'from': config['from_number'],
                'text': script,
                'account_sid': config['account_sid'],
            }
            headers = {
                'Authorization': 'Bearer ' + config['auth_token'],
                'Content-Type': 'application/json',
            }
            url = config['api_domain'].rstrip('/') + '/call'
            resp = http_requests.post(url, json=payload, headers=headers, timeout=20)
            data = resp.json() if resp.content else {}
            if resp.status_code < 300 and str(data.get('code', 0)) in ('0', '200'):
                result.update(ok=True, call_sid=str(data.get('call_id') or data.get('sid') or ''),
                              status='initiated')
            else:
                result['error_msg'] = str(data.get('message') or data.get('msg') or ('HTTP ' + str(resp.status_code)))[:500]
            return result
        except Exception as e:
            result['error_msg'] = str(e)[:500]
            return result

    if provider == 'infin8linx':
        # infin8linx click-to-call: rings a SIP extension first, then connects
        # to destnumber. There is no TTS broadcast in this API; the agent on the
        # extension reads the script. If the calling user has a fixed extension
        # assigned (forced_ext), use it; otherwise pick randomly from the pool.
        ok, ref, err, used_ext = infin8linx_make_call(
            normalized, config, forced_ext=forced_ext, customuuid=customuuid)
        if ok:
            result.update(ok=True, call_sid=ref, status='initiated',
                          extnumber=used_ext)
        else:
            result['error_msg'] = err or 'Error al iniciar la llamada'
            result['extnumber'] = used_ext
        return result

    result['error_msg'] = 'Proveedor de voz no soportado: ' + provider
    return result


def voice_query_status(call_sid, config=None):
    """Query a call's live status from the provider. Returns dict or None."""
    if not call_sid or call_sid.startswith('SIM'):
        return None
    if config is None:
        config = get_voice_config()
    provider = (config or {}).get('provider', '')
    if provider == 'twilio':
        try:
            from twilio.rest import Client
            client = Client(config['account_sid'], config['auth_token'])
            call = client.calls(call_sid).fetch()
            status = (getattr(call, 'status', '') or '').lower()
            mapped = {
                'queued': 'initiated', 'initiated': 'initiated', 'ringing': 'ringing',
                'in-progress': 'answered', 'answered': 'answered',
                'completed': 'completed', 'busy': 'busy', 'failed': 'failed',
                'no-answer': 'no-answer', 'canceled': 'canceled',
            }.get(status, status)
            duration = 0
            try:
                duration = int(getattr(call, 'duration', 0) or 0)
            except Exception:
                duration = 0
            price = 0.0
            try:
                price = float(getattr(call, 'price', 0) or 0)
            except Exception:
                price = 0.0
            return {'status': mapped, 'duration': duration, 'price': abs(price)}
        except Exception:
            return None
    return None


# ----- Voice API routes -----

@app.route('/api/voice/call', methods=['POST'])
@login_required
def voice_place_call_route():
    """Place one or more outbound voice calls with a TTS script."""
    data = request.get_json() or {}
    phones = data.get('phones') or []
    script = (data.get('script') or '').strip()
    contact_names = data.get('contact_names', {}) or {}
    if isinstance(phones, str):
        phones = [phones]
    phones = [str(p).strip() for p in phones if str(p).strip()]
    if not phones:
        return jsonify({'error': 'Telefono(s) requerido(s)'}), 400
    if not script:
        script = '(sin guion)'
    if len(phones) > 200:
        return jsonify({'error': 'Maximo 200 numeros por llamada masiva'}), 400

    # Resolve the voice config for the calling agent's country. Each country
    # (mx/co/pe) has its own Infinity credentials/extension pool; agents
    # without a country use the global (legacy) config.
    caller_country = normalize_country(g.user.get('country')) if hasattr(g, 'user') else ''
    cfg = resolve_voice_config(caller_country)
    simulated = not is_voice_configured(caller_country)
    provider = (cfg or {}).get('provider', 'simulation')
    # Extension fija del agente que realiza la llamada.
    caller_ext = (g.user.get('extnumber') or '').strip() if hasattr(g, 'user') else ''
    # Con infin8linx (click-to-call real) cada agente DEBE tener su propia
    # extension fija asignada. Quien no la tiene no puede llamar.
    if not simulated and provider == 'infin8linx' and not caller_ext:
        return jsonify({'error': 'No tienes una extension/telefono fijo asignado. Contacta a un administrador para asignar tu extension antes de realizar llamadas.'}), 403
    db = get_db()
    contact_cache = build_contact_template_cache(db, phones)
    results = []
    errors = []
    for raw in phones:
        phone = normalize_phone(raw)
        name = contact_names.get(raw, '') or contact_names.get(phone, '')
        text = apply_template_vars(script, raw, contact_names, contact_cache)
        res = voice_place_call(phone, text, config=cfg, contact_name=name, forced_ext=caller_ext)
        status = res['status']
        db.execute(
            "INSERT INTO voice_records (phone, contact_name, script, status, call_sid, provider, extnumber, country, customuuid, duration, price, error_msg, initiated_at, finished_at, created_by) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), CASE WHEN ? IN ('completed','failed','no-answer','busy','canceled') THEN datetime('now') ELSE NULL END, ?)",
            (phone, name, text, status, res['call_sid'],
             (cfg or {}).get('provider', 'simulation'),
             res.get('extnumber', ''),
             caller_country or ((cfg or {}).get('country') or ''),
             res.get('customuuid', ''),
             res['duration'], res['price'], res['error_msg'],
             status, g.user['id'])
        )
        results.append({'phone': phone, 'status': status, 'call_sid': res['call_sid'],
                        'extnumber': res.get('extnumber', ''),
                        'customuuid': res.get('customuuid', ''),
                        'error': res['error_msg'], 'simulated': simulated})
        if status in ('failed', 'no-answer', 'busy'):
            errors.append(phone + ': ' + (res['error_msg'] or VOICE_STATUS_LABELS.get(status, status)))
    db.commit()
    db.execute(
        "INSERT INTO send_logs (action, details, status) VALUES (?, ?, ?)",
        ('voice_call',
         json.dumps({'count': len(results), 'user': g.user['username'],
                     'simulated': simulated, 'errors': errors[:5]}),
         'success' if not errors else ('partial' if results else 'error'))
    )
    db.commit()
    ok = sum(1 for r in results if r['status'] in ('completed', 'initiated', 'ringing', 'answered'))
    msg = f'{ok} llamada(s) iniciada(s)'
    if errors:
        msg += f', {len(errors)} fallida(s)'
    if simulated:
        msg += ' (modo simulacion - API de voz no configurada)'
    return jsonify({'message': msg, 'results': results, 'errors': errors[:10], 'simulated': simulated})


@app.route('/api/voice/records', methods=['GET'])
@login_required
def voice_list_records():
    db = get_db()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', '').strip()
    search = request.args.get('search', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    offset = (page - 1) * per_page

    where = ['1=1']
    params = []
    if g.user['role'] == 'team_member':
        where.append('r.created_by = ?')
        params.append(g.user['id'])
    elif g.user['role'] == 'team_admin':
        where.append('r.created_by IN (SELECT id FROM users WHERE id=? OR team_creator_id=?)')
        params.extend([g.user['id'], g.user['id']])
    if status:
        where.append('r.status = ?')
        params.append(status)
    if search:
        where.append('(r.phone LIKE ? OR r.contact_name LIKE ? OR r.script LIKE ?)')
        like = f'%{search}%'
        params.extend([like, like, like])
    if date_from:
        where.append("date(r.created_at) >= date(?)")
        params.append(date_from)
    if date_to:
        where.append("date(r.created_at) <= date(?)")
        params.append(date_to)
    where_sql = ' AND '.join(where)

    total = db.execute(f"SELECT COUNT(*) AS c FROM voice_records r WHERE {where_sql}", params).fetchone()['c']
    rows = db.execute(
        f"SELECT r.*, u.username AS sender_username, u.full_name AS sender_full_name "
        f"FROM voice_records r LEFT JOIN users u ON u.id=r.created_by "
        f"WHERE {where_sql} ORDER BY r.id DESC LIMIT ? OFFSET ?",
        params + [per_page, offset]
    ).fetchall()
    items = []
    for r in rows:
        items.append({
            'id': r['id'], 'phone': r['phone'], 'contact_name': r['contact_name'] or '',
            'script': r['script'], 'status': r['status'],
            'status_label': VOICE_STATUS_LABELS.get(r['status'], r['status']),
            'call_sid': r['call_sid'] or '', 'provider': r['provider'] or '',
            'ext_used': r['extnumber'] or '',
            'duration': r['duration'] or 0, 'price': float(r['price'] or 0),
            'error_msg': r['error_msg'] or '',
            'record_file': (r['record_file'] or '') if 'record_file' in r.keys() else '',
            'provider_uuid': (r['provider_uuid'] or '') if 'provider_uuid' in r.keys() else '',
            'hangupcause': (r['hangupcause'] or 0) if 'hangupcause' in r.keys() else 0,
            'answer_at': r['answer_at'] if 'answer_at' in r.keys() else None,
            'initiated_at': r['initiated_at'], 'finished_at': r['finished_at'],
            'created_at': r['created_at'],
            'created_by': r['created_by'],
            'sender_username': r['sender_username'] or '',
            'sender_full_name': r['sender_full_name'] or '',
        })
    return jsonify({'records': items, 'total': total, 'page': page, 'per_page': per_page})


@app.route('/api/voice/statistics', methods=['GET'])
@login_required
def voice_statistics():
    db = get_db()
    scope = ''
    params = []
    if g.user['role'] == 'team_member':
        scope = ' AND created_by = ?'
        params = [g.user['id']]
    elif g.user['role'] == 'team_admin':
        scope = ' AND created_by IN (SELECT id FROM users WHERE id=? OR team_creator_id=?)'
        params = [g.user['id'], g.user['id']]
    today = datetime.now().strftime('%Y-%m-%d')
    today_calls = db.execute(
        f"SELECT COUNT(*) AS c FROM voice_records WHERE date(initiated_at)=? {scope}", [today] + params
    ).fetchone()['c']
    total = db.execute(f"SELECT COUNT(*) AS c FROM voice_records WHERE 1=1 {scope}", params).fetchone()['c']
    completed = db.execute(
        f"SELECT COUNT(*) AS c FROM voice_records WHERE status='completed' {scope}", params
    ).fetchone()['c']
    failed = db.execute(
        f"SELECT COUNT(*) AS c FROM voice_records WHERE status IN ('failed','no-answer','busy','canceled') {scope}",
        params
    ).fetchone()['c']
    pending = db.execute(
        f"SELECT COUNT(*) AS c FROM voice_records WHERE status IN ('pending','initiated','ringing','answered') {scope}",
        params
    ).fetchone()['c']
    total_duration = db.execute(
        f"SELECT COALESCE(SUM(duration),0) AS s FROM voice_records WHERE status='completed' {scope}", params
    ).fetchone()['s'] or 0
    last7 = []
    for i in range(6, -1, -1):
        day = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        c = db.execute(
            f"SELECT COUNT(*) AS c FROM voice_records WHERE date(initiated_at)=? {scope}", [day] + params
        ).fetchone()['c']
        last7.append({'date': day, 'count': c})
    answer_rate = (completed / total * 100) if total else 0
    caller_country = normalize_country(g.user.get('country')) if hasattr(g, 'user') else ''
    vcfg = resolve_voice_config(caller_country)
    provider = (vcfg or {}).get('provider', 'simulation')
    caller_ext = (g.user.get('extnumber') or '').strip() if hasattr(g, 'user') else ''
    configured = is_voice_configured(caller_country)
    # Con infin8linx, el agente debe tener extension fija propia para llamar.
    can_call = True if not configured or provider != 'infin8linx' else bool(caller_ext)
    return jsonify({
        'today_calls': today_calls, 'total': total, 'completed': completed,
        'failed': failed, 'pending': pending,
        'total_duration': int(total_duration),
        'answer_rate': round(answer_rate, 1),
        'last_7_days': last7,
        'configured': configured,
        'provider': provider,
        'caller_country': caller_country,
        'caller_ext': caller_ext,
        'can_call': can_call,
    })


@app.route('/api/voice/hangup', methods=['POST'])
@login_required
def voice_hangup_route():
    """Hang up an active Infinity call on the record's extension.

    Infinity exposes App.Sip_Call.HangupCall (per the voice API docs);
    the dial button uses MakeCall and this route completes the lifecycle
    by releasing the agent's extension/line.
    """
    data = request.get_json() or {}
    record_id = data.get('id')
    call_sid = (data.get('call_sid') or '').strip()
    if not record_id and not call_sid:
        return jsonify({'error': 'id o call_sid requerido'}), 400

    db = get_db()
    user = g.user
    role = user['role']
    uid = user['id']
    params = []
    where = ''
    if record_id:
        where = 'id=?'
        params.append(record_id)
    else:
        where = 'call_sid=?'
        params.append(call_sid)
    if role == 'team_member':
        where += ' AND created_by=?'
        params.append(uid)
    elif role == 'team_admin':
        where += ' AND (created_by=? OR created_by IN (SELECT id FROM users WHERE team_creator_id=?))'
        params.extend([uid, uid])
    row = db.execute(
        "SELECT id, call_sid, extnumber, phone, status, country FROM voice_records WHERE " + where,
        tuple(params)
    ).fetchone()
    if not row:
        return jsonify({'error': 'Llamada no encontrada'}), 404

    ext = (row['extnumber'] or '').strip()
    if not ext:
        return jsonify({'error': 'La llamada no tiene extension registrada; no se puede colgar'}), 400

    rec_country = (row['country'] if 'country' in row.keys() else '') or ''
    user_country = (user['country'] if 'country' in user.keys() else '') or ''
    config = resolve_voice_config(rec_country or user_country or '')
    if (config or {}).get('provider') != 'infin8linx':
        return jsonify({'error': 'El proveedor actual no soporta colgado en tiempo real'}), 400

    ok, err = infin8linx_hangup(config, ext)
    if not ok:
        return jsonify({'error': err or 'No se pudo colgar la llamada'}), 502

    db.execute(
        "UPDATE voice_records SET status='canceled', finished_at=datetime('now') WHERE id=?",
        (row['id'],)
    )
    db.commit()
    try:
        db.execute(
            "INSERT INTO send_logs (action, details, status) VALUES (?, ?, ?)",
            ('voice_hangup', 'Colgo llamada %s ext %s' % (row['call_sid'] or row['id'], ext), 'success')
        )
        db.commit()
    except Exception:
        pass
    return jsonify({'success': True, 'status': 'canceled', 'message': 'Llamada colgada'})


def _map_infinity_cdr_status(cdr):
    """Map an Infinity CDR payload to a local voice_records status.

    billsec > 0  -> answered (the agent/customer spoke)
    answertime present but billsec 0 -> ringing then dropped
    otherwise use hangupcause heuristics.
    """
    try:
        billsec = int(cdr.get('billsec') or 0)
    except (TypeError, ValueError):
        billsec = 0
    answer = str(cdr.get('answertime') or '').strip()
    if billsec > 0:
        return 'answered'
    if answer:
        return 'busy'
    # No answer / failed / canceled. hangupcause is provider-specific; we keep
    # it generic and store the code for reference.
    return 'no-answer'


def _parse_cdr_datetime(value):
    """Return a DB-friendly datetime string from 'YYYY-MM-DD HH:MM:SS', or None."""
    v = str(value or '').strip()
    if not v:
        return None
    try:
        return datetime.strptime(v[:19], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        return None


@app.route('/api/voice/cdr', methods=['POST'])
def voice_cdr_callback():
    """CDR/callback endpoint called by Infinity when a call ends.

    Public (no login): Infinity servers push JSON here. We correlate via
    ``customuuid`` (sent on MakeCall) and fall back to destnumber + ext +
    recent initiated record. Updates status, durations and recording file.
    """
    cdr = request.get_json(silent=True) or request.form.to_dict() or {}
    app.logger.info('[voice-cdr] received: %s', json.dumps(cdr, ensure_ascii=False)[:800])
    try:
        customuuid = str(cdr.get('customuuid') or '').strip()
        dest = ''.join(ch for ch in str(cdr.get('destnumber') or '') if ch.isdigit())
        ext = str(cdr.get('extnumber') or '').strip()
        provider_uuid = str(cdr.get('uuid') or '').strip()
        record_file = str(cdr.get('recordfilename') or '').strip()
        try:
            hangupcause = int(cdr.get('hangupcause') or 0)
        except (TypeError, ValueError):
            hangupcause = 0
        try:
            billsec = int(cdr.get('billsec') or 0)
        except (TypeError, ValueError):
            billsec = 0
        try:
            duration = int(cdr.get('duration') or 0)
        except (TypeError, ValueError):
            duration = 0
        answer_at = _parse_cdr_datetime(cdr.get('answertime'))
        endtime = _parse_cdr_datetime(cdr.get('endtime'))
        status = _map_infinity_cdr_status(cdr)
        # A hangup before the call is answered by either side is a no-answer/
        # cancel, not a completed talked call.
        if status == 'answered':
            final_status = 'completed'
        else:
            final_status = status

        db = get_db()
        row = None
        if customuuid:
            row = db.execute(
                "SELECT id FROM voice_records WHERE customuuid=? ORDER BY id DESC LIMIT 1",
                (customuuid,)
            ).fetchone()
        if row is None and dest and ext:
            row = db.execute(
                "SELECT id FROM voice_records WHERE phone LIKE ? AND extnumber=? "
                "AND status IN ('pending','initiated','ringing') ORDER BY id DESC LIMIT 1",
                ('%' + dest[-10:], ext)
            ).fetchone()
        if row is None and provider_uuid:
            row = db.execute(
                "SELECT id FROM voice_records WHERE provider_uuid=? ORDER BY id DESC LIMIT 1",
                (provider_uuid,)
            ).fetchone()

        if row is None:
            app.logger.warning('[voice-cdr] no matching record for %s', customuuid or dest)
            return jsonify({'ok': True, 'matched': False}), 200

        rid = row['id']
        db.execute(
            "UPDATE voice_records SET status=?, duration=?, provider_uuid=COALESCE(NULLIF(?, ''), provider_uuid), "
            "record_file=COALESCE(NULLIF(?, ''), record_file), hangupcause=?, "
            "answer_at=COALESCE(?, answer_at), "
            "finished_at=COALESCE(?, datetime('now')) WHERE id=?",
            (final_status, billsec or duration, provider_uuid, record_file,
             hangupcause, answer_at, endtime, rid)
        )
        db.commit()
        try:
            db.execute(
                "INSERT INTO send_logs (action, details, status) VALUES (?, ?, ?)",
                ('voice_cdr',
                 json.dumps({'id': rid, 'status': final_status, 'billsec': billsec,
                             'cause': hangupcause, 'uuid': provider_uuid}, ensure_ascii=False),
                 'success')
            )
            db.commit()
        except Exception:
            pass
        return jsonify({'ok': True, 'matched': True, 'id': rid, 'status': final_status}), 200
    except Exception as exc:
        app.logger.exception('[voice-cdr] processing failed')
        # Always 200 so Infinity does not retry-storm; log the error instead.
        return jsonify({'ok': False, 'error': str(exc)[:200]}), 200


@app.route('/api/voice/recording', methods=['GET'])
@login_required
def voice_recording_url():
    """Return a (short-lived) recording download URL for a call.

    Uses Infinity App.Sip_Cdr.GetRecodeFile. Requires record_file from CDR.
    """
    rid = request.args.get('id', type=int)
    if not rid:
        return jsonify({'error': 'id requerido'}), 400
    db = get_db()
    row = db.execute(
        "SELECT id, record_file, country, created_by FROM voice_records WHERE id=?", (rid,)
    ).fetchone()
    if not row:
        return jsonify({'error': 'Llamada no encontrada'}), 404
    record_file = (row['record_file'] if 'record_file' in row.keys() else '') or ''
    if not record_file:
        return jsonify({'error': 'Esta llamada no tiene grabacion disponible'}), 404
    rec_country = (row['country'] if 'country' in row.keys() else '') or ''
    user_country = (g.user['country'] if 'country' in g.user.keys() else '') or ''
    config = resolve_voice_config(rec_country or user_country or '')
    if (config or {}).get('provider') != 'infin8linx':
        return jsonify({'error': 'El proveedor no soporta grabaciones'}), 400
    url, err = infin8linx_get_record_url(config, record_file)
    if err:
        return jsonify({'error': err}), 502
    return jsonify({'url': url, 'record_file': record_file})


@app.route('/api/voice/query-status', methods=['POST'])
@login_required
def voice_query_status_route():
    data = request.get_json() or {}
    call_sid = (data.get('call_sid') or '').strip()
    record_id = data.get('id')
    if not call_sid and record_id:
        db = get_db()
        row = db.execute("SELECT call_sid FROM voice_records WHERE id=?", (record_id,)).fetchone()
        if row:
            call_sid = row['call_sid'] or ''
    if not call_sid:
        return jsonify({'error': 'call_sid requerido'}), 400
    if not call_sid.startswith('INF'):
        return jsonify({
            'success': True,
            'status': 'initiated',
            'live_status_supported': False,
            'message': 'No se pudo consultar el estado en tiempo real; proveedor no soportado o llamada simulada.'
        }), 200

    info = voice_query_status(call_sid)
    if not info:
        return jsonify({
            'success': True,
            'status': 'initiated',
            'live_status_supported': False,
            'message': 'No se pudo consultar el estado en tiempo real con Infinity.'
        }), 200
    db = get_db()
    db.execute(
        "UPDATE voice_records SET status=?, duration=?, price=?, "
        "finished_at=CASE WHEN ? IN ('completed','failed','no-answer','busy','canceled') THEN datetime('now') ELSE finished_at END "
        "WHERE call_sid=?",
        (info['status'], info['duration'], info['price'], info['status'], call_sid)
    )
    db.commit()
    return jsonify({'success': True, 'status': info['status'], 'duration': info['duration'], 'price': info['price']})


@app.route('/api/config/voice', methods=['GET'])
@admin_required
def voice_get_config():
    # List all per-country voice configs (like sms_api_configs).
    return jsonify({'configs': list_voice_configs()})


@app.route('/api/config/voice', methods=['POST'])
@admin_required
def voice_create_config():
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    country = normalize_country(data.get('country'))
    if not name:
        return jsonify({'error': 'Nombre es requerido'}), 400
    if not country:
        return jsonify({'error': 'Pais invalido (use mx/co/pe)'}), 400
    db = get_db()
    exists = db.execute("SELECT id FROM voice_configs WHERE country = ?", (country,)).fetchone()
    if exists:
        return jsonify({'error': 'Ya existe una configuracion para este pais'}), 409
    cursor = db.execute(
        "INSERT INTO voice_configs (name, country, provider, is_active) VALUES (?, ?, 'infin8linx', 1)",
        (name, country)
    )
    db.commit()
    return jsonify({'message': 'Configuracion creada', 'id': cursor.lastrowid}), 201


@app.route('/api/config/voice/<int:config_id>', methods=['PUT'])
@admin_required
def voice_update_config(config_id):
    data = request.get_json() or {}
    db = get_db()
    row = db.execute("SELECT * FROM voice_configs WHERE id = ?", (config_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Configuracion no encontrada'}), 404
    row = dict(row)
    name = (data.get('name') or row.get('name') or '').strip()
    country = normalize_country(data.get('country', row.get('country')))
    provider = (data.get('provider') or row.get('provider') or 'infin8linx').strip().lower()
    if provider not in ('simulation', 'infin8linx'):
        return jsonify({'error': 'Proveedor invalido'}), 400
    api_domain = (data.get('api_domain') if 'api_domain' in data else row.get('api_domain')) or ''
    # Normalize: store host[:port] without scheme so the UI stays clean;
    # _voice_api_url() prepends https:// on every request.
    api_domain = api_domain.strip()
    if api_domain.startswith(('http://', 'https://')):
        api_domain = re.sub(r'^https?://', '', api_domain)
    api_domain = api_domain.rstrip('/')
    voice_appid = (data.get('voice_appid') if 'voice_appid' in data else row.get('voice_appid')) or ''
    from_number = (data.get('from_number') if 'from_number' in data else row.get('from_number')) or ''
    new_ak = data.get('voice_accesskey')
    if new_ak is not None and new_ak.strip() == '':
        voice_accesskey = row.get('voice_accesskey') or ''  # write-only: keep
    elif new_ak is not None:
        voice_accesskey = new_ak.strip()
    else:
        voice_accesskey = row.get('voice_accesskey') or ''
    is_active = 1 if data.get('is_active', True) else 0
    # Only Infinity is exposed; when all three credentials are present treat the
    # row as a live Infinity config, otherwise fall back to simulation (like the
    # SMS "no credentials -> simulation" behavior).
    provider = 'infin8linx' if (api_domain.strip() and voice_appid.strip() and voice_accesskey.strip()) else 'simulation'

    token_reset = (
        ('voice_appid' in data and (voice_appid or '') != (row.get('voice_appid') or ''))
        or (new_ak is not None and new_ak.strip() != '')
    )
    sets = ("name=?, country=?, provider=?, api_domain=?, voice_appid=?, voice_accesskey=?,"
            " from_number=?, is_active=?, updated_at=datetime('now')")
    params = [name, country, provider, api_domain.strip(), voice_appid.strip(),
              voice_accesskey, from_number.strip(), is_active]
    if token_reset:
        sets += ", voice_token='', voice_token_expiry=0"
        _INFIN_TOKEN_CACHE.pop('vcfg:%s' % config_id, None)
    params.append(config_id)
    db.execute("UPDATE voice_configs SET " + sets + " WHERE id=?", params)
    db.commit()
    return jsonify({'message': 'Configuracion actualizada'})


@app.route('/api/config/voice/<int:config_id>', methods=['DELETE'])
@admin_required
def voice_delete_config(config_id):
    db = get_db()
    db.execute("DELETE FROM voice_configs WHERE id=?", (config_id,))
    db.commit()
    _INFIN_TOKEN_CACHE.pop('vcfg:%s' % config_id, None)
    return jsonify({'message': 'Configuracion eliminada'})


@app.route('/api/config/voice/test', methods=['POST'])
@admin_required
def voice_test_config():
    """Validate Infinity credentials for a given config id or country."""
    data = request.get_json(silent=True) or {}
    config_id = data.get('config_id')
    db = get_db()
    if config_id:
        row = db.execute("SELECT * FROM voice_configs WHERE id = ?", (config_id,)).fetchone()
        if not row:
            return jsonify({'error': 'Configuracion no encontrada'}), 404
        cfg = dict(row)
        cfg['provider'] = (cfg.get('provider') or '').lower()
    else:
        country = normalize_country(data.get('country') or request.args.get('country'))
        cfg = resolve_voice_config(country)
    if not cfg or not is_voice_configured(cfg.get('country')):
        return jsonify({'error': 'API de voz no configurada para este pais'}), 400
    if cfg['provider'] == 'infin8linx':
        raw_domain = (cfg.get('api_domain') or '').strip()
        appid = (cfg.get('voice_appid') or '').strip()
        accesskey = (cfg.get('voice_accesskey') or '').strip()
        if not (raw_domain and appid and accesskey):
            return jsonify({'error': 'Faltan URL, AppID o AccessKey para probar'}), 400
        # Probe connectivity over https then http (whichever answers first).
        scheme, resp, conn_err = _voice_probe_host(cfg, appid, accesskey, timeout=10)
        if not scheme:
            return jsonify({
                'error': (conn_err or 'No se pudo conectar') +
                         '. Verifique que el servidor tenga salida al host/puerto '
                         '(firewall/security group) y que la URL sea correcta.'
            }), 400
        # Parse the answer from the reachable endpoint.
        try:
            data = resp.json() if resp is not None else {}
        except Exception:
            return jsonify({'error': 'Respuesta no JSON en %s://%s (revisar URL/puerto)'
                                     % (scheme, re.sub(r'^https?://', '', raw_domain))}), 400
        if int(data.get('ret', 0)) != 200:
            return jsonify({'error': str(data.get('msg') or ('HTTP %s - revisar credenciales/URL' % (resp.status_code if resp else '?')))[:300]}), 400
        body = data.get('data') or {}
        if int(body.get('status', 1)) != 0:
            errs = body.get('errors') or {}
            return jsonify({'error': str(body.get('desc') or errs.get('codemsg') or 'Credenciales rechazadas')[:300]}), 400
        token = ((body.get('result') or {}).get('token') or '').strip()
        if not token:
            return jsonify({'error': 'El servidor no devolvio token'}), 400
        # Persist the working scheme + token so real calls use the right protocol.
        try:
            db.execute(
                "UPDATE voice_configs SET voice_scheme = ?, voice_token = ?, "
                "voice_token_expiry = ? WHERE id = ?",
                (scheme, token, int(time.time()) + 12 * 3600, cfg['id']))
            db.commit()
        except Exception:
            pass
        _INFIN_TOKEN_CACHE[_token_cache_key(cfg)] = {
            'token': token, 'expiry': int(time.time()) + 12 * 3600}
        label = COUNTRY_LABELS.get(cfg.get('country'), 'General')
        return jsonify({'success': True,
                        'message': 'Conexion OK con Infinity (%s) via %s' % (label, scheme.upper()),
                        'scheme': scheme,
                        'token_preview': (token[:6] + '...' + token[-4:]) if len(token) > 12 else 'OK'})
    return jsonify({'error': 'Proveedor no soportado'}), 400


# ---------------------------------------------------------------------------
# Extensions management (separate page)
# ---------------------------------------------------------------------------

@app.route('/api/extensions', methods=['GET'])
@admin_required
def list_extensions():
    """List extensions, optionally filtered by country.

    Returns each extension with its status (free/assigned) and the agent it is
    assigned to. Query param: ?country=mx|co|pe (empty = general pool).
    """
    _extensions_seed_from_config()
    country = normalize_country(request.args.get('country'))
    db = get_db()
    rows = db.execute(
        "SELECT e.id, e.extnumber, e.country, e.assigned_to, "
        "u.username AS agent_username, u.full_name AS agent_full_name "
        "FROM extensions e "
        "LEFT JOIN users u ON u.id = e.assigned_to "
        "WHERE e.country = ? "
        "ORDER BY e.extnumber",
        (country,)
    ).fetchall()
    items = []
    for r in rows:
        d = dict(r) if not isinstance(r, tuple) else {
            'id': r[0], 'extnumber': r[1], 'country': r[2], 'assigned_to': r[3],
            'agent_username': r[4], 'agent_full_name': r[5]
        }
        d['status'] = 'assigned' if d.get('assigned_to') else 'free'
        items.append(d)
    total = len(items)
    assigned = sum(1 for i in items if i['status'] == 'assigned')
    return jsonify({
        'country': country,
        'country_label': COUNTRY_LABELS.get(country, 'General'),
        'extensions': items,
        'total': total,
        'free': total - assigned,
        'assigned': assigned,
    })


@app.route('/api/extensions/template', methods=['GET'])
@admin_required
def download_extensions_template():
    """Download an Excel template (.xlsx) listing extensions to upload.

    The file has a single column 'Extension' with a few example rows. The
    uploaded country is taken from the active tab, so the template itself
    does not carry a country.
    """
    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Extensiones'
    ws.append(['Extension'])
    for sample in ('8001', '8002', '8003'):
        ws.append([sample])
    ws.column_dimensions['A'].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='plantilla_extensiones.xlsx'
    )


def _parse_extensions_from_upload(file_storage):
    """Parse an uploaded file (.xlsx/.xls/.csv/.txt) into a list of raw extension
    strings. Raises ValueError with a user-facing Spanish message on failure.
    """
    filename = (file_storage.filename or '').lower()
    if filename.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError('El servidor no tiene openpyxl instalado')
        try:
            wb = load_workbook(file_storage, read_only=True, data_only=True)
        except Exception as e:
            raise ValueError('No se pudo leer el archivo Excel: %s' % e)
        values = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    val = str(cell).strip()
                    # Excel may store whole numbers as floats (8001.0)
                    if val.endswith('.0') and val[:-2].isdigit():
                        val = val[:-2]
                    # A single cell may hold several comma/semicolon-separated values
                    for piece in val.replace(';', ',').replace('\t', ',').split(','):
                        values.append(piece)
        return values
    if filename.endswith('.xls'):
        raise ValueError('El formato .xls (Excel antiguo) no es compatible. Guarde el archivo como .xlsx y vuelva a intentarlo.')
    # .csv / .txt / any text
    try:
        raw = file_storage.read()
        text = raw.decode('utf-8-sig', errors='replace') if isinstance(raw, (bytes, bytearray)) else str(raw)
    except Exception as e:
        raise ValueError('No se pudo leer el archivo: %s' % e)
    return text.replace(';', ',').replace('\n', ',').replace('\r', ',').replace('\t', ',').split(',')


def _bulk_insert_extensions(country, candidates):
    """Shared insertion logic for bulk-adding extensions.

    Returns (added, duplicates, invalid). Each candidate is normalized,
    de-duplicated (case-insensitive, globally across countries), and linked
    to an existing active user that already holds the extension.
    """
    parsed = []
    seen = set()
    for c in candidates:
        ext = _normalize_extnumber(c)
        if not ext:
            continue
        key = ext.lower()
        if key in seen:
            continue
        seen.add(key)
        parsed.append(ext)
    _extensions_seed_from_config()
    db = get_db()
    added = []
    duplicates = []
    invalid = []
    for ext in parsed:
        clash = db.execute(
            "SELECT country FROM extensions WHERE LOWER(TRIM(extnumber))=LOWER(?)",
            (ext,)
        ).fetchone()
        if clash:
            duplicates.append({'extnumber': ext, 'country': (clash['country'] if not isinstance(clash, tuple) else clash[0])})
            continue
        owner = _find_user_by_extnumber(ext)
        try:
            db.execute(
                "INSERT INTO extensions (extnumber, country, assigned_to) VALUES (?,?,?)",
                (ext, country, owner['id'] if owner else None)
            )
            added.append({'extnumber': ext, 'auto_linked': bool(owner)})
        except Exception as e:
            invalid.append({'extnumber': ext, 'error': str(e)})
    db.commit()
    return added, duplicates, invalid


@app.route('/api/extensions', methods=['POST'])
@admin_required
def add_extensions():
    """Bulk add extensions for a country.

    Accepts either:
      - multipart/form-data with an uploaded file (file=<...>) plus country;
        supports .xlsx, .csv and .txt (one column/list of extensions).
      - JSON {country, extensions: "8001, 8002"} or {country, extensions:[...]}.
    Duplicates (within the request or already stored) are reported and skipped.
    """
    country = normalize_country(
        request.form.get('country') if request.files else (request.get_json(silent=True) or {}).get('country')
    )
    if request.files and 'file' in request.files:
        up = request.files['file']
        if not up or not up.filename:
            return jsonify({'error': 'No se selecciono ningun archivo'}), 400
        try:
            candidates = _parse_extensions_from_upload(up)
        except ValueError as ve:
            return jsonify({'error': str(ve)}), 400
    else:
        data = request.get_json(silent=True) or {}
        raw = data.get('extensions', '')
        if isinstance(raw, list):
            candidates = [str(x) for x in raw]
        else:
            candidates = str(raw).replace(';', ',').replace('\n', ',').replace('\t', ',').split(',')
    added, duplicates, invalid = _bulk_insert_extensions(country, candidates)
    if not added and not duplicates and not invalid:
        return jsonify({'error': 'Debe proporcionar al menos una extension valida'}), 400
    return jsonify({
        'message': 'Extensiones procesadas',
        'added_count': len(added),
        'duplicate_count': len(duplicates),
        'invalid_count': len(invalid),
        'added': added,
        'duplicates': duplicates,
        'invalid': invalid,
    }), 201 if added else 200


@app.route('/api/extensions/<int:ext_id>', methods=['DELETE'])
@admin_required
def delete_extension(ext_id):
    """Delete a single extension. Only free (unassigned) extensions can be removed."""
    _extensions_seed_from_config()
    db = get_db()
    row = db.execute("SELECT * FROM extensions WHERE id=?", (ext_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Extension no encontrada'}), 404
    assigned_to = row['assigned_to'] if 'assigned_to' in row.keys() else (row[3] if len(row) > 3 else None)
    if assigned_to:
        agent = db.execute("SELECT username, full_name FROM users WHERE id=?", (assigned_to,)).fetchone()
        name = ''
        if agent:
            name = agent['full_name'] if 'full_name' in agent.keys() and agent['full_name'] else (agent['username'] if not isinstance(agent, tuple) else agent[0])
        return jsonify({'error': f'La extension esta asignada a {name}. Liberela primero desde el usuario.'}), 409
    db.execute("DELETE FROM extensions WHERE id=?", (ext_id,))
    db.commit()
    return jsonify({'message': 'Extension eliminada'})


init_db()

# Start daily auto-clear worker after the database is ready. In production,
# Gunicorn imports this module once per worker; the advisory/SQLite lock
# guarantees only one worker performs the actual deletion per day.
_auto_clear_thread = threading.Thread(target=_auto_clear_loop, name='auto-clear-contacts', daemon=True)
_auto_clear_thread.start()
app.logger.info('Daily auto-clear contacts worker thread started')

if __name__ == '__main__':
    port = int(os.environ.get('DEPLOY_RUN_PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
